from typing import List, Optional, Union, Dict, Any
from sqlmodel import Session, select, text
from app.schemas.purchase import (PurchaseOrder,PurchaseOrderQuery,PurchaseWip,PurchaseWipQuery)
from app.schemas.assy import (AssyOrder,
                              AssyOrderQuery,
                              AssyWip,
                              AssyWipQuery,
                              AssyOrderItemsQuery,
                              AssyOrderPackageTypeQuery,
                              AssyOrderSupplierQuery,
                              AssyBomQuery,
                              AssyBom,
                              AssyAnalyzeTotalResponse,
                              AssyAnalyzeLoadingResponse,
                              AssyYearTrendResponse,
                              AssySupplyAnalyzeResponse,
                              ItemWaferInfoResponse,
                              SalesResponse,
                              AssySubmitOrdersRequest,
                              AssySubmitOrdersResponse,
                              CpTestOrdersQuery,
                              CpTestOrders,
                              AssyRequireOrdersQuery,
                              AssyRequireOrdersList,
                              AssyRequireOrdersCancel)
from app.schemas.stock import (StockQuery,
                             Stock,
                             WaferIdQtyDetailQuery,
                             WaferIdQtyDetail,
                             StockSummaryQuery,
                             StockSummary)
from app.schemas.e10 import (FeatureGroupNameQuery,
                             ItemCodeQuery,
                             ItemNameQuery,
                             WarehouseNameQuery,
                             TestingProgramQuery,
                             BurningProgramQuery,
                             LotCodeQuery)
from app.schemas.report import GlobalReport,SopAnalyzeResponse,ChipInfoTraceQuery,ChipInfoTraceResponse,ChipInfoTrace
from app.core.exceptions import CustomException
from app.core.logger import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


class CRUDE10:
    """E10 CRUD操作类"""

    def _clean_input(self, value: str) -> str:
        """清理输入参数，移除潜在的危险字符
        
        Args:
            value: 输入字符串
            
        Returns:
            清理后的字符串
        """
        # 移除常见的 SQL 注入字符
        dangerous_chars = ["'", '"', ';', '--', '/*', '*/', 'xp_']
        cleaned = value
        for char in dangerous_chars:
            cleaned = cleaned.replace(char, '')
        return cleaned

    def get_feature_group_name(self,db:Session,params:FeatureGroupNameQuery)->Dict[str,Any]:
        """获取品号群组"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT FEATURE_GROUP_NAME
                FROM FEATURE_GROUP
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.feature_group_name:
                feature_group_name = params.feature_group_name.upper()
                conditions.append("AND FEATURE_GROUP_NAME LIKE :feature_group_name")
                query_params["feature_group_name"] = f"%{self._clean_input(feature_group_name)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY FEATURE_GROUP_NAME"

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()

            # 转换为模型实例
            feature_group_names = [
                {
                    "label": row.FEATURE_GROUP_NAME,
                    "value": row.FEATURE_GROUP_NAME
                } for row in result
            ]
            return {
                "list": feature_group_names
            }
        except Exception as e:
            logger.error(f"获取品号群组失败: {str(e)}")
            raise CustomException("获取品号群组失败")
        
    def get_item_code(self,db:Session,params:ItemCodeQuery)->Dict[str,Any]:
        """获取品号"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT ITEM_CODE
                FROM ITEM
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.item_code:
                item_code = params.item_code.upper()
                conditions.append("AND ITEM_CODE LIKE :item_code")
                query_params["item_code"] = f"%{self._clean_input(item_code)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序  
            query += " ORDER BY ITEM_CODE"

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()
            
            # 转换为模型实例
            item_codes = [
                {
                    "label": row.ITEM_CODE,
                    "value": row.ITEM_CODE
                } for row in result
            ]
            return {
                "list": item_codes
            }
        except Exception as e:
            logger.error(f"获取品号失败: {str(e)}")
            raise CustomException("获取品号失败")
        
    def get_item_name(self,db:Session,params:ItemNameQuery)->Dict[str,Any]:
        """获取品名"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT ITEM_NAME
                FROM ITEM
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.item_name:
                item_name = params.item_name.upper()
                conditions.append("AND ITEM_NAME LIKE :item_name")
                query_params["item_name"] = f"%{self._clean_input(item_name)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY ITEM_NAME"            

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()
            
            # 转换为模型实例
            item_names = [
                {
                    "label": row.ITEM_NAME,
                    "value": row.ITEM_NAME
                } for row in result
            ]
            return {
                "list": item_names
            }
        except Exception as e:
            logger.error(f"获取品名失败: {str(e)}")
            raise CustomException("获取品名失败")
    
    def get_lot_code(self,db:Session,params:LotCodeQuery)->Dict[str,Any]:
        """获取批号"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT LOT_CODE
                FROM ITEM_LOT
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.lot_code:
                lot_code = params.lot_code.upper()
                conditions.append("AND LOT_CODE LIKE :lot_code")
                query_params["lot_code"] = f"%{self._clean_input(lot_code)}%"
            
            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY LOT_CODE"

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()

            # 转换为模型实例
            lot_codes = [
                {
                    "label": row.LOT_CODE,
                    "value": row.LOT_CODE
                } for row in result
            ]
            return {
                "list": lot_codes
            }
        except Exception as e:
            logger.error(f"获取批号失败: {str(e)}")
            raise CustomException("获取批号失败")
    
    def get_warehouse_name(self,db:Session,params:WarehouseNameQuery)->Dict[str,Any]:
        """获取仓库"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT WAREHOUSE_NAME
                FROM WAREHOUSE
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.warehouse_name:
                warehouse_name = params.warehouse_name.upper()
                conditions.append("AND WAREHOUSE_NAME LIKE :warehouse_name")
                query_params["warehouse_name"] = f"%{self._clean_input(warehouse_name)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY WAREHOUSE_NAME" 

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()
            
            # 转换为模型实例
            warehouse_names = [
                {
                    "label": row.WAREHOUSE_NAME,
                    "value": row.WAREHOUSE_NAME
                } for row in result
            ]
            return {
                "list": warehouse_names
            }
        except Exception as e:
            logger.error(f"获取仓库失败: {str(e)}")
            raise CustomException("获取仓库失败")
        
    def get_testing_program(self,db:Session,params:TestingProgramQuery)->Dict[str,Any]:
        """获取测试程序"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT Z_TESTING_PROGRAM_NAME
                FROM Z_TESTING_PROGRAM
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.testing_program:
                testing_program = params.testing_program.upper()
                conditions.append("AND Z_TESTING_PROGRAM_NAME LIKE :testing_program")
                query_params["testing_program"] = f"%{self._clean_input(testing_program)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY Z_TESTING_PROGRAM_NAME"

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()

            # 转换为模型实例
            testing_programs = [
                {
                    "label": row.Z_TESTING_PROGRAM_NAME,
                    "value": row.Z_TESTING_PROGRAM_NAME
                } for row in result
            ]
            return {
                "list": testing_programs
            }
        except Exception as e:
            logger.error(f"获取测试程序失败: {str(e)}")
            raise CustomException("获取测试程序失败")
        
    def get_burning_program(self,db:Session,params:BurningProgramQuery)->Dict[str,Any]:
        """获取烧录程序"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT Z_BURNING_PROGRAM_NAME
                FROM Z_BURNING_PROGRAM
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.burning_program:
                burning_program = params.burning_program.upper()
                conditions.append("AND Z_BURNING_PROGRAM_NAME LIKE :burning_program")
                query_params["burning_program"] = f"%{self._clean_input(burning_program)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY Z_BURNING_PROGRAM_NAME"

            # 执行查询
            result = db.exec(text(query).bindparams(**query_params)).all()

            # 转换为模型实例
            burning_programs = [
                {
                    "label": row.Z_BURNING_PROGRAM_NAME,
                    "value": row.Z_BURNING_PROGRAM_NAME
                } for row in result
            ]
            return {
                "list": burning_programs
            }
        except Exception as e:
            logger.error(f"获取烧录程序失败: {str(e)}")
            raise CustomException("获取烧录程序失败")
        
    def get_purchase_order_by_params(self, db: Session, params: PurchaseOrderQuery) -> List[PurchaseOrder]:
        """根据参数获取采购订单"""
        try:
            # 构建基础查询
            base_query = """
            SELECT
                po.SUPPLIER_FULL_NAME,
                po.DOC_NO,
                CAST(po.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
                p.REMARK,
                i.ITEM_CODE,
                i.ITEM_NAME,
                i.SHORTCUT,
                CAST(p.BUSINESS_QTY AS INT) AS 'BUSINESS_QTY',
                CAST(p.SECOND_QTY AS INT) AS 'SECOND_QTY',
                CAST(s.RECEIPTED_BUSINESS_QTY AS INT) AS 'RECEIPTED_BUSINESS_QTY',
                    CASE 
                        WHEN s.RECEIPT_CLOSE = 0 OR s.RECEIPT_CLOSE IS NULL THEN CAST(p.BUSINESS_QTY - ISNULL(s.RECEIPTED_BUSINESS_QTY, 0) AS INT)
                        ELSE 0
                    END AS WIP_QTY,
                CAST(p.PRICE AS FLOAT(4)) AS PRICE,
                CAST(p.AMOUNT AS FLOAT(4)) AS AMOUNT,
                s.RECEIPT_CLOSE
            FROM PURCHASE_ORDER po
            LEFT JOIN PURCHASE_ORDER_D p
            ON po.PURCHASE_ORDER_ID = p.PURCHASE_ORDER_ID 
            LEFT JOIN ITEM i
            ON p.ITEM_ID = i.ITEM_BUSINESS_ID
            LEFT JOIN PURCHASE_ORDER_SD s
            ON s.PURCHASE_ORDER_D_ID = p.PURCHASE_ORDER_D_ID
            WHERE (p.PURCHASE_TYPE=1) AND (i.ITEM_CODE LIKE N'CL%WF')
            """
            
            # 构建查询条件
            conditions = []
            query_params = {}
            
            # 参数验证和清理
            if params.doc_no and isinstance(params.doc_no, str):
                conditions.append("AND po.DOC_NO LIKE :doc_no")
                query_params["doc_no"] = f"%{self._clean_input(params.doc_no)}%"
                
            if params.item_code and isinstance(params.item_code, str):
                conditions.append("AND i.ITEM_CODE LIKE :item_code")
                query_params["item_code"] = f"%{self._clean_input(params.item_code)}%"
                
            if params.item_name and isinstance(params.item_name, str):
                conditions.append("AND i.ITEM_NAME LIKE :item_name")
                query_params["item_name"] = f"%{self._clean_input(params.item_name)}%"
                
            if params.supplier and isinstance(params.supplier, str):
                conditions.append("AND po.SUPPLIER_FULL_NAME LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"
                
            if params.purchase_date_start:
                conditions.append("AND CAST(po.PURCHASE_DATE AS DATE) >= :purchase_date_start")
                query_params["purchase_date_start"] = params.purchase_date_start
            
            if params.purchase_date_end:
                conditions.append("AND CAST(po.PURCHASE_DATE AS DATE) <= :purchase_date_end")
                query_params["purchase_date_end"] = params.purchase_date_end
            # 添加收货结束状态查询条件
            if params.receipt_close is not None:
                if params.receipt_close == 0:
                    conditions.append("AND (s.RECEIPT_CLOSE = 0 OR s.RECEIPT_CLOSE IS NULL)")
                else:
                    conditions.append("AND s.RECEIPT_CLOSE > 0")

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加排序
            query += " ORDER BY po.PURCHASE_DATE, po.DOC_NO"

            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.exec(stmt).all()
            
            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1)
                FROM PURCHASE_ORDER po
                LEFT JOIN PURCHASE_ORDER_D p ON po.PURCHASE_ORDER_ID = p.PURCHASE_ORDER_ID 
                LEFT JOIN ITEM i ON p.ITEM_ID = i.ITEM_BUSINESS_ID
                LEFT JOIN PURCHASE_ORDER_SD s ON s.PURCHASE_ORDER_D_ID = p.PURCHASE_ORDER_D_ID
                WHERE (p.PURCHASE_TYPE=1) AND (i.ITEM_CODE LIKE N'CL%WF')
                {' '.join(conditions)}
            """
            total = db.exec(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()
            
            # 转换为模型实例
            purchase_orders = [
                PurchaseOrder(
                    SUPPLIER_FULL_NAME=row.SUPPLIER_FULL_NAME,
                    DOC_NO=row.DOC_NO,
                    PURCHASE_DATE=row.PURCHASE_DATE,
                    REMARK=row.REMARK,
                    ITEM_CODE=row.ITEM_CODE,
                    ITEM_NAME=row.ITEM_NAME,
                    SHORTCUT=row.SHORTCUT,
                    BUSINESS_QTY=row.BUSINESS_QTY,
                    SECOND_QTY=row.SECOND_QTY,
                    RECEIPTED_BUSINESS_QTY=row.RECEIPTED_BUSINESS_QTY,
                    WIP_QTY=row.WIP_QTY,
                    PRICE=round(row.PRICE, 4) if row.PRICE else None,
                    AMOUNT=round(row.AMOUNT, 4) if row.AMOUNT else None,
                    RECEIPT_CLOSE=row.RECEIPT_CLOSE
                ) for row in result
            ]
            
            return {
                "list": purchase_orders,
                "total": total or 0
            }
            
        except Exception as e:
            # 记录错误但不暴露详细信息
            logger.error(f"查询采购订单失败: {str(e)}")
            raise CustomException("查询采购订单失败")
        
    def get_purchase_wip_by_params(self,db:Session,params:PurchaseWipQuery)->List[PurchaseWip]:
        """根据参数获取采购在途"""
        try:
            # 构建基础查询
            base_query = """
                SELECT
                    purchaseOrder,
                    itemName,
                    lot,
                    qty,
                    status,
                    stage,
                    layerCount,
                    remainLayer,
                    currentPosition,
                    forecastDate,
                    supplier,
                    finished_at,
                    CASE 
                        WHEN forecastDate IS NULL THEN 0
                        ELSE DATEDIFF(DAY, modified_at, GETDATE())
                    END AS stranded,
                    CASE 
                        WHEN NOT finished_at IS NULL THEN DATEDIFF(DAY, create_at, GETDATE())
                        ELSE NULL
                    END AS leadTime
                FROM huaxinAdmin_wip_fab
                WHERE 1=1
            """
            # 构建查询条件
            conditions = []
            query_params = {}

            # 参数验证和清理
            if params.purchase_order and isinstance(params.purchase_order, str):
                conditions.append("AND purchaseOrder LIKE :purchase_order")
                query_params["purchase_order"] = f"%{self._clean_input(params.purchase_order)}%"

            if params.item_name and isinstance(params.item_name, str):
                conditions.append("AND itemName LIKE :item_name")
                query_params["item_name"] = f"%{self._clean_input(params.item_name)}%"
            
            if params.supplier and isinstance(params.supplier, str):
                conditions.append("AND supplier LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"
            
            if params.status and isinstance(params.status, str):
                conditions.append("AND status LIKE :status")
                query_params["status"] = f"%{self._clean_input(params.status)}%"

            if params.is_finished is not None:
                if params.is_finished == 1:
                    conditions.append("AND finished_at IS NOT NULL")
                else:
                    conditions.append("AND finished_at IS NULL")

            if params.is_stranded is not None:
                if params.is_stranded == 1:
                    conditions.append("AND CASE WHEN forecastDate IS NULL THEN 0 ELSE DATEDIFF(DAY, modified_at, GETDATE()) END > 0")
                else:
                    conditions.append("AND CASE WHEN forecastDate IS NULL THEN 0 ELSE DATEDIFF(DAY, modified_at, GETDATE()) END <= 0")

            if params.days is not None:
                conditions.append("AND forecastDate <= DATEADD(DAY, :days, GETDATE())")
                query_params["days"] = params.days

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY purchaseOrder"

            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.exec(stmt).all()

            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1)
                FROM huaxinAdmin_wip_fab
                WHERE 1=1
                {' '.join(conditions)}
            """
            total = db.exec(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()

            # 转换为模型实例
            purchase_wips = [
                PurchaseWip(
                    purchaseOrder=row.purchaseOrder,
                    itemName=row.itemName,
                    lot=row.lot,
                    qty=row.qty,
                    status=row.status,
                    stage=row.stage,
                    layerCount=row.layerCount,
                    remainLayerCount=row.remainLayer,
                    currentPosition=row.currentPosition,
                    forecastDate=row.forecastDate,
                    supplier=row.supplier,
                    finished_at=row.finished_at,
                    stranded=row.stranded,
                    leadTime=row.leadTime
                ) for row in result
            ]

            return {
                "list": purchase_wips,
                "total": total or 0
            }
            
        except Exception as e:
            logger.error(f"查询采购在途失败: {str(e)}")
            raise CustomException("查询采购在途失败")
    
    def get_purchase_supplier(self,db:Session)->List[str]:
        """获取采购供应商"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT po.SUPPLIER_FULL_NAME
                FROM PURCHASE_ORDER po
                LEFT JOIN PURCHASE_ORDER_D p
                ON po.PURCHASE_ORDER_ID = p.PURCHASE_ORDER_ID 
                LEFT JOIN ITEM i
                ON p.ITEM_ID = i.ITEM_BUSINESS_ID
                LEFT JOIN PURCHASE_ORDER_SD s
                ON s.PURCHASE_ORDER_D_ID = p.PURCHASE_ORDER_D_ID
                WHERE (p.PURCHASE_TYPE=1) AND (i.ITEM_CODE LIKE N'CL%WF')
            """
            result = db.exec(text(base_query)).all()
            # 提取供应商名称列表
            return [row[0] for row in result if row[0]]  # 确保返回非空的供应商名称列表
        except Exception as e:
            logger.error(f"获取采购供应商失败: {str(e)}")
            raise CustomException("获取采购供应商失败")
        
    def get_purchase_wip_supplier(self,db:Session)->List[str]:
        """获取采购在制供应商"""
        try:
            # 构建基础查询
            base_query = """
                SELECT DISTINCT supplier
                FROM huaxinAdmin_wip_fab
                WHERE supplier IS NOT NULL
                ORDER BY supplier
                """
            result = db.exec(text(base_query)).all()
            # 提取供应商名称列表
            return [row[0] for row in result if row[0]]  # 确保返回非空的供应商名称列表
        except Exception as e:
            logger.error(f"获取采购在制供应商失败: {str(e)}")
            raise CustomException("获取采购在制供应商失败")

    def get_assy_order_by_params(self,db:Session,params:AssyOrderQuery)->Dict[str,Any]:
        """获取封装订单列表"""
        try:
            # 构建基础查询
            base_query = """
                SELECT *
                FROM (
                    SELECT
                        hpl.ID,
                        hpl.DOC_NO,
                        hpl.ITEM_CODE,
                        hpl.Z_PACKAGE_TYPE_NAME,
                        hpl.LOT_CODE,
                        hpl.BUSINESS_QTY,
                        hpl.RECEIPTED_PRICE_QTY,
                        0 AS WIP_QTY,
                        hpl.Z_PROCESSING_PURPOSE_NAME,
                        hpl.Z_TESTING_PROGRAM_NAME,
                        hpl.Z_ASSEMBLY_CODE,
                        hpl.Z_WIRE_NAME,
                        hpl.REMARK,
                        hpl.PURCHASE_DATE,
                        ISNULL(hpl.FIRST_ARRIVAL_DATE, DATEADD(MONTH, 2, hpl.PURCHASE_DATE)) AS FIRST_ARRIVAL_DATE,
                        hpl.SUPPLIER_FULL_NAME,
                        hpl.RECEIPT_CLOSE
                    FROM HSUN_PACKAGE_LIST hpl
                    UNION ALL
                    SELECT
                        ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE, PO.DOC_NO) + 115617 AS ID,
                        PO.DOC_NO,
                        ITEM.ITEM_CODE,
                        ITEM.UDF025 AS Z_PACKAGE_TYPE_NAME,
                        ITEM_LOT.LOT_CODE,
                        CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
                        CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
                        CASE 
                            WHEN PO.[CLOSE] = N'2' THEN 0
                            WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                            ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
                        END AS WIP_QTY,
                        Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_NAME,
                        ZTP.Z_TESTING_PROGRAM_NAME,
                        Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE,
                        Z_WIRE.Z_WIRE_NAME,
                        Z_PACKAGE.REMARK,
                        CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
                        CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
                        PO.SUPPLIER_FULL_NAME,
                        PO_SD.RECEIPT_CLOSE
                    FROM PURCHASE_ORDER PO
                    LEFT JOIN PURCHASE_ORDER_D PO_D 
                        ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                    LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                    LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                    LEFT JOIN Z_OUT_MO_D 
                        ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                    LEFT JOIN ITEM 
                        ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                    LEFT JOIN ITEM_LOT 
                        ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                    LEFT JOIN Z_ASSEMBLY_CODE 
                        ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_ASSEMBLY_CODE ZAC
                        ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_TESTING_PROGRAM ZTP
                        ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                    LEFT JOIN Z_PACKAGE 
                        ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                    LEFT JOIN Z_PROCESSING_PURPOSE 
                        ON Z_ASSEMBLY_CODE.Z_PROCESSING_PURPOSE_ID = Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_ID
                    LEFT JOIN Z_LOADING_METHOD 
                        ON Z_LOADING_METHOD.Z_LOADING_METHOD_ID = Z_PACKAGE.Z_LOADING_METHOD_ID
                    LEFT JOIN Z_WIRE 
                        ON Z_WIRE.Z_WIRE_ID = Z_PACKAGE.Z_WIRE_ID
                    LEFT JOIN FEATURE_GROUP 
                        ON FEATURE_GROUP.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                    OUTER APPLY (
                        SELECT TOP 1 *
                        FROM PURCHASE_RECEIPT_D PRD
                        WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                        ORDER BY PRD.CreateDate
                    ) PR
                    WHERE PO.PURCHASE_TYPE = 2 
                        AND PO.PURCHASE_DATE > '2024-10-21' 
                        AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                        AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                ) AS CombinedResults
                WHERE 1=1
            """

            # 构建查询条件
            conditions = []
            query_params = {}
            
            # 参数验证和清理
            if params.doc_no and isinstance(params.doc_no, str):
                conditions.append("AND UPPER(DOC_NO) LIKE UPPER(:doc_no)")
                query_params["doc_no"] = f"%{self._clean_input(params.doc_no)}%"
                
            if params.item_code and isinstance(params.item_code, str):
                conditions.append("AND UPPER(ITEM_CODE) LIKE UPPER(:item_code)")
                query_params["item_code"] = f"%{self._clean_input(params.item_code)}%"
            
            if params.lot_code and isinstance(params.lot_code, str):
                conditions.append("AND UPPER(LOT_CODE) LIKE UPPER(:lot_code)")
                query_params["lot_code"] = f"%{self._clean_input(params.lot_code)}%"
            
            if params.supplier and isinstance(params.supplier, str):
                conditions.append("AND SUPPLIER_FULL_NAME LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"
                
            if params.package_type and isinstance(params.package_type, str):
                conditions.append("AND UPPER(Z_PACKAGE_TYPE_NAME) LIKE UPPER(:package_type)")
                query_params["package_type"] = f"%{self._clean_input(params.package_type)}%"
                
            if params.assembly_code and isinstance(params.assembly_code, str):
                conditions.append("AND UPPER(Z_ASSEMBLY_CODE) LIKE UPPER(:assembly_code)")
                query_params["assembly_code"] = f"%{self._clean_input(params.assembly_code)}%"
                
            if params.is_closed is not None:
                if params.is_closed == 0:
                    conditions.append("AND RECEIPT_CLOSE = 0")
                else:
                    conditions.append("AND RECEIPT_CLOSE != 0")

            if params.order_date_start:
                conditions.append("AND PURCHASE_DATE >= :order_date_start")
                query_params["order_date_start"] = params.order_date_start

            if params.order_date_end:
                conditions.append("AND PURCHASE_DATE <= :order_date_end")
                query_params["order_date_end"] = params.order_date_end
                
            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加排序
            query += " ORDER BY PURCHASE_DATE, DOC_NO"
            
            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()
            
            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1) 
                FROM (
                    SELECT
                        hpl.ID,
                        hpl.DOC_NO,
                        hpl.ITEM_CODE,
                        hpl.Z_PACKAGE_TYPE_NAME,
                        hpl.LOT_CODE,
                        hpl.BUSINESS_QTY,
                        hpl.RECEIPTED_PRICE_QTY,
                        0 AS WIP_QTY,
                        hpl.Z_PROCESSING_PURPOSE_NAME,
                        hpl.Z_TESTING_PROGRAM_NAME,
                        hpl.Z_ASSEMBLY_CODE,
                        hpl.Z_WIRE_NAME,
                        hpl.REMARK,
                        hpl.PURCHASE_DATE,
                        ISNULL(hpl.FIRST_ARRIVAL_DATE, DATEADD(MONTH, 2, hpl.PURCHASE_DATE)) AS FIRST_ARRIVAL_DATE,
                        hpl.SUPPLIER_FULL_NAME,
                        hpl.RECEIPT_CLOSE
                    FROM HSUN_PACKAGE_LIST hpl
                    UNION ALL
                    SELECT
                        ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE, PO.DOC_NO) + 115617 AS ID,
                        PO.DOC_NO,
                        ITEM.ITEM_CODE,
                        Z_PACKAGE_TYPE.Z_PACKAGE_TYPE_NAME,
                        ITEM_LOT.LOT_CODE,
                        CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
                        CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
                        CASE 
                            WHEN PO.[CLOSE] = N'2' THEN 0
                            WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                            ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
                        END AS WIP_QTY,
                        Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_NAME,
                        ZTP.Z_TESTING_PROGRAM_NAME,
                        Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE,
                        Z_WIRE.Z_WIRE_NAME,
                        Z_PACKAGE.REMARK,
                        CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
                        CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
                        PO.SUPPLIER_FULL_NAME,
                        PO_SD.RECEIPT_CLOSE
                    FROM PURCHASE_ORDER PO
                    LEFT JOIN PURCHASE_ORDER_D PO_D 
                        ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                    LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                    LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                    LEFT JOIN Z_OUT_MO_D 
                        ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                    LEFT JOIN ITEM 
                        ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                    LEFT JOIN ITEM_LOT 
                        ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                    LEFT JOIN Z_ASSEMBLY_CODE 
                        ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_ASSEMBLY_CODE ZAC
                        ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_TESTING_PROGRAM ZTP
                        ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                    LEFT JOIN Z_PACKAGE 
                        ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                    LEFT JOIN Z_PROCESSING_PURPOSE 
                        ON Z_ASSEMBLY_CODE.Z_PROCESSING_PURPOSE_ID = Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_ID
                    LEFT JOIN Z_LOADING_METHOD 
                        ON Z_LOADING_METHOD.Z_LOADING_METHOD_ID = Z_PACKAGE.Z_LOADING_METHOD_ID
                    LEFT JOIN Z_WIRE 
                        ON Z_WIRE.Z_WIRE_ID = Z_PACKAGE.Z_WIRE_ID
                    LEFT JOIN Z_PACKAGE_TYPE 
                        ON Z_PACKAGE_TYPE.Z_PACKAGE_TYPE_ID = Z_PACKAGE.Z_PACKAGE_TYPE_ID
                    LEFT JOIN FEATURE_GROUP 
                        ON FEATURE_GROUP.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                    OUTER APPLY (
                        SELECT TOP 1 *
                        FROM PURCHASE_RECEIPT_D PRD
                        WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                        ORDER BY PRD.CreateDate
                    ) PR
                    WHERE PO.PURCHASE_TYPE = 2 
                        AND PO.PURCHASE_DATE > '2024-10-21' 
                        AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                        AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                ) t
                WHERE 1=1 {' '.join(conditions)}
            """
            total = db.execute(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()
            
            # 转换为响应对象
            assy_orders = [
                AssyOrder(
                    ID=row.ID,
                    DOC_NO=row.DOC_NO,
                    ITEM_CODE=row.ITEM_CODE,
                    Z_PACKAGE_TYPE_NAME=row.Z_PACKAGE_TYPE_NAME,
                    LOT_CODE=row.LOT_CODE,
                    BUSINESS_QTY=row.BUSINESS_QTY,
                    RECEIPTED_PRICE_QTY=row.RECEIPTED_PRICE_QTY,
                    WIP_QTY=row.WIP_QTY,
                    Z_PROCESSING_PURPOSE_NAME=row.Z_PROCESSING_PURPOSE_NAME,
                    Z_TESTING_PROGRAM_NAME=row.Z_TESTING_PROGRAM_NAME,
                    Z_ASSEMBLY_CODE=row.Z_ASSEMBLY_CODE,
                    Z_WIRE_NAME=row.Z_WIRE_NAME,
                    REMARK=row.REMARK,
                    PURCHASE_DATE=row.PURCHASE_DATE,
                    FIRST_ARRIVAL_DATE=row.FIRST_ARRIVAL_DATE,
                    SUPPLIER_FULL_NAME=row.SUPPLIER_FULL_NAME,
                    RECEIPT_CLOSE=row.RECEIPT_CLOSE
                ) for row in result
            ]
            return {
                "list": assy_orders,
                "total": total or 0
            }
        except Exception as e:
            logger.error(f"查询封装订单失败: {str(e)}")
            raise CustomException("查询封装订单失败")
    
    def get_assy_bom_by_params(self, db: Session, params: AssyBomQuery) -> Dict[str, Any]:
        """根据参数获取封装订单BOM"""
        try:
            # 构建基础查询
            base_query = """
                SELECT * FROM
                (
                SELECT * FROM HSUN_BOM_LIST
                UNION ALL
                SELECT 
                ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE,PO.DOC_NO) + 16820 AS ID,
                PO.DOC_NO,
                ZOMSD.Z_MAIN_CHIP,
                ITEM.ITEM_CODE,
                ITEM.ITEM_NAME,
                IL.LOT_CODE,
                CAST(ZOMSD.BUSINESS_QTY AS FLOAT) AS BUSINESS_QTY,
                CAST(ZOMSD.SECOND_QTY AS FLOAT) AS SECOND_QTY,
                ZOMSD.Z_WF_ID_STRING
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D
                ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD
                ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD
                ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D ZOMD
                ON ZOMD.Z_OUT_MO_D_ID = PO_SSD.REFERENCE_SOURCE_ID_ROid
                LEFT JOIN Z_OUT_MO_SD ZOMSD
                ON ZOMSD.Z_OUT_MO_D_ID = ZOMD.Z_OUT_MO_D_ID
                LEFT JOIN ITEM
                ON ZOMSD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT IL
                ON IL.ITEM_LOT_ID = ZOMSD.ITEM_LOT_ID
                WHERE PO_D.PURCHASE_TYPE=2)
                AS ALL_BOM
                WHERE 1=1
            """

            # 添加查询条件
            if params.doc_no:
                base_query += " AND DOC_NO = :doc_no"

            # 执行查询
            result = db.execute(text(base_query).bindparams(doc_no=params.doc_no)).all()
            
            # 构造返回结果
            return {
                "list": [
                    AssyBom(
                        MAIN_CHIP=row.MAIN_CHIP,
                        ITEM_CODE=row.ITEM_CODE,
                        ITEM_NAME=row.ITEM_NAME,
                        LOT_CODE_NAME=row.LOT_CODE_NAME,
                        BUSINESS_QTY=row.BUSINESS_QTY,
                        SECOND_QTY=row.SECOND_QTY,
                        WAFER_ID=row.WAFER_ID
                    ) for row in result
                ]
            }
        except Exception as e:
            logger.error(f"获取封装订单BOM失败: {str(e)}")
            raise CustomException("获取封装订单BOM失败")
    
    def export_assy_order_to_excel(self, db: Session, params: AssyOrderQuery) -> bytes:
        """导出封装订单数据到Excel"""
        try:
            # 获取数据
            params.pageIndex = 1
            params.pageSize = 10000000
            result = self.get_assy_order_by_params(db, params)
            assy_orders = result["list"]
            
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "封装订单"
            
            # 定义表头
            headers = [
                "订单编号", "芯片名称", "封装形式", "打印批号", "订单数量", 
                "收货数量", "在制数量", "加工方式", "测试程序", 
                "打线图号", "打线材料", "备注", "订单日期", "首到日期", 
                "供应商", "状态"
            ]
            
            # 设置列宽
            column_widths = {
                'A': 15,  # 订单号
                'B': 20,  # 品号
                'C': 15,  # 封装形式
                'D': 15,  # 打印批号
                'E': 12,  # 订单数量
                'F': 12,  # 收货数量
                'G': 12,  # 在制数量
                'H': 15,  # 加工方式
                'I': 15,  # 测试程序
                'J': 15,  # 打线图号
                'K': 10,  # 打线材料
                'L': 20,  # 备注
                'M': 12,  # 订单日期
                'N': 12,  # 首到日期
                'O': 25,  # 供应商
                'P': 10   # 状态
            }
            
            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            for row, order in enumerate(assy_orders, 2):
                data = [
                    order.DOC_NO,
                    order.ITEM_CODE,
                    order.Z_PACKAGE_TYPE_NAME,
                    order.LOT_CODE,
                    order.BUSINESS_QTY,
                    order.RECEIPTED_PRICE_QTY,
                    order.WIP_QTY,
                    order.Z_PROCESSING_PURPOSE_NAME,
                    order.Z_TESTING_PROGRAM_NAME,
                    order.Z_ASSEMBLY_CODE,
                    order.Z_WIRE_NAME,
                    order.REMARK,
                    order.PURCHASE_DATE.strftime('%Y-%m-%d') if order.PURCHASE_DATE else '',
                    order.FIRST_ARRIVAL_DATE.strftime('%Y-%m-%d') if order.FIRST_ARRIVAL_DATE else '',
                    order.SUPPLIER_FULL_NAME,
                    '已关闭' if order.RECEIPT_CLOSE else '未关闭'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border
                    
                    # 设置数字列的格式
                    if col in [5, 6, 7]:  # 订单数量、已收货数量、在制数量
                        cell.number_format = '#,##0'
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"导出封装订单Excel失败: {str(e)}")
            raise CustomException("导出封装订单Excel失败")
        
    def get_assy_wip_by_params(self,db:Session,params:AssyWipQuery)->Dict[str,Any]:
        """获取封装在制"""
        try:
            # 构建基础查询
            base_query = """
                SELECT 
                PO.DOC_NO AS '订单号',
                PO.SUPPLIER_FULL_NAME,
                ITEM.ITEM_CODE,
                CASE
                  WHEN ZPP.Z_PROCESSING_PURPOSE_NAME IS NULL THEN '封装'
                  ELSE ZPP.Z_PROCESSING_PURPOSE_NAME
                 END AS Z_PROCESSING_PURPOSE_NAME,
                CASE 
                    WHEN WIP.[当前工序] = N'已完成' THEN NULL
                    ELSE DATEDIFF(DAY, WIP.modified_at, GETDATE()) 
                END
                AS STRANDED,
                ISNULL(WIP.[当前工序], '需确认') AS 当前工序,
                WIP.[预计交期],
                WIP.finished_at,
                CAST((PO_D.BUSINESS_QTY * 0.9989- PO_D.RECEIPTED_PRICE_QTY) AS INT) AS 在线合计,
                ISNULL(WIP.[仓库库存], 0) AS 仓库库存,
                ISNULL(WIP.[扣留信息], '') AS 扣留信息,
                ISNULL(WIP.[次日预计], 0) AS 次日预计,
                ISNULL(WIP.[三日预计], 0) AS 三日预计,
                ISNULL(WIP.[七日预计], 0) AS 七日预计,
                ISNULL(WIP.[研磨], 0) AS 研磨,
                ISNULL(WIP.[切割], 0) AS 切割,
                ISNULL(WIP.[待装片], 0) AS 待装片,
                ISNULL(WIP.[装片], 0) AS 装片,
                ISNULL(WIP.[银胶固化], 0) AS 银胶固化,
                ISNULL(WIP.[等离子清洗1], 0) AS 等离子清洗1,
                ISNULL(WIP.[键合], 0) AS 键合,
                ISNULL(WIP.[三目检], 0) AS 三目检,
                ISNULL(WIP.[等离子清洗2], 0) AS 等离子清洗2,
                ISNULL(WIP.[塑封], 0) AS 塑封,
                ISNULL(WIP.[后固化], 0) AS 后固化,
                ISNULL(WIP.[回流焊], 0) AS 回流焊,
                ISNULL(WIP.[电镀], 0) AS 电镀,
                ISNULL(WIP.[打印], 0) AS 打印,
                ISNULL(WIP.[后切割], 0) AS 后切割,
                ISNULL(WIP.[切筋成型], 0) AS 切筋成型,
                ISNULL(WIP.[测编打印], 0) AS 测编打印,
                ISNULL(WIP.[外观检], 0) AS 外观检,
                ISNULL(WIP.[包装], 0) AS 包装,
                ISNULL(WIP.[待入库], 0) AS 待入库
                FROM PURCHASE_ORDER PO
                LEFT JOIN huaxinAdmin_wip_assy WIP ON PO.DOC_NO = WIP.[订单号]
                LEFT JOIN PURCHASE_ORDER_D PO_D ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D ZOMD ON PO_SSD.REFERENCE_SOURCE_ID_ROid = ZOMD.Z_OUT_MO_D_ID
                LEFT JOIN ITEM ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT ON ZOMD.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC ON ZOMD.Z_PACKAGE_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_PACKAGE ON ZAC.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                LEFT JOIN Z_PROCESSING_PURPOSE ZPP ON ZAC.Z_PROCESSING_PURPOSE_ID = ZPP.Z_PROCESSING_PURPOSE_ID
                WHERE 1=1 AND 
                  PO.[CLOSE]=0 AND 
                  PO.SUPPLIER_FULL_NAME<>'苏州荐恒电子科技有限公司' AND 
                  (PO.DOC_NO NOT LIKE '3501-%' AND 
                  PO.DOC_NO != 'HX-20240430001') AND 
                  ITEM.ITEM_CODE LIKE N'BC%AB' AND
                  (PO_D.BUSINESS_QTY * 0.9989- PO_D.RECEIPTED_PRICE_QTY)>0
            """
            result = db.execute(text(base_query)).all()
            
            # 构建查询条件
            conditions = []
            query_params = {}
            
            # 参数验证和清理
            if params.doc_no and isinstance(params.doc_no, str):
                conditions.append("AND PO.DOC_NO LIKE :doc_no")
                query_params["doc_no"] = f"%{self._clean_input(params.doc_no)}%"
            
            if params.item_code and isinstance(params.item_code, str):
                conditions.append("AND ITEM.ITEM_CODE LIKE :item_code")
                query_params["item_code"] = f"%{self._clean_input(params.item_code)}%"
                
            if params.supplier and isinstance(params.supplier, str):
                conditions.append("AND PO.SUPPLIER_FULL_NAME LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"
                
            if params.current_process and isinstance(params.current_process, str):
                conditions.append("AND WIP.[当前工序] LIKE :current_process")
                query_params["current_process"] = f"%{self._clean_input(params.current_process)}%"
            
            if params.is_tr is not None:
                if params.is_tr == 1:
                    conditions.append("AND (CASE WHEN ZPP.Z_PROCESSING_PURPOSE_NAME IS NULL THEN '封装' ELSE ZPP.Z_PROCESSING_PURPOSE_NAME END) LIKE N'%编带'")
                else:
                    conditions.append("AND (CASE WHEN ZPP.Z_PROCESSING_PURPOSE_NAME IS NULL THEN '封装' ELSE ZPP.Z_PROCESSING_PURPOSE_NAME END) NOT LIKE N'%编带'")
                    
            if params.is_stranded is not None:
                if params.is_stranded == 0:
                    conditions.append("""
                                      AND CASE 
                                      WHEN WIP.[当前工序] = N'已完成' THEN NULL
                                      ELSE DATEDIFF(DAY, WIP.modified_at, GETDATE()) 
                                      END = 0
                                      """)
                else:
                    conditions.append("""
                                      AND CASE 
                                      WHEN WIP.[当前工序] = N'已完成' THEN NULL
                                      ELSE DATEDIFF(DAY, WIP.modified_at, GETDATE()) 
                                      END != 0
                                      """)
                    
            if params.days is not None:
                if params.days == 1:
                    conditions.append("AND WIP.[次日预计] > 0")
                elif params.days == 3:
                    conditions.append("AND WIP.[三日预计] > 0")
                elif params.days == 7:
                    conditions.append("AND WIP.[七日预计] > 0")
                    
            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加排序
            query += " ORDER BY PO.DOC_DATE"
            
            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize
                
            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()
            
            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1)
                FROM PURCHASE_ORDER PO
                LEFT JOIN huaxinAdmin_wip_assy WIP ON PO.DOC_NO = WIP.[订单号]
                LEFT JOIN PURCHASE_ORDER_D PO_D ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID 
                LEFT JOIN Z_OUT_MO_D ZOMD ON PO_SSD.REFERENCE_SOURCE_ID_ROid = ZOMD.Z_OUT_MO_D_ID
                LEFT JOIN ITEM ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT ON ZOMD.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC ON ZOMD.Z_PACKAGE_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_PACKAGE ON ZAC.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                LEFT JOIN Z_PROCESSING_PURPOSE ZPP ON ZAC.Z_PROCESSING_PURPOSE_ID = ZPP.Z_PROCESSING_PURPOSE_ID 
                WHERE 1=1 AND 
                  PO.[CLOSE]=0 AND 
                  PO.SUPPLIER_FULL_NAME<>'苏州荐恒电子科技有限公司' AND 
                  (PO.DOC_NO NOT LIKE '3501-%' AND 
                  PO.DOC_NO != 'HX-20240430001') AND 
                  ITEM.ITEM_CODE LIKE N'BC%AB' AND
                  (PO_D.BUSINESS_QTY * 0.9989- PO_D.RECEIPTED_PRICE_QTY)>0
                {' '.join(conditions)}
            """
            total = db.execute(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()
            
            # 转换为响应对象
            assy_wips = [
                AssyWip(
                    DOC_NO=row.订单号,
                    SUPPLIER_FULL_NAME=row.SUPPLIER_FULL_NAME,
                    ITEM_CODE=row.ITEM_CODE,
                    Z_PROCESSING_PURPOSE_NAME=row.Z_PROCESSING_PURPOSE_NAME,
                    STRANDED=row.STRANDED,
                    CURRENT_PROCESS=row.当前工序,
                    EXPECTED_DELIVERY_DATE=row.预计交期,
                    FINISHED_AT=row.finished_at,
                    ONLINE_TOTAL=row.在线合计,
                    WAREHOUSE_INVENTORY=row.仓库库存,
                    HOLD_INFO=row.扣留信息,
                    NEXT_DAY_EXPECTED=row.次日预计,
                    THREE_DAY_EXPECTED=row.三日预计,
                    SEVEN_DAY_EXPECTED=row.七日预计,
                    POLISHING=row.研磨,
                    CUTTING=row.切割,
                    WAITING_FOR_INSTALLATION=row.待装片,
                    INSTALLATION=row.装片,
                    SILVER_GLUE_CURE=row.银胶固化,
                    PLASMA_CLEANING_1=row.等离子清洗1,
                    BONDING=row.键合,
                    THREE_POINT_INSPECTION=row.三目检,
                    PLASMA_CLEANING_2=row.等离子清洗2,
                    SEALING=row.塑封,
                    POST_CURE=row.后固化,
                    REFLOW_SOLDERING=row.回流焊,
                    ELECTROPLATING=row.电镀,
                    PRINTING=row.打印,
                    POST_CUTTING=row.后切割,
                    CUTTING_AND_SHAPING=row.切筋成型,
                    MEASUREMENT_AND_PRINTING=row.测编打印,
                    APPEARANCE_INSPECTION=row.外观检,    
                    PACKING=row.包装,
                    WAITING_FOR_WAREHOUSE_INVENTORY=row.待入库
                ) for row in result
            ]
            
            return {
                "list": assy_wips,
                "total": total or 0
            }
            
        except Exception as e:
            logger.error(f"查询封装在制失败: {str(e)}")
            raise CustomException("查询封装在制失败")
    
    def get_assy_order_items(self,db:Session,params:AssyOrderItemsQuery)->Dict[str,Any]:
        """获取封装在制品号"""
        try:
            base_query = """
                SELECT DISTINCT ITEM_CODE
                FROM ITEM
                WHERE ITEM_CODE LIKE N'BC%AB'
            """

            # 构建查询条件
            conditions = []
            query_params = {}

            if params.item_code:
                # 将输入的品号转换为大写
                item_code = params.item_code.upper()
                conditions.append("AND ITEM_CODE LIKE :item_code")
                query_params["item_code"] = f"%{self._clean_input(item_code)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions) 

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 转换为响应对象
            items = [
                {
                    "label": row.ITEM_CODE,
                    "value": row.ITEM_CODE
                } for row in result
            ]

            return {
                "list": items
            }
        except Exception as e:
            logger.error(f"获取封装在制品号失败: {str(e)}")
            raise CustomException("获取封装在制品号失败")
    
    def get_assy_order_package_type(self,db:Session,params:AssyOrderPackageTypeQuery)->Dict[str,Any]:
        """获取封装订单类型"""
        try:
            base_query = """
                SELECT DISTINCT Z_PACKAGE_TYPE_NAME
                FROM Z_PACKAGE_TYPE
                WHERE 1=1
            """

            # 构建查询条件
            conditions = []
            query_params = {}

            if params.package_type:
                # 将输入的封装类型转换为大写
                package_type = params.package_type.upper()
                conditions.append("AND Z_PACKAGE_TYPE_NAME LIKE :package_type")
                query_params["package_type"] = f"%{self._clean_input(package_type)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 转换为响应对象
            package_types = [
                {
                    "label": row.Z_PACKAGE_TYPE_NAME,
                    "value": row.Z_PACKAGE_TYPE_NAME
                } for row in result
            ]   

            return {
                "list": package_types
            }
        except Exception as e:
            logger.error(f"获取封装订单类型失败: {str(e)}")
            raise CustomException("获取封装订单类型失败")
        
    def get_assy_order_supplier(self,db:Session,params:AssyOrderSupplierQuery)->Dict[str,Any]:
        """获取封装订单供应商"""
        try:
            base_query = """
                SELECT DISTINCT SUPPLIER_FULL_NAME
                FROM PURCHASE_ORDER
                WHERE 1=1
            """

            # 构建查询条件
            conditions = []
            query_params = {}

            if params.supplier:
                conditions.append("AND SUPPLIER_FULL_NAME LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 转换为响应对象
            suppliers = [
                {
                    "label": row.SUPPLIER_FULL_NAME,
                    "value": row.SUPPLIER_FULL_NAME
                } for row in result
            ]

            return {
                "list": suppliers
            }
        except Exception as e:
            logger.error(f"获取封装订单供应商失败: {str(e)}")
            raise CustomException("获取封装订单供应商失败")
        
    def get_stock_by_params(self,db:Session,params:StockQuery)->Dict[str,Any]:
        """获取库存"""
        try:
            base_query = """
                        SELECT 
                        FG.FEATURE_GROUP_NAME,
                        ITEM.ITEM_CODE,
                        ITEM.ITEM_NAME,
                        IL.LOT_CODE,
                        W.WAREHOUSE_NAME,
                        CAST(SUM(A.INVENTORY_QTY) AS INT) AS INVENTORY_QTY,
                        CAST(SUM(ISNULL(A.SECOND_QTY,0)) AS FLOAT) AS SECOND_QTY,
                        B.Z_BIN_LEVEL_NAME,
                        T.Z_TESTING_PROGRAM_NAME,
                        BP.Z_BURNING_PROGRAM_NAME
                        FROM Z_WF_IC_WAREHOUSE_BIN A
                        LEFT JOIN ITEM
                        ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                        LEFT JOIN ITEM_LOT IL
                        ON IL.ITEM_LOT_ID = A.ITEM_LOT_ID
                        LEFT JOIN FEATURE_GROUP FG
                        ON FG.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                        LEFT JOIN Z_BIN_LEVEL B
                        ON B.Z_BIN_LEVEL_ID = A.Z_BIN_LEVEL_ID
                        LEFT JOIN Z_TESTING_PROGRAM T
                        ON T.Z_TESTING_PROGRAM_ID = A.Z_TESTING_PROGRAM_ID
                        LEFT JOIN Z_BURNING_PROGRAM BP
                        ON BP.Z_BURNING_PROGRAM_ID = A.Z_BURNING_PROGRAM_ID
                        LEFT JOIN WAREHOUSE W
                        ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                        GROUP BY FG.FEATURE_GROUP_NAME,ITEM.ITEM_CODE,ITEM.ITEM_NAME,IL.LOT_CODE,W.WAREHOUSE_NAME,B.Z_BIN_LEVEL_NAME,
                        T.Z_TESTING_PROGRAM_NAME,BP.Z_BURNING_PROGRAM_NAME
                        HAVING SUM(A.INVENTORY_QTY) > 0
                    """
            # 构建查询条件
            having_conditions = []
            query_params = {}

            # 参数验证和清理
            if params.feature_group_name:
                feature_group_names = [self._clean_input(name) for name in params.feature_group_name]
                feature_group_name_conditions = []
                for i, name in enumerate(feature_group_names):
                    feature_group_name_conditions.append(f"FG.FEATURE_GROUP_NAME LIKE :feature_group_name_{i}")
                    query_params[f"feature_group_name_{i}"] = f"%{name}%"
                having_conditions.append(f"AND ({' OR '.join(feature_group_name_conditions)})")

            if params.item_code:
                item_codes = [self._clean_input(code) for code in params.item_code]
                item_code_conditions = []
                for i, code in enumerate(item_codes):
                    item_code_conditions.append(f"UPPER(ITEM.ITEM_CODE) LIKE UPPER(:item_code_{i})")
                    query_params[f"item_code_{i}"] = f"%{code}%"
                having_conditions.append(f"AND ({' OR '.join(item_code_conditions)})")


            if params.item_name:
                item_names = [self._clean_input(name) for name in params.item_name]
                item_name_conditions = []
                for i, name in enumerate(item_names):
                    item_name_conditions.append(f"UPPER(ITEM.ITEM_NAME) LIKE UPPER(:item_name_{i})")
                    query_params[f"item_name_{i}"] = f"%{name}%"
                having_conditions.append(f"AND ({' OR '.join(item_name_conditions)})")

            if params.lot_code:
                lot_codes = [self._clean_input(lot) for lot in params.lot_code]
                lot_code_conditions = []
                for i, lot in enumerate(lot_codes):
                    lot_code_conditions.append(f"UPPER(IL.LOT_CODE) LIKE UPPER(:lot_code_{i})")
                    query_params[f"lot_code_{i}"] = f"%{lot}%"
                having_conditions.append(f"AND ({' OR '.join(lot_code_conditions)})")

            if params.warehouse_name:
                warehouse_names = [self._clean_input(name) for name in params.warehouse_name]
                warehouse_name_conditions = []
                for i, name in enumerate(warehouse_names):
                    warehouse_name_conditions.append(f"W.WAREHOUSE_NAME LIKE :warehouse_name_{i}")
                    query_params[f"warehouse_name_{i}"] = f"%{name}%"
                having_conditions.append(f"AND ({' OR '.join(warehouse_name_conditions)})")

            if params.testing_program:
                testing_programs = [self._clean_input(name) for name in params.testing_program]
                placeholders = [f":testing_program_{i}" for i in range(len(testing_programs))]
                having_conditions.append(f"AND UPPER(T.Z_TESTING_PROGRAM_NAME) IN ({','.join([f'UPPER({p})' for p in placeholders])})")
                for i, name in enumerate(testing_programs):
                    query_params[f"testing_program_{i}"] = name

            if params.burning_program:
                burning_programs = [self._clean_input(name) for name in params.burning_program]
                placeholders = [f":burning_program_{i}" for i in range(len(burning_programs))]
                having_conditions.append(f"AND UPPER(BP.Z_BURNING_PROGRAM_NAME) IN ({','.join([f'UPPER({p})' for p in placeholders])})")
                for i, name in enumerate(burning_programs):
                    query_params[f"burning_program_{i}"] = name
        
            query = base_query + " " + " ".join(having_conditions)

            # 构建完整的 SQL 查询语句
            if having_conditions:
                query = f"{base_query} {' '.join(having_conditions)} ORDER BY ITEM.ITEM_CODE,IL.LOT_CODE"
            else:
                query = f"{base_query} ORDER BY ITEM.ITEM_CODE,IL.LOT_CODE"

            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            result = db.execute(text(query), query_params)
            rows = result.fetchall()

            # 获取总记录数 - 使用COUNT聚合函数
            count_query = f"""
                SELECT COUNT(*) as total_count
                FROM (
                    {base_query} {''.join(having_conditions)}
                ) AS subquery
            """
            total_count = db.execute(text(count_query), query_params).scalar_one()

            # 转换为响应对象
            stocks = [
                Stock(
                    FEATURE_GROUP_NAME=row.FEATURE_GROUP_NAME,
                    ITEM_CODE=row.ITEM_CODE,
                    ITEM_NAME=row.ITEM_NAME,
                    LOT_CODE=row.LOT_CODE,
                    WAREHOUSE_NAME=row.WAREHOUSE_NAME,
                    INVENTORY_QTY=row.INVENTORY_QTY,
                    SECOND_QTY=row.SECOND_QTY,
                    Z_BIN_LEVEL_NAME=row.Z_BIN_LEVEL_NAME,
                    Z_TESTING_PROGRAM_NAME=row.Z_TESTING_PROGRAM_NAME,
                    Z_BURNING_PROGRAM_NAME=row.Z_BURNING_PROGRAM_NAME
                ) for row in rows
            ]
            return {
                "list": stocks,
                "total": total_count
            }
        except Exception as e:
            logger.error(f"查询库存失败: {str(e)}")
            raise CustomException("查询库存失败")
    
    def get_wafer_id_qty_detail_by_params(self,db:Session,params:WaferIdQtyDetailQuery)->List[WaferIdQtyDetail]:
        """根据参数获取晶圆ID数量明细"""
        try:
            base_query = """
            SELECT
                ITEM.ITEM_CODE,ITEM.ITEM_NAME,
                IL.LOT_CODE,
                T.Z_TESTING_PROGRAM_NAME,
                (A.WF_ID + '#') AS WF_ID,
                B.Z_BIN_LEVEL_NAME,
                CAST(A.INVENTORY_QTY AS INT) AS INVENTORY_QTY,CAST(A.SECOND_QTY AS FLOAT) AS SECOND_QTY,
                W.WAREHOUSE_NAME
                FROM Z_WF_IC_WAREHOUSE_BIN A
                LEFT JOIN ITEM
                ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                LEFT JOIN ITEM_LOT IL
                ON IL.ITEM_LOT_ID = A.ITEM_LOT_ID
                LEFT JOIN FEATURE_GROUP FG
                ON FG.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                LEFT JOIN Z_BIN_LEVEL B
                ON B.Z_BIN_LEVEL_ID = A.Z_BIN_LEVEL_ID
                LEFT JOIN Z_TESTING_PROGRAM T
                ON T.Z_TESTING_PROGRAM_ID = A.Z_TESTING_PROGRAM_ID
                LEFT JOIN WAREHOUSE W
                ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                WHERE A.INVENTORY_QTY > 0
                """
            # 构建查询条件
            conditions = []
            query_params = {}

            if params.item_code:
                conditions.append("AND ITEM.ITEM_CODE LIKE :item_code")
                query_params["item_code"] = f"%{self._clean_input(params.item_code)}%"

            if params.lot_code:
                conditions.append("AND IL.LOT_CODE LIKE :lot_code")
                query_params["lot_code"] = f"%{self._clean_input(params.lot_code)}%"
                
            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            # 添加排序
            query += " ORDER BY ITEM.ITEM_CODE,IL.LOT_CODE,CAST(A.WF_ID AS INT)"
            
            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 转换为响应对象
            wafer_id_qty_details = [
                WaferIdQtyDetail(
                    ITEM_CODE=row.ITEM_CODE,
                    LOT_CODE=row.LOT_CODE,
                    Z_TESTING_PROGRAM_NAME=row.Z_TESTING_PROGRAM_NAME,
                    WF_ID=row.WF_ID,
                    Z_BIN_LEVEL_NAME=row.Z_BIN_LEVEL_NAME,
                    INVENTORY_QTY=row.INVENTORY_QTY,
                    SECOND_QTY=row.SECOND_QTY,
                    WAREHOUSE_NAME=row.WAREHOUSE_NAME
                ) for row in result
            ]

            return {"list": wafer_id_qty_details}
        
        except Exception as e:
            logger.error(f"获取晶圆ID数量明细失败: {str(e)}")
            raise CustomException(status_code=500, message="获取晶圆ID数量明细失败")
    
    def get_stock_summary_by_params(self,db:Session,params:StockSummaryQuery)->List[StockSummary]:
        """根据参数获取库存汇总"""
        try:
            base_query = """
                SELECT 
                FG.FEATURE_GROUP_NAME,
                ITEM.ITEM_NAME,
                W.WAREHOUSE_NAME,
                CAST(SUM(A.INVENTORY_QTY) AS INT) AS INVENTORY_QTY,
                AVG(DATEDIFF(Month, ITEM.CreateDate , GETDATE())) AS AVERAGE_STOCK_AGE
                FROM Z_WF_IC_WAREHOUSE_BIN A
                LEFT JOIN ITEM
                ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                LEFT JOIN FEATURE_GROUP FG
                ON FG.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                LEFT JOIN WAREHOUSE W
                ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                WHERE W.WAREHOUSE_NAME NOT LIKE N'%原材料%'
            """
            # 构建查询条件
            conditions = []
            query_params = {}
            if params.item_name:
                conditions.append("AND ITEM.ITEM_NAME LIKE :item_name")
                query_params["item_name"] = f"%{self._clean_input(params.item_name)}%"

            if params.feature_group_name:
                conditions.append("AND FG.FEATURE_GROUP_NAME LIKE :feature_group_name")
                query_params["feature_group_name"] = f"%{self._clean_input(params.feature_group_name)}%"

            if params.warehouse_name:
                conditions.append("AND UPPER(W.WAREHOUSE_NAME) LIKE UPPER(:warehouse_name)")
                query_params["warehouse_name"] = f"%{self._clean_input(params.warehouse_name)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)

            query += """ GROUP BY FG.FEATURE_GROUP_NAME,ITEM.ITEM_NAME,W.WAREHOUSE_NAME
                        HAVING SUM(A.INVENTORY_QTY) > 0
                        ORDER BY ITEM.ITEM_NAME
                    """
            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 转换为响应对象
            stock_summaries = [
                StockSummary(
                    FEATURE_GROUP_NAME=row.FEATURE_GROUP_NAME,
                    ITEM_NAME=row.ITEM_NAME,
                    WAREHOUSE_NAME=row.WAREHOUSE_NAME,
                    INVENTORY_QTY=row.INVENTORY_QTY,
                    AVERAGE_STOCK_AGE=row.AVERAGE_STOCK_AGE
                ) for row in result
            ]
            return {"list": stock_summaries}
        except Exception as e:
            logger.error(f"获取库存汇总失败: {str(e)}")
            raise CustomException("获取库存汇总失败")

    def export_stock_by_params(self,db:Session,params:StockQuery)->bytes:
        """导出库存数据到Excel"""
        try:
            # 获取库存数据
            result = self.get_stock_by_params(db, params)
            stock_list = result["list"]
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"
            
            # 定义表头
            headers = ["品号群组", "物料名称", "物料编码", "批号", "BIN等级", "测试程序", "烧录程序","仓库","库存数量","数量片数"]
            
            # 设置列宽
            column_widths = {
               'A': 15,
               'B': 25,
               'C': 30,
               'D': 20,
               'E': 10,
               'F': 20,
               'G': 20,
               'H': 20,
               'I': 12,
               'J': 10
            }
            
            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            for row, order in enumerate(stock_list, 2):
                data = [
                    order.FEATURE_GROUP_NAME if order.FEATURE_GROUP_NAME else '',
                    order.ITEM_NAME if order.ITEM_NAME else '',
                    order.ITEM_CODE if order.ITEM_CODE else '',
                    order.LOT_CODE if order.LOT_CODE else '',
                    order.Z_BIN_LEVEL_NAME if order.Z_BIN_LEVEL_NAME else '',
                    order.Z_TESTING_PROGRAM_NAME if order.Z_TESTING_PROGRAM_NAME else '',
                    order.Z_BURNING_PROGRAM_NAME if order.Z_BURNING_PROGRAM_NAME else '',
                    order.WAREHOUSE_NAME if order.WAREHOUSE_NAME else '',
                    order.INVENTORY_QTY if order.INVENTORY_QTY is not None else 0,
                    order.SECOND_QTY if order.SECOND_QTY is not None else 0.0
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border
                    
                    # 设置数字列的格式
                    if col in [9, 10]:  # 库存数量、数量片数
                        cell.number_format = '#,##0'
                    
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"导出库存Excel失败: {str(e)}")
            raise CustomException("导出库存Excel失败")
        
    def get_global_report(self,db:Session)->List[GlobalReport]:
        """获取综合报表"""
        try:
            base_query = text("""
                WITH SD AS(
                    SELECT 
                        STK.ITEM_NAME,
                        STK.TOTAL_FINISHED_GOODS,
                        STK.TOP_FINISHED_GOODS,
                        STK.BACK_FINISHED_GOODS,
                        STK.TOTAL_SEMI_MANUFACTURED,
                        STK.TOP_SEMI_MANUFACTURED,
                        STK.BACK_SEMI_MANUFACTURED,
                        SGK.SG_QTY,
                        SGK.SG_FINISHED_GOODS,
                        SGK.SG_SEMI_MANUFACTURED,
                        STK.TOTAL_RAW_MATERIALS,
                        NO_TESTED_WAFER,
                        TESTED_WAFER,
                        OUTSOURCING_WAFER
                    FROM
                        (
                        SELECT 
                        ITEM_NAME,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CP%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS TOTAL_FINISHED_GOODS,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CP%ZY%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS TOP_FINISHED_GOODS,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CP%BY%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS BACK_FINISHED_GOODS,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'BC%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS TOTAL_SEMI_MANUFACTURED,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'BC%ZY%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS TOP_SEMI_MANUFACTURED,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'BC%BY%' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS BACK_SEMI_MANUFACTURED,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CL%' THEN SECOND_QTY ELSE 0 END) AS FLOAT) AS TOTAL_RAW_MATERIALS,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CL%WF' THEN SECOND_QTY ELSE 0 END) AS FLOAT) AS NO_TESTED_WAFER,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CL%CP' THEN SECOND_QTY ELSE 0 END) AS FLOAT) AS TESTED_WAFER,
                        CAST(SUM(CASE WHEN ITEM_CODE LIKE N'CL%WG' THEN INVENTORY_QTY ELSE 0 END) AS INT) AS OUTSOURCING_WAFER
                        FROM
                        (
                            SELECT 
                            ITEM.ITEM_CODE,
                            CASE 
                                WHEN ITEM.ITEM_NAME = N'HS6601MX-16H-SOP8-H' THEN N'HS6601MX-SOP8-H'
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-BY' THEN REPLACE(ITEM.ITEM_NAME, '-BY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZY' THEN REPLACE(ITEM.ITEM_NAME, '-ZY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZB' THEN REPLACE(ITEM.ITEM_NAME, '-ZB', '')
                                ELSE ITEM.ITEM_NAME
                            END AS ITEM_NAME,
                            SUM(A.INVENTORY_QTY) AS INVENTORY_QTY,
                            SUM(CASE WHEN A.SECOND_QTY = 0 THEN A.INVENTORY_QTY ELSE A.SECOND_QTY END) AS SECOND_QTY
                            FROM ITEM_WAREHOUSE_BIN A
                            LEFT JOIN ITEM
                                ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                            LEFT JOIN WAREHOUSE W
                                ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                            WHERE A.INVENTORY_QTY > 0 AND (W.WAREHOUSE_NAME != '实验仓-TH（产成品）') AND (W.WAREHOUSE_NAME NOT LIKE 'HOLD仓%') AND (W.WAREHOUSE_NAME NOT LIKE '华新源%')
                            GROUP BY
                                ITEM.ITEM_CODE,
                                CASE 
                                    WHEN ITEM.ITEM_NAME = N'HS6601MX-16H-SOP8-H' THEN N'HS6601MX-SOP8-H'
                                    WHEN RIGHT(ITEM.ITEM_NAME,3)='-BY' THEN REPLACE(ITEM.ITEM_NAME, '-BY', '') 
                                    WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZY' THEN REPLACE(ITEM.ITEM_NAME, '-ZY', '') 
                                    WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZB' THEN REPLACE(ITEM.ITEM_NAME, '-ZB', '')
                                    ELSE ITEM.ITEM_NAME
                                END
                            ) AS NT
                        GROUP BY ITEM_NAME
                        ) AS STK
                    LEFT JOIN 
                        (SELECT 
                            CASE 
                                WHEN ITEM.ITEM_NAME = N'HS6601MX-16H-SOP8-H' THEN N'HS6601MX-SOP8-H'
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-BY' THEN REPLACE(ITEM.ITEM_NAME, '-BY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZY' THEN REPLACE(ITEM.ITEM_NAME, '-ZY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZB' THEN REPLACE(ITEM.ITEM_NAME, '-ZB', '')
                                ELSE ITEM.ITEM_NAME
                            END AS ITEM_NAME,
                            CAST(SUM(A.INVENTORY_QTY) AS INT) AS SG_QTY,
                            CAST(SUM(CASE WHEN W.WAREHOUSE_NAME=N'苏工院（半成品）' THEN A.INVENTORY_QTY ELSE 0 END) AS INT) AS SG_SEMI_MANUFACTURED,
                            CAST(SUM(CASE WHEN W.WAREHOUSE_NAME=N'苏工院（产成品）' THEN A.INVENTORY_QTY ELSE 0 END) AS INT) AS SG_FINISHED_GOODS
                            FROM ITEM_WAREHOUSE_BIN A
                            LEFT JOIN ITEM
                            ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                            LEFT JOIN WAREHOUSE W
                            ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                            WHERE A.INVENTORY_QTY > 0 AND (W.WAREHOUSE_NAME = N'苏工院（半成品）' OR W.WAREHOUSE_NAME = N'苏工院（产成品）')
                            GROUP BY 
                            CASE 
                                WHEN ITEM.ITEM_NAME = N'HS6601MX-16H-SOP8-H' THEN N'HS6601MX-SOP8-H'
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-BY' THEN REPLACE(ITEM.ITEM_NAME, '-BY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZY' THEN REPLACE(ITEM.ITEM_NAME, '-ZY', '') 
                                WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZB' THEN REPLACE(ITEM.ITEM_NAME, '-ZB', '')
                                ELSE ITEM.ITEM_NAME
                            END) AS SGK
                        ON STK.ITEM_NAME = SGK.ITEM_NAME
                        ),
                        
                    WIP AS
                    (	
                        SELECT
                        ITEM_NAME,
                        SUM(CASE WHEN (ITEM_CODE LIKE N'CL%WF' OR ITEM_CODE LIKE N'CL%WG') THEN WIP_QTY ELSE 0 END ) AS PURCHASE_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'CL%WG' THEN WIP_QTY ELSE 0 END ) AS OUTSOURCING_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'BC%AB' THEN WIP_QTY ELSE 0 END ) AS PACKAGE_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'BC%ZY%AB' THEN WIP_QTY ELSE 0 END ) AS PACKAGE_TOP_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'BC%BY%AB' THEN WIP_QTY ELSE 0 END ) AS PACKAGE_BACK_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'CL%CP' THEN WIP_QTY ELSE 0 END ) AS CP_WIP_QTY,
                        SUM(CASE WHEN ITEM_CODE LIKE N'CP%' THEN WIP_QTY ELSE 0 END ) AS SECONDARY_OUTSOURCING_WIP_QTY
                        FROM
                        (	
                        SELECT NT.ITEM_CODE,NT.ITEM_NAME,SUM(NT.WIP_QTY) AS WIP_QTY
                        FROM
                        (
                            SELECT
                            ITEM.ITEM_CODE,
                            CASE 
                            WHEN ITEM.ITEM_NAME='GC1808-TSSOP14-D-ZY' THEN 'GC1808D-TSSOP14-D-ZY'
                            WHEN RIGHT(ITEM.ITEM_NAME,3)='-BY' THEN REPLACE(ITEM.ITEM_NAME, '-BY', '') 
                            WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZY' THEN REPLACE(ITEM.ITEM_NAME, '-ZY', '') 
                            WHEN RIGHT(ITEM.ITEM_NAME,3)='-ZB' THEN REPLACE(ITEM.ITEM_NAME, '-ZB', '')
                            ELSE ITEM.ITEM_NAME
                            END AS ITEM_NAME,
                            CASE 
                            WHEN PO.[CLOSE] <> N'0' THEN 0
                            WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.9989 THEN 0
                            ELSE CAST((PO_D.BUSINESS_QTY - PO_D.RECEIPTED_PRICE_QTY) AS INT)
                            END AS WIP_QTY
                            FROM PURCHASE_ORDER PO
                            LEFT JOIN PURCHASE_ORDER_D PO_D ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                            LEFT JOIN PURCHASE_ORDER_SD PO_SD ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                            LEFT JOIN ITEM ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                            WHERE PO_SD.RECEIPT_CLOSE = 0
                            UNION ALL 
                            SELECT ITEM.ITEM_CODE,ITEM.ITEM_NAME,(MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY) AS WIP_QTY
                            FROM MO
                            LEFT JOIN MO_D ON MO.MO_ID = MO_D.MO_ID
                            LEFT JOIN MO_PRODUCT MP ON MP.MO_ID = MO.MO_ID
                            LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = MO.ITEM_ID
                            WHERE ITEM.ITEM_CODE LIKE N'CL%CP' AND MO.STATUS NOT IN ('Y','y') AND MO.DOC_NO LIKE '5105%'
                            ) AS NT
                            GROUP BY NT.ITEM_CODE,NT.ITEM_NAME
                        ) AS NNT
                        GROUP BY ITEM_NAME
                    ),
                    BL AS
                    (
                    SELECT
                    ROW_NUMBER ( ) OVER ( ORDER BY MAIN_CHIP, DEPUTY_CHIP, CHIP_NAME ) + 10000 AS ROW,BB.*
                    FROM
                    (
                    SELECT 
                    DISTINCT(ITEM_NAME) AS MAIN_CHIP,BOM.DEPUTY_CHIP,BOM.CHIP_NAME,
                    ROW_NUMBER() OVER (PARTITION BY NT.ITEM_NAME ORDER BY NT.ITEM_NAME, BOM.DEPUTY_CHIP, BOM.CHIP_NAME) AS RN,
                    COUNT(*) OVER (PARTITION BY NT.ITEM_NAME) AS MAIN_CHIP_COUNT
                    FROM
                    (
                        SELECT DISTINCT(ITEM.ITEM_NAME) AS ITEM_NAME
                        FROM ITEM
                        LEFT JOIN FEATURE_GROUP FG ON FG.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                        WHERE ITEM.ITEM_CODE LIKE N'CL%' AND ITEM.FEATURE_GROUP_ID NOT IN ('16CC4C4F-7563-405C-37AF-1B442416133B','8848D890-F146-4542-98ED-1B521A585688','61A2E333-0BBB-4CE9-D47E-1B45AF842140','61A2E333-0BBB-4CE9-D47E-1B45AF842140') AND ITEM.ITEM_NAME != '待失效'
                    ) AS NT
                    LEFT JOIN 
                    (
                        SELECT 
                        MAX(CASE WHEN RowNum = 1 THEN NT.WAFER_NAME END) AS MAIN_CHIP,
                        MAX(CASE WHEN RowNum = 2 THEN NT.WAFER_NAME END) AS DEPUTY_CHIP,
                            CHIP_NAME
                        FROM
                        (
                        SELECT 
                        CASE
                        WHEN IT.ITEM_NAME='GC1808-TSSOP14-D-ZY' THEN 'GC1808D-TSSOP14-D'
                        WHEN (RIGHT(IT.ITEM_NAME,3)='-ZY' OR RIGHT(IT.ITEM_NAME,3)='-BY' OR RIGHT(IT.ITEM_NAME,3)='-ZB') THEN LEFT(IT.ITEM_NAME, LEN(IT.ITEM_NAME)-3)
                        ELSE IT.ITEM_NAME
                        END AS CHIP_NAME,
                        ROW_NUMBER() OVER(PARTITION BY IT.ITEM_CODE ORDER BY BD.Z_MAIN_CHIP) AS RowNum,ITEM.ITEM_NAME AS WAFER_NAME
                        FROM BOM_PRODUCT BP
                        LEFT JOIN ITEM IT
                        ON BP.ITEM_ID = IT.ITEM_BUSINESS_ID
                        LEFT JOIN BOM_D BD
                        ON BD.BOM_ID =BP.BOM_ID 
                        LEFT JOIN ITEM 
                        ON BD.SOURCE_ID_ROid = ITEM.ITEM_BUSINESS_ID
                        WHERE IT.STATUS != 3 AND IT.ITEM_NAME <> N'AZ1' AND IT.ITEM_CODE LIKE N'BC%AB'
                        ) AS NT
                        GROUP BY NT.CHIP_NAME
                        ) AS BOM  ON BOM.MAIN_CHIP = NT.ITEM_NAME
                    ) AS BB
                    )

                    SELECT
                    BL.ROW,
                    BL.RN,
                    BL.MAIN_CHIP_COUNT,
                    BL.MAIN_CHIP,
                    NNT.ITEM_CODE AS WAFER_CODE,
                    BL.CHIP_NAME,
                    ISNULL(SD1.TOTAL_FINISHED_GOODS, 0) AS TOTAL_FINISHED_GOODS,
                    ISNULL(SD1.TOP_FINISHED_GOODS, 0) AS TOP_FINISHED_GOODS,
                    ISNULL(SD1.BACK_FINISHED_GOODS, 0) AS BACK_FINISHED_GOODS,
                    ISNULL(SD1.TOTAL_SEMI_MANUFACTURED, 0) AS TOTAL_SEMI_MANUFACTURED,
                    ISNULL(SD1.TOP_SEMI_MANUFACTURED, 0) AS TOP_SEMI_MANUFACTURED,
                    ISNULL(SD1.BACK_SEMI_MANUFACTURED, 0) AS BACK_SEMI_MANUFACTURED,
                    ISNULL(WIP3.PACKAGE_WIP_QTY, 0) AS PACKAGE_WIP_QTY,
                    ISNULL(WIP3.PACKAGE_TOP_WIP_QTY, 0) AS PACKAGE_TOP_WIP_QTY,
                    ISNULL(WIP3.PACKAGE_BACK_WIP_QTY, 0) AS PACKAGE_BACK_WIP_QTY,
                    ISNULL(SD1.SG_QTY, 0) AS SG_QTY,
                    ISNULL(SD1.SG_FINISHED_GOODS, 0) AS SG_FINISHED_GOODS,
                    ISNULL(SD1.SG_SEMI_MANUFACTURED, 0) AS SG_SEMI_MANUFACTURED,
                    ISNULL(WIP3.SECONDARY_OUTSOURCING_WIP_QTY, 0) AS SECONDARY_OUTSOURCING_WIP_QTY,
                    ISNULL(WIP1.PURCHASE_WIP_QTY, 0) AS PURCHASE_WIP_QTY,
                    ISNULL(SD2.TOTAL_RAW_MATERIALS,0) AS TOTAL_RAW_MATERIALS,
                    ISNULL(WIP1.CP_WIP_QTY, 0) AS CP_WIP_QTY,
                    ISNULL(SD2.NO_TESTED_WAFER, 0) AS NO_TESTED_WAFER,
                    ISNULL(SD2.TESTED_WAFER, 0) AS TESTED_WAFER,
                    BL.DEPUTY_CHIP,
                    ISNULL(WIP2.PURCHASE_WIP_QTY, 0) AS OUTSOURCING_WIP_QTY,
                    ISNULL(SD3.TOTAL_RAW_MATERIALS, 0) AS TOTAL_B_RAW_MATERIALS
                    FROM BL
                    LEFT JOIN SD SD1 ON SD1.ITEM_NAME = BL.CHIP_NAME
                    LEFT JOIN SD SD2 ON SD2.ITEM_NAME = BL.MAIN_CHIP
                    LEFT JOIN SD SD3 ON SD3.ITEM_NAME = BL.DEPUTY_CHIP
                    LEFT JOIN WIP WIP1 ON WIP1.ITEM_NAME = BL.MAIN_CHIP
                    LEFT JOIN WIP WIP2 ON WIP2.ITEM_NAME = BL.DEPUTY_CHIP
                    LEFT JOIN WIP WIP3 ON WIP3.ITEM_NAME = BL.CHIP_NAME
                    INNER JOIN 
                    (
                        SELECT ITEM_NAME,ITEM_CODE
                        FROM ITEM
                        WHERE ITEM_CODE LIKE N'%WG' OR ITEM_CODE LIKE N'%WF'
                    )NNT 
                    ON NNT.ITEM_NAME = BL.MAIN_CHIP
                    ORDER BY BL.ROW
            """)
            result = db.execute(base_query).all()
            
            # 将查询结果转换为GlobalReport对象列表
            reports = []
            for row in result:
                report_data = {
                    "ROW": row.ROW,
                    "RN": row.RN,
                    "MAIN_CHIP_COUNT": row.MAIN_CHIP_COUNT,
                    "MAIN_CHIP": row.MAIN_CHIP,
                    "WAFER_CODE": row.WAFER_CODE,
                    "CHIP_NAME": row.CHIP_NAME,
                    "TOTAL_FINISHED_GOODS": row.TOTAL_FINISHED_GOODS if row.TOTAL_FINISHED_GOODS != 0 else None,
                    "TOP_FINISHED_GOODS": row.TOP_FINISHED_GOODS if row.TOP_FINISHED_GOODS != 0 else None,
                    "BACK_FINISHED_GOODS": row.BACK_FINISHED_GOODS if row.BACK_FINISHED_GOODS != 0 else None,
                    "TOTAL_SEMI_MANUFACTURED": row.TOTAL_SEMI_MANUFACTURED if row.TOTAL_SEMI_MANUFACTURED != 0 else None,
                    "TOP_SEMI_MANUFACTURED": row.TOP_SEMI_MANUFACTURED if row.TOP_SEMI_MANUFACTURED != 0 else None,
                    "BACK_SEMI_MANUFACTURED": row.BACK_SEMI_MANUFACTURED if row.BACK_SEMI_MANUFACTURED != 0 else None,
                    "PACKAGE_WIP_QTY": row.PACKAGE_WIP_QTY if row.PACKAGE_WIP_QTY != 0 else None,
                    "PACKAGE_TOP_WIP_QTY": row.PACKAGE_TOP_WIP_QTY if row.PACKAGE_TOP_WIP_QTY != 0 else None,
                    "PACKAGE_BACK_WIP_QTY": row.PACKAGE_BACK_WIP_QTY if row.PACKAGE_BACK_WIP_QTY != 0 else None,
                    "SG_QTY": row.SG_QTY if row.SG_QTY != 0 else None,
                    "SG_FINISHED_GOODS": row.SG_FINISHED_GOODS if row.SG_FINISHED_GOODS != 0 else None,
                    "SG_SEMI_MANUFACTURED": row.SG_SEMI_MANUFACTURED if row.SG_SEMI_MANUFACTURED != 0 else None,
                    "SECONDARY_OUTSOURCING_WIP_QTY": row.SECONDARY_OUTSOURCING_WIP_QTY if row.SECONDARY_OUTSOURCING_WIP_QTY != 0 else None,
                    "PURCHASE_WIP_QTY": row.PURCHASE_WIP_QTY if row.PURCHASE_WIP_QTY != 0 else None,
                    "TOTAL_RAW_MATERIALS": row.TOTAL_RAW_MATERIALS if row.TOTAL_RAW_MATERIALS != 0 else None,
                    "CP_WIP_QTY": row.CP_WIP_QTY if row.CP_WIP_QTY != 0 else None,
                    "NO_TESTED_WAFER": row.NO_TESTED_WAFER if row.NO_TESTED_WAFER != 0 else None,
                    "TESTED_WAFER": row.TESTED_WAFER if row.TESTED_WAFER != 0 else None,
                    "DEPUTY_CHIP": row.DEPUTY_CHIP,
                    "OUTSOURCING_WIP_QTY": row.OUTSOURCING_WIP_QTY if row.OUTSOURCING_WIP_QTY != 0 else None,
                    "TOTAL_B_RAW_MATERIALS": row.TOTAL_B_RAW_MATERIALS if row.TOTAL_B_RAW_MATERIALS != 0 else None
                }
                reports.append(GlobalReport(**report_data))
            
            return reports
        except Exception as e:
            logger.error(f"获取综合报表失败: {str(e)}")
            raise CustomException("获取综合报表失败")

    def export_global_report(self,db:Session)->bytes:
        """导出综合报表Excel"""
        try:
            reports = self.get_global_report(db)
            wb = Workbook()
            ws = wb.active
            ws.title = "外协报表"

            headers = [
                "行号",
                "主芯片",
                "晶圆编码",
                "芯片名称",
                "产成品数量",
                "产成品正印数量",
                "产成品背印数量",
                "半成品数量",
                "半成品正印数量",
                "半成品背印数量",
                "封装在制数量",
                "封装正印在制数量",
                "封装背印在制数量",
                "芯片总和(产半封/万颗)",
                "苏工院库存数量",
                "苏工院产成品数量",
                "苏工院半成品数量",
                "二次委外在制数量",
                "在线晶圆",
                "库存晶圆",
                "中测晶圆",
                "未测晶圆",
                "已测晶圆",
                "副芯片",
                "B芯在途数量",
                "B芯圆片数量"
            ]
            
            # 设置列宽
            column_widths = {
                'A': 15,
                'B': 15,
                'C': 25,
                'D': 35,
                'E': 10,
                'F': 10,
                'G': 10,
                'H': 10,
                'I': 10,
                'J': 10,
                'K': 10,
                'L': 10,
                'M': 10,
                'N': 10,
                'O': 10,
                'P': 10,
                'Q': 10,
                'R': 10,
                'S': 10,
                'T': 10,
                'U': 10,
                'V': 10,
                'W': 10,
                'X': 35,
                'Y': 10,
                'Z': 10
            }

            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            for row, report in enumerate(reports, 2):
                # 处理可能为None的值
                total_finished = report.TOTAL_FINISHED_GOODS or 0
                total_semi = report.TOTAL_SEMI_MANUFACTURED or 0
                package_wip = report.PACKAGE_WIP_QTY or 0
                
                data = [
                    report.ROW,
                    report.MAIN_CHIP,
                    report.WAFER_CODE,
                    report.CHIP_NAME,
                    report.TOTAL_FINISHED_GOODS,
                    report.TOP_FINISHED_GOODS,
                    report.BACK_FINISHED_GOODS,
                    report.TOTAL_SEMI_MANUFACTURED,
                    report.TOP_SEMI_MANUFACTURED,
                    report.BACK_SEMI_MANUFACTURED,
                    report.PACKAGE_WIP_QTY,
                    report.PACKAGE_TOP_WIP_QTY,
                    report.PACKAGE_BACK_WIP_QTY,
                    round((total_finished + total_semi + package_wip)/10000, 2),
                    report.SG_QTY,
                    report.SG_FINISHED_GOODS,
                    report.SG_SEMI_MANUFACTURED,
                    report.SECONDARY_OUTSOURCING_WIP_QTY,
                    report.PURCHASE_WIP_QTY,
                    report.TOTAL_RAW_MATERIALS,
                    report.CP_WIP_QTY,
                    report.NO_TESTED_WAFER,
                    report.TESTED_WAFER,
                    report.DEPUTY_CHIP,
                    report.OUTSOURCING_WIP_QTY,
                    report.TOTAL_B_RAW_MATERIALS
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border

            # 冻结第一行
            ws.freeze_panes = 'A2'

            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"导出外协报表Excel失败: {str(e)}")
            raise CustomException("导出外协报表Excel失败")
        
    def get_assy_analyze_total(self,db:Session)->List[AssyAnalyzeTotalResponse]:
        """获取封装分析总表"""
        RECEIPT_QUERY = text(f"""
            SELECT CAST(SUM(PRD.INVENTORY_QTY) AS INT) AS receipt
            FROM PURCHASE_RECEIPT_D PRD
            LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = PRD.ITEM_ID
            WHERE ITEM.ITEM_CODE LIKE N'BC-%' 
            AND 
            PRD.CreateDate BETWEEN DATEFROMPARTS(YEAR(DATEADD(MONTH, -1, GETDATE())),MONTH(DATEADD(MONTH, -1, GETDATE())),1) AND EOMONTH(DATEADD(MONTH, -1, GETDATE()))
        """)
        WIP_QUERY = text(f"""
            SELECT SUM(WIP_QTY) AS wip
            FROM 
            (
            SELECT
            CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
            CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
            CASE 
                    WHEN PO.[CLOSE] = N'2' THEN 0
                    WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                    ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
            END AS WIP_QTY,
            CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
            CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
            DATEDIFF(DAY, PO.PURCHASE_DATE, PR.CreateDate) AS leadtime
            FROM PURCHASE_ORDER PO
            LEFT JOIN PURCHASE_ORDER_D PO_D 
            ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
            LEFT JOIN PURCHASE_ORDER_SD PO_SD 
            ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
            LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
            ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
            LEFT JOIN Z_OUT_MO_D 
            ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
            LEFT JOIN ITEM 
            ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
            OUTER APPLY (
            SELECT TOP 1 *
            FROM PURCHASE_RECEIPT_D PRD
            WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
            ORDER BY PRD.CreateDate
            ) PR
            WHERE ITEM.ITEM_CODE LIKE N'BC%AB' 
            AND PO.PURCHASE_DATE >= N'2025-01-01' 
            AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
            AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
            AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            ) AS WIP
        """)
        YIELD_LEADTIME_EXCEED_QUERY = text(f"""
            SELECT 
            ROUND(CAST(AVG(YIELD) AS FLOAT)*100,2) AS yields,
            ROUND(CAST(AVG(leadtime) AS FLOAT),0) AS leadtime,
            ROUND(CAST(SUM(
                CASE
                WHEN leadtime > 30 THEN 1
                ELSE 0
                END
              )AS FLOAT) /COUNT(*)*100,2)  AS exceed
            FROM
                      (
            SELECT
            CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
            CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
            CAST(PO_D.RECEIPTED_PRICE_QTY/PO_D.BUSINESS_QTY AS FLOAT) AS YIELD,
            CASE 
                    WHEN PO.[CLOSE] = N'2' THEN 0
                    WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                    ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
            END AS WIP_QTY,
            CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
            CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
            DATEDIFF(DAY, PO.PURCHASE_DATE, PR.CreateDate) AS leadtime
            FROM PURCHASE_ORDER PO
            LEFT JOIN PURCHASE_ORDER_D PO_D 
            ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
            LEFT JOIN PURCHASE_ORDER_SD PO_SD 
            ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
            LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
            ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
            LEFT JOIN Z_OUT_MO_D 
            ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
            LEFT JOIN ITEM 
            ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
            OUTER APPLY (
            SELECT TOP 1 *
            FROM PURCHASE_RECEIPT_D PRD
            WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
            ORDER BY PRD.CreateDate
            ) PR
            WHERE ITEM.ITEM_CODE LIKE N'BC%AB' 
            AND PO.PURCHASE_DATE BETWEEN DATEFROMPARTS(YEAR(DATEADD(MONTH, -4, GETDATE())),MONTH(DATEADD(MONTH, -4, GETDATE())),1) AND EOMONTH(DATEADD(MONTH, -2, GETDATE()))
            AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
            AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
            AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            ) AS WIP
            WHERE WIP_QTY = 0 AND YIELD < 1 AND RECEIPTED_PRICE_QTY > 0
        """)
        THIS_MONTH_RECEIPT_QUERY = text(f"""
            SELECT CAST(SUM(PRD.INVENTORY_QTY) AS INT) AS receipt
            FROM PURCHASE_RECEIPT_D PRD
            LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = PRD.ITEM_ID
            WHERE ITEM.ITEM_CODE LIKE N'BC-%' 
            AND PRD.CreateDate >= DATEFROMPARTS(YEAR(GETDATE()),MONTH(GETDATE()),1)
        """)
        try:
            receipt_result = db.execute(RECEIPT_QUERY).fetchone()
            wip_result = db.execute(WIP_QUERY).fetchone()
            yield_leadtime_exceed_result = db.execute(YIELD_LEADTIME_EXCEED_QUERY).fetchone()
            this_month_receipt = db.execute(THIS_MONTH_RECEIPT_QUERY).fetchone()
            
            # 将结果转换为可序列化的字典
            result = {
                "receipt": receipt_result.receipt if receipt_result else 0,
                "wip": wip_result.wip if wip_result else 0,
                "yields": yield_leadtime_exceed_result.yields if yield_leadtime_exceed_result else 0.0,
                "leadTime": yield_leadtime_exceed_result.leadtime if yield_leadtime_exceed_result else 0,
                "exceed": yield_leadtime_exceed_result.exceed if yield_leadtime_exceed_result else 0.0,
                "this_month_receipt": this_month_receipt.receipt / 500000 if this_month_receipt and this_month_receipt.receipt is not None else 0
            }
            
            return result
        except Exception as e:
            logger.error(f"获取封装分析总表失败: {str(e)}")
            raise CustomException("获取封装分析总表失败")

    def get_assy_analyze_loading(self,db:Session,range_type:str)->List[AssyAnalyzeLoadingResponse]:
        """获取封装分析装载"""
        ASSY_ANALYZE_LOADING_QUERY = text("""
            SELECT 
            [Date],
            [SOP8(12R)] AS SOP8_12R,
            [SOP8] AS SOP8,
            [DFN8L(2X2X0.5-P0.5)] AS DFN8,
            [SOP16(12R)] AS SOP16_12R,
            [SOP16] AS SOP16,
            [SOP14(12R)] AS SOP14_12R,
            [SOP14] AS SOP14,
            [TSSOP20L] AS TSSOP20,
            [SOT26(14R)] AS SOT26_14R,
            [SOT25(20R)] AS SOT25_20R,
            [SOT25(14R)] AS SOT25_14R,
            [SSOP24] AS SSOP24,
            [ESSOP10] AS ESSOP10,
            [QFN20L(3X3X0.5-P0.4)] AS QFN20,
            [LQFP32L(7X7)] AS LQFP32
            FROM huaxinAdmin_hisemi_loading_analyze
            WHERE 1=1
        """)
        try:
            # 构造查询条件
            """
            '0': 近1个月
            '1': 近3个月
            '2': 近6个月
            '3': 近1年
            '4': 近2年
            """
            condition = ''

            if range_type == "0":
                condition = " AND Date >= DATEADD(MONTH, -1, GETDATE())"
            elif range_type == "1":
                condition = " AND Date >= DATEADD(MONTH, -3, GETDATE())"
            elif range_type == "2":
                condition = " AND Date >= DATEADD(MONTH, -6, GETDATE())"
            elif range_type == "3":
                condition = " AND Date >= DATEADD(MONTH, -12, GETDATE())"
            elif range_type == "4":
                condition = " AND Date >= DATEADD(MONTH, -24, GETDATE())"
            ASSY_ANALYZE_LOADING_QUERY = text(str(ASSY_ANALYZE_LOADING_QUERY) + condition)
            result = db.execute(ASSY_ANALYZE_LOADING_QUERY).fetchall()
            # 构造AssyAnalyzeLoadingResponse对象列表
            return [
                AssyAnalyzeLoadingResponse(
                    Date = row.Date,
                    SOP8_12R = row.SOP8_12R,
                    SOP8 = row.SOP8,
                    DFN8 = row.DFN8,
                    SOP16_12R = row.SOP16_12R,
                    SOP16 = row.SOP16,
                    SOP14_12R = row.SOP14_12R,
                    SOP14 = row.SOP14,
                    TSSOP20 = row.TSSOP20,
                    SOT26 = row.SOT26_14R,
                    SOT25_20R = row.SOT25_20R,
                    SOT25_14R = row.SOT25_14R,
                    SSOP24 = row.SSOP24,
                    ESSOP10 = row.ESSOP10,
                    QFN20 = row.QFN20,
                    LQFP32 = row.LQFP32
                ) for row in result
            ]
        except Exception as e:
            logger.error(f"获取封装分析装载失败: {str(e)}")
            raise CustomException("获取封装分析装载失败")
        
    def get_assy_year_trend(self,db:Session)->List[AssyYearTrendResponse]:
        """获取封装年趋势"""
        ASSY_YEAR_TREND_QUERY = text("""
            SELECT ROUND(SUM(BUSINESS_QTY)/1000000,2) AS qty,Z_PACKAGE_TYPE_NAME AS packageType,[YEAR] AS year
            FROM (
                SELECT
                    hpl.BUSINESS_QTY,
                    CASE
                        WHEN LEFT(hpl.Z_PACKAGE_TYPE_NAME,4)='SOP8' THEN 'SOP8'
                        WHEN LEFT(hpl.Z_PACKAGE_TYPE_NAME,4)='DFN8' THEN 'DFN8'
                        WHEN LEFT(hpl.Z_PACKAGE_TYPE_NAME,5)='SOP14' THEN 'SOP14'
                        WHEN LEFT(hpl.Z_PACKAGE_TYPE_NAME,5)='SOP16' THEN 'SOP16'
                        WHEN LEFT(hpl.Z_PACKAGE_TYPE_NAME,7)='TSSOP20' THEN 'TSSOP20'
                    ELSE hpl.Z_PACKAGE_TYPE_NAME
                    END AS Z_PACKAGE_TYPE_NAME,
                    YEAR(hpl.PURCHASE_DATE) AS [YEAR]
                FROM HSUN_PACKAGE_LIST hpl 
                UNION ALL
                SELECT
                    PO_D.BUSINESS_QTY AS BUSINESS_QTY,
                    CASE
                        WHEN LEFT(ITEM.UDF025,4)='SOP8' THEN 'SOP8'
                        WHEN LEFT(ITEM.UDF025,4)='DFN8' THEN 'DFN8'
                        WHEN LEFT(ITEM.UDF025,5)='SOP14' THEN 'SOP14'
                        WHEN LEFT(ITEM.UDF025,5)='SOP16' THEN 'SOP16'
                        WHEN LEFT(ITEM.UDF025,7)='TSSOP20' THEN 'TSSOP20'
                    ELSE ITEM.UDF025
                    END AS Z_PACKAGE_TYPE_NAME,
                    YEAR(CAST(PO.PURCHASE_DATE AS DATE)) AS [YEAR]
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D 
                    ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                LEFT JOIN ITEM 
                    ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                WHERE PO.PURCHASE_TYPE = 2 
                    AND PO.PURCHASE_DATE > '2024-10-21' 
                    AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                    AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            ) AS CombinedResults
            WHERE YEAR <= YEAR(GETDATE())-1
            AND (Z_PACKAGE_TYPE_NAME = N'SOP8' OR Z_PACKAGE_TYPE_NAME = N'SOP14' OR Z_PACKAGE_TYPE_NAME = N'SOP16' OR Z_PACKAGE_TYPE_NAME = N'DFN8' OR Z_PACKAGE_TYPE_NAME = N'TSSOP20')
            GROUP BY Z_PACKAGE_TYPE_NAME,[YEAR]
            ORDER BY [YEAR]
        """)

        try:
            result = db.execute(ASSY_YEAR_TREND_QUERY).fetchall()
            return [AssyYearTrendResponse(qty=row.qty, packageType=row.packageType, year=row.year) for row in result]
        except Exception as e:
            logger.error(f"获取封装年趋势失败: {str(e)}")
            raise CustomException("获取封装年趋势失败")

    def get_assy_supply_analyze(self,db:Session)->List[AssySupplyAnalyzeResponse]:
        """获取封装供应分析"""
        ASSY_SUPPLY_ANALYZE_QUERY = text("""
            WITH SMR AS(
            SELECT
                ITEM.ITEM_CODE,
                ITEM.UDF025 AS packageType,
                CAST(PO_D.BUSINESS_QTY AS INT) AS orderQty,
                CAST(PO.PURCHASE_DATE AS DATE) AS purchaseDate,
                YEAR(PO.PURCHASE_DATE) AS year,
                MONTH(PO.PURCHASE_DATE) AS month,
                PO.SUPPLIER_FULL_NAME AS supply
            FROM PURCHASE_ORDER PO
            LEFT JOIN PURCHASE_ORDER_D PO_D 
                ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
            LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
            LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
            LEFT JOIN Z_OUT_MO_D 
                ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
            LEFT JOIN ITEM 
                ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
            WHERE PO.PURCHASE_TYPE = 2 
                AND PO.PURCHASE_DATE > '2024-12-31' AND PO.PURCHASE_DATE > DATEADD(YEAR, -1, GETDATE())
                AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            )

            SELECT
                supply AS Supplier,
                COUNT(*) AS DataRowCount,
                SUM(orderQty) AS TotalOrderQty,
                COUNT(DISTINCT ITEM_CODE) AS PackageTypeCount
            FROM SMR
            GROUP BY supply
        """)
        try:
            result = db.execute(ASSY_SUPPLY_ANALYZE_QUERY).fetchall()
            return [AssySupplyAnalyzeResponse(Supplier=row.Supplier, DataRowCount=row.DataRowCount, TotalOrderQty=row.TotalOrderQty, PackageTypeCount=row.PackageTypeCount) for row in result]
        except Exception as e:
            logger.error(f"获取封装供应分析失败: {str(e)}")
            raise CustomException("获取封装供应分析失败")

    def get_sop_analyze(self,db:Session)->List[SopAnalyzeResponse]:
        """获取SOP分析"""
        SOP_ANALYZE_QUERY = text("""
                WITH 
                SALES AS 
                (
                SELECT 
                ABTR,
                ITEM_NAME,
                SUM(PRICE_QTY) AS PRICE_QTY
                FROM (
                SELECT
                    CASE
                    WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                    ELSE '管装'
                    END AS ABTR,
                    dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB') AS ITEM_NAME,
                    SID.PRICE_QTY
                FROM SALES_ISSUE SI
                LEFT JOIN SALES_ISSUE_D SID
                ON SID.SALES_ISSUE_ID = SI.SALES_ISSUE_ID
                LEFT JOIN ITEM
                ON SID.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                WHERE SI.DOC_DATE BETWEEN DATEFROMPARTS(YEAR(DATEADD(MONTH, -1, GETDATE())),MONTH(DATEADD(MONTH, -1, GETDATE())),1) AND EOMONTH(DATEADD(MONTH, -1, GETDATE()))
                UNION ALL
                SELECT
                    CASE
                    WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                    ELSE '管装'
                    END AS ABTR,
                    dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB') AS ITEM_NAME,
                    SRRD.PRICE_QTY*-1
                FROM SALES_RETURN_RECEIPT SRR
                LEFT JOIN SALES_RETURN_RECEIPT_D SRRD
                ON SRRD.SALES_RETURN_RECEIPT_ID = SRR.SALES_RETURN_RECEIPT_ID
                LEFT JOIN ITEM
                ON SRRD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                WHERE SRR.DOC_DATE BETWEEN DATEFROMPARTS(YEAR(DATEADD(MONTH, -1, GETDATE())),MONTH(DATEADD(MONTH, -1, GETDATE())),1) AND EOMONTH(DATEADD(MONTH, -1, GETDATE()))
                ) AS NT
                GROUP BY ABTR,ITEM_NAME
                ),
                SAFE_STOCK AS 
                (
                SELECT 
                CASE
                WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                ELSE '管装'
                END AS ABTR,
                CASE
                WHEN RIGHT(ITEM.ITEM_NAME,3) = '-BY' OR RIGHT(ITEM.ITEM_NAME,3) = '-ZY' THEN ITEM.ITEM_NAME
                ELSE (
                    CASE 
                    WHEN ITEM.ITEM_NAME = 'HS6601MX-16H-SOP8-H' THEN 'HS6601MX-SOP8-H-BY'
                    WHEN ITEM.ITEM_NAME = 'HS6601L-25 可重复触发 LDO:2.5V' THEN 'HS6601L-25-SOP8-A-ZY'
                    WHEN ITEM.ITEM_CODE LIKE '%-ZY-%' THEN ITEM.ITEM_NAME + '-ZY'
                    WHEN ITEM.ITEM_CODE LIKE '%-BY-%' THEN ITEM.ITEM_NAME + '-BY'
                    END
                )
                END AS ITEM_NAME,
                CAST(SUM(IW.SAFE_STOCK) AS INT) AS SAFE_STOCK
                FROM ITEM_WAREHOUSE IW
                LEFT JOIN ITEM ON IW.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                WHERE SAFE_STOCK > 0
                GROUP BY 
                CASE
                    WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                    ELSE '管装'
                END,
                CASE
                    WHEN RIGHT(ITEM.ITEM_NAME,3) = '-BY' OR RIGHT(ITEM.ITEM_NAME,3) = '-ZY' THEN ITEM.ITEM_NAME
                    ELSE (
                    CASE 
                        WHEN ITEM.ITEM_NAME = 'HS6601MX-16H-SOP8-H' THEN 'HS6601MX-SOP8-H-BY'
                        WHEN ITEM.ITEM_NAME = 'HS6601L-25 可重复触发 LDO:2.5V' THEN 'HS6601L-25-SOP8-A-ZY'
                        WHEN ITEM.ITEM_CODE LIKE '%-ZY-%' THEN ITEM.ITEM_NAME + '-ZY'
                        WHEN ITEM.ITEM_CODE LIKE '%-BY-%' THEN ITEM.ITEM_NAME + '-BY'
                    END
                    )
                END
                ),
                STOCK AS 
                (
                SELECT 
                CASE
                    WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                    ELSE '管装'
                END AS ABTR,
                CASE
                    WHEN ITEM.ITEM_CODE LIKE 'CP-%' THEN '产成品'
                    WHEN ITEM.ITEM_CODE LIKE 'DG-%' THEN '代工品'
                    ELSE '半成品'
                END AS CPBC,
                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'BC-,CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB') AS ITEM_NAME,
                SUM(CAST(A.INVENTORY_QTY AS INT)) AS INVENTORY_QTY
                FROM Z_WF_IC_WAREHOUSE_BIN A
                LEFT JOIN ITEM
                    ON ITEM.ITEM_BUSINESS_ID = A.ITEM_ID
                LEFT JOIN WAREHOUSE W
                    ON A.WAREHOUSE_ID = W.WAREHOUSE_ID
                WHERE 
                A.INVENTORY_QTY > 0  AND 
                (W.WAREHOUSE_NAME != '实验仓-TH（产成品）') AND 
                (W.WAREHOUSE_NAME NOT LIKE 'HOLD仓%') AND 
                (W.WAREHOUSE_NAME NOT LIKE '华新源%') AND 
                ITEM.ITEM_CODE NOT LIKE 'CL-%'
                GROUP BY
                CASE
                    WHEN ITEM.ITEM_CODE LIKE '%TR' THEN '编带'
                    ELSE '管装'
                END,
                CASE
                    WHEN ITEM.ITEM_CODE LIKE 'CP-%' THEN '产成品'
                    WHEN ITEM.ITEM_CODE LIKE 'DG-%' THEN '代工品'
                    ELSE '半成品'
                END,
                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'BC-,CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB')
                ),
                WIP AS 
                (
                SELECT 
                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'BC-,CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB') AS ITEM_NAME,
                CASE
                WHEN ZPP.Z_PROCESSING_PURPOSE_NAME LIKE '%编带' THEN '编带'
                ELSE '管装'
                END AS ABTR,
                SUM(CAST((PO_D.BUSINESS_QTY * 0.9989 - PO_D.RECEIPTED_PRICE_QTY - ISNULL(wip.[仓库库存],0)) AS INT)) AS WIP_QTY_WITHOUT_STOCK,
                SUM(ISNULL(wip.[仓库库存],0)) AS ASSY_STOCK
                FROM PURCHASE_ORDER PO
                LEFT JOIN huaxinAdmin_wip_assy wip ON wip.[订单号]=PO.DOC_NO
                LEFT JOIN PURCHASE_ORDER_D PO_D ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D ZOMD ON PO_SSD.REFERENCE_SOURCE_ID_ROid = ZOMD.Z_OUT_MO_D_ID
                LEFT JOIN ITEM ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT ON ZOMD.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC ON ZOMD.Z_PACKAGE_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_PACKAGE ON ZAC.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                LEFT JOIN Z_PROCESSING_PURPOSE ZPP ON ZAC.Z_PROCESSING_PURPOSE_ID = ZPP.Z_PROCESSING_PURPOSE_ID
                WHERE 1=1 AND 
                PO.[CLOSE]=0 AND 
                PO.SUPPLIER_FULL_NAME<>'苏州荐恒电子科技有限公司' AND 
                (PO.DOC_NO NOT LIKE '3501-%' AND 
                PO.DOC_NO != 'HX-20240430001') AND 
                ITEM.ITEM_CODE LIKE N'BC%AB' AND
                (PO_D.BUSINESS_QTY * 0.9989- PO_D.RECEIPTED_PRICE_QTY)>5000
                GROUP BY
                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'BC-,CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB'),
                CASE
                WHEN ZPP.Z_PROCESSING_PURPOSE_NAME LIKE '%编带' THEN '编带'
                ELSE '管装'
                END
                )


                SELECT 
                ROW_NUMBER() OVER (ORDER BY SS.ITEM_NAME, SS.ABTR) + 100 AS ID,
                SS.ITEM_NAME,
                SS.ABTR,
                SS.SAFE_STOCK,
                ISNULL(CAST(SALES.PRICE_QTY AS INT),0) AS LAST_MONTH_SALE,
                ISNULL(S1.INVENTORY_QTY,0) AS CP_QTY,
                ISNULL(S2.INVENTORY_QTY,0) AS BC_QTY,
                ISNULL(WIP.WIP_QTY_WITHOUT_STOCK,0) AS WIP_QTY_WITHOUT_STOCK,
                ISNULL(WIP.ASSY_STOCK,0) AS ASSY_STOCK,
                (ISNULL(S1.INVENTORY_QTY,0) + ISNULL(S2.INVENTORY_QTY,0) + ISNULL(WIP.WIP_QTY_WITHOUT_STOCK,0) + ISNULL(WIP.ASSY_STOCK,0)) AS TOTAL_STOCK,
                CAST((ISNULL(S1.INVENTORY_QTY,0) + ISNULL(S2.INVENTORY_QTY,0) + ISNULL(WIP.WIP_QTY_WITHOUT_STOCK,0) + ISNULL(WIP.ASSY_STOCK,0) - SS.SAFE_STOCK * 1.5 ) AS INT)  AS                INVENTORY_GAP,
                CAST(SUM(
                    (ISNULL(S1.INVENTORY_QTY, 0) + ISNULL(S2.INVENTORY_QTY, 0) + ISNULL(WIP.WIP_QTY_WITHOUT_STOCK, 0) + ISNULL(WIP.ASSY_STOCK, 0) - SS.SAFE_STOCK * 1.5)
                ) OVER (PARTITION BY SS.ITEM_NAME) AS INT)  AS INVENTORY_GAP_TOTAL
                FROM SAFE_STOCK SS
                LEFT JOIN SALES ON (REPLACE(SS.ITEM_NAME, '_', '-') = SALES.ITEM_NAME AND SS.ABTR = SALES.ABTR)
                LEFT JOIN STOCK S1 ON (S1.ITEM_NAME = REPLACE(SS.ITEM_NAME, '_', '-') AND S1.ABTR = SS.ABTR AND S1.CPBC = '产成品')
                LEFT JOIN STOCK S2 ON (S2.ITEM_NAME = REPLACE(SS.ITEM_NAME, '_', '-') AND S2.ABTR = SS.ABTR AND S2.CPBC = '半成品')
                LEFT JOIN WIP ON (WIP.ITEM_NAME = SS.ITEM_NAME AND WIP.ABTR = SS.ABTR)
                ORDER BY SS.ITEM_NAME,SS.ABTR 
        """)

        try:
            result = db.execute(SOP_ANALYZE_QUERY).fetchall()
            return [SopAnalyzeResponse(ID=row.ID, ITEM_NAME=row.ITEM_NAME, ABTR=row.ABTR, SAFE_STOCK=row.SAFE_STOCK, LAST_MONTH_SALE=row.LAST_MONTH_SALE, CP_QTY=row.CP_QTY, BC_QTY=row.BC_QTY, WIP_QTY_WITHOUT_STOCK=row.WIP_QTY_WITHOUT_STOCK, ASSY_STOCK=row.ASSY_STOCK, TOTAL_STOCK=row.TOTAL_STOCK, INVENTORY_GAP=row.INVENTORY_GAP,INVENTORY_GAP_TOTAL=row.INVENTORY_GAP_TOTAL) for row in result]
        except Exception as e:
            logger.error(f"获取SOP分析失败: {str(e)}")
            raise CustomException("获取SOP分析失败")

    def export_sop_report(self,db:Session)->bytes:
        """导出SOP报表"""
        try:
            results = self.get_sop_analyze(db)
            wb = Workbook()
            ws = wb.active
            ws.title = "SOP报表"
            headers = [
                "ID",
                "品名",
                "管装/编带",
                "安全库存值",
                "上月销售量",
                "产成品库存",
                "半成品库存",
                "封装数量",
                "封装厂库存",
                "总库存",
                "库存缺口",
                "缺口合计"
            ]

            # 设置列宽
            column_widths = {
                "A": 10,
                "B": 40,
                "C": 20,
                "D": 20,
                "E": 20,
                "F": 20,
                "G": 20,
                "H": 20,
                "I": 20,
                "J": 20,
                "K": 20,
                "L": 20
            }

            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            for row, result in enumerate(results, 2):
                data = [
                    result.ID,
                    result.ITEM_NAME,
                    result.ABTR,
                    result.SAFE_STOCK,
                    result.LAST_MONTH_SALE,
                    result.CP_QTY,
                    result.BC_QTY,
                    result.WIP_QTY_WITHOUT_STOCK,
                    result.ASSY_STOCK,
                    result.TOTAL_STOCK,
                    result.INVENTORY_GAP,
                    result.INVENTORY_GAP_TOTAL,
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border

            # 冻结第一行
            ws.freeze_panes = 'A2'

            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"导出SOP报表Excel失败: {str(e)}")
            raise CustomException("导出SOP报表Excel失败")

    def get_item_wafer_info(self,db:Session,item_name:str)->List[ItemWaferInfoResponse]:
        """获取晶圆信息"""
        try:
            # 构建基础 SQL 语句
            sql_query = """
            WITH 
            BOM AS
            (
                SELECT 
                MAX(CASE WHEN RowNum = 1 THEN NT.WAFER_NAME END) AS MAIN_CHIP,
                MAX(CASE WHEN RowNum = 2 THEN NT.WAFER_NAME END) AS DEPUTY_CHIP,
                        CHIP_NAME
                FROM
                (
                SELECT 
                CASE
                    WHEN IT.ITEM_NAME='GC1808-TSSOP14-D-ZY' THEN 'GC1808D-TSSOP14-D-ZY'
                    ELSE IT.ITEM_NAME
                END AS CHIP_NAME,
                ROW_NUMBER() OVER(PARTITION BY IT.ITEM_CODE ORDER BY BD.Z_MAIN_CHIP) AS RowNum,ITEM.ITEM_NAME AS WAFER_NAME
                FROM BOM_PRODUCT BP
                LEFT JOIN ITEM IT
                ON BP.ITEM_ID = IT.ITEM_BUSINESS_ID
                LEFT JOIN BOM_D BD
                ON BD.BOM_ID =BP.BOM_ID 
                LEFT JOIN ITEM 
                ON BD.SOURCE_ID_ROid = ITEM.ITEM_BUSINESS_ID
                WHERE IT.STATUS != 3 AND IT.ITEM_NAME <> N'AZ1' AND IT.ITEM_CODE LIKE N'BC%AB'
                ) AS NT
                GROUP BY NT.CHIP_NAME
            ),
            GD AS
            (
                SELECT ITEM_NAME,Z_GROSS_DIE
                FROM ITEM
                WHERE ITEM_CODE LIKE N'CL%WF' AND Z_GROSS_DIE>0
            )

            SELECT BOM.CHIP_NAME,BOM.MAIN_CHIP,GD1.Z_GROSS_DIE AS MAIN_CHIP_GROSS_DIE,BOM.DEPUTY_CHIP,GD2.Z_GROSS_DIE AS DEPUTY_CHIP_GROSS_DIE
            FROM BOM
            LEFT JOIN GD GD1 ON GD1.ITEM_NAME = BOM.MAIN_CHIP
            LEFT JOIN GD GD2 ON GD2.ITEM_NAME = BOM.DEPUTY_CHIP
            """
            
            # 添加条件
            if item_name:
                sql_query += f" WHERE BOM.CHIP_NAME LIKE '%{item_name}%'"
            
            # 执行查询
            result = db.execute(text(sql_query)).fetchall()
            
            # 转换结果为响应对象
            responses = []
            for row in result:
                responses.append(ItemWaferInfoResponse(
                    CHIP_NAME=row.CHIP_NAME, 
                    MAIN_CHIP=row.MAIN_CHIP, 
                    MAIN_CHIP_GROSS_DIE=row.MAIN_CHIP_GROSS_DIE, 
                    DEPUTY_CHIP=row.DEPUTY_CHIP, 
                    DEPUTY_CHIP_GROSS_DIE=row.DEPUTY_CHIP_GROSS_DIE
                ))
            
            return responses
        except Exception as e:
            logger.error(f"获取晶圆信息失败: {str(e)}")
            raise CustomException("获取晶圆信息失败")

    def get_sales(self,db:Session)->List[SalesResponse]:
        """获取销售员名称"""
        SALES_QUERY = text("""
            SELECT USER_NAME
            FROM [USER]
            WHERE REMARK = '销售' OR REMARK = '销售部长' OR REMARK = '销售组长' OR REMARK = '销售助理' 
        """)
        try:
            result = db.execute(SALES_QUERY).fetchall()
            return [SalesResponse(label=row.USER_NAME, value=row.USER_NAME) for row in result]
        except Exception as e:
            logger.error(f"获取销售员名称失败: {str(e)}")
            raise CustomException("获取销售员名称失败")
    
    def get_assy_require_orders(self, db: Session, params: AssyRequireOrdersQuery) -> Dict[str, Any]:
        try:
            # 基础查询
            base_query = """
                SELECT 
                    ASSY_REQUIREMENTS_ID,
                    ITEM_NAME,
                    ITEM_CODE,
                    ABTR,
                    BUSINESS_QTY,
                    REQUIREMENT_TYPE,
                    EMERGENCY,
                    SALES,
                    REMARK,
                    CHIP_A,
                    CHIP_A_QTY,
                    CHIP_B,
                    CHIP_B_QTY,
                    STATUS,
                    CreateDate,
                    CreateBy
                FROM huaxinAdmin_Requirement_DOC
                WHERE 1=1
            """
            
            # 构建查询条件
            conditions = []
            query_params = {}

            # 参数验证和清理
            if params.assy_requirements_id:
                conditions.append("AND ASSY_REQUIREMENTS_ID = :assy_requirements_id")
                query_params["assy_requirements_id"] = self._clean_input(params.assy_requirements_id)

            if params.itemName:
                conditions.append("AND UPPER(ITEM_NAME) LIKE UPPER(:item_name)")
                query_params["item_name"] = f"%{self._clean_input(params.itemName)}%"

            if params.requirementType:
                conditions.append("AND REQUIREMENT_TYPE = :requirement_type")
                query_params["requirement_type"] = self._clean_input(params.requirementType)

            if params.abtr:
                conditions.append("AND ABTR = :abtr")
                query_params["abtr"] = self._clean_input(params.abtr)

            if params.status:
                conditions.append("AND STATUS = :status")
                query_params["status"] = self._clean_input(params.status)
                
            if params.sales:
                conditions.append("AND SALES = :sales")
                query_params["sales"] = self._clean_input(params.sales)

            if params.order_date_start:
                conditions.append("AND CAST(CreateDate AS DATE) >= :order_date_start")
                query_params["order_date_start"] = params.order_date_start
            
            if params.order_date_end:
                conditions.append("AND CAST(CreateDate AS DATE) <= :order_date_end")
                query_params["order_date_end"] = params.order_date_end
            
            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加排序
            query += " ORDER BY CreateDate"

            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).fetchall()
            
            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1) as total_count
                FROM huaxinAdmin_Requirement_DOC
                WHERE 1=1
                {' '.join(conditions)}
            """
            total = db.execute(
                text(count_query).bindparams(
                    **{k: v for k, v in query_params.items() if k not in ['offset', 'pageSize']}
                )
            ).scalar_one()
            
            # 转换为响应对象
            assy_orders = [
                AssyRequireOrdersList(
                    ASSY_REQUIREMENTS_ID=row.ASSY_REQUIREMENTS_ID,
                    ITEM_NAME=row.ITEM_NAME,
                    ITEM_CODE=row.ITEM_CODE,
                    ABTR=row.ABTR,
                    BUSINESS_QTY=row.BUSINESS_QTY,
                    REQUIREMENT_TYPE=row.REQUIREMENT_TYPE,
                    EMERGENCY=row.EMERGENCY,
                    SALES=row.SALES,
                    REMARK=row.REMARK,
                    CHIP_A=row.CHIP_A,
                    CHIP_A_QTY=row.CHIP_A_QTY,
                    CHIP_B=row.CHIP_B,
                    CHIP_B_QTY=row.CHIP_B_QTY,
                    STATUS=row.STATUS,
                    CreateDate=row.CreateDate.date() if row.CreateDate else None,
                    CreateBy=row.CreateBy
                ) for row in result
            ]
            
            return {
                "list": assy_orders,
                "total": total
            }
            
        except Exception as e:
            logger.error(f"查询封装需求订单失败: {str(e)}")
            raise CustomException("查询封装需求订单失败")
                
    def cancel_assy_require_orders(self, db: Session, data: AssyRequireOrdersCancel) -> AssySubmitOrdersResponse:
        try:
            if not data.id:
                raise CustomException("订单ID不能为空")
                
            query = """
                UPDATE huaxinAdmin_Requirement_DOC
                SET STATUS = '2'
                WHERE ASSY_REQUIREMENTS_ID = :assy_requirements_id
            """
            
            result = db.execute(
                text(query).bindparams(assy_requirements_id=self._clean_input(data.id))
            )
            
            if result.rowcount == 0:
                raise CustomException("未找到对应的订单记录")
                
            db.commit()
            
            return AssySubmitOrdersResponse(
                message="作废成功！",
                success=True
            )
            
        except CustomException as e:
            db.rollback()
            raise e
        except Exception as e:
            db.rollback()
            logger.error(f"作废失败: {str(e)}")
            raise CustomException("作废失败")

    def delete_assy_require_orders(self, db: Session, data: AssyRequireOrdersCancel) -> AssySubmitOrdersResponse:
        try:
            if not data.id:
                raise CustomException("订单ID不能为空")
                
            query = """
                DELETE FROM huaxinAdmin_Requirement_DOC
                WHERE STATUS = '0' AND ASSY_REQUIREMENTS_ID = :assy_requirements_id
            """
            
            result = db.execute(
                text(query).bindparams(assy_requirements_id=self._clean_input(data.id))
            )
            
            if result.rowcount == 0:
                raise CustomException("未找到对应的订单记录或订单非待处理状态！")
                
            db.commit()
            
            return AssySubmitOrdersResponse(
                message="删除成功！",
                success=True
            )
            
        except CustomException as e:
            db.rollback()
            raise e
        except Exception as e:
            db.rollback()
            logger.error(f"删除失败: {str(e)}")
            raise CustomException("删除失败")

    def batch_submit_assy_orders(self,db:Session,data:AssySubmitOrdersRequest,current_user:str)->AssySubmitOrdersResponse:
        """批量提交封装单"""
        try:
            # 记录日志
            logger.info(f"开始处理批量封装单提交，共 {len(data.orders)} 条数据")
            
            # 对每个订单进行插入操作
            for order in data.orders:
                # 将业务数量乘以10000
                actual_qty = order.businessQty
                
                # 芯片用量保持原样
                chip_a_qty = order.mainChipUsage if order.mainChipUsage is not None else None
                chip_b_qty = order.deputyChipUsage if order.deputyChipUsage is not None else None
                
                # 构建参数化查询
                insert_sql = text("""
                INSERT INTO huaxinAdmin_Requirement_DOC (
                    ASSY_REQUIREMENTS_ID, ITEM_NAME, ITEM_CODE, ABTR, BUSINESS_QTY, 
                    REQUIREMENT_TYPE, EMERGENCY, SALES, REMARK, CHIP_A, CHIP_A_QTY, 
                    CHIP_B, CHIP_B_QTY, CreateBy, CreateDate, UpdateDate, STATUS
                ) VALUES (
                    NEWID(), :item_name, :item_code, :abtr, :business_qty,
                    :requirement_type, :emergency, :sales, :remark, :chip_a, :chip_a_qty,
                    :chip_b, :chip_b_qty, :create_by, SYSDATETIME(), SYSDATETIME(), :status
                )
                """)
                
                # 执行插入
                db.execute(
                    insert_sql,
                    {
                        "item_name": order.itemName,
                        "item_code": order.itemCode,
                        "abtr": order.abtr,
                        "business_qty": actual_qty,
                        "requirement_type": order.requirementType,
                        "emergency": order.emergency,
                        "sales": order.sales,
                        "remark": order.remark,
                        "chip_a": order.mainChip,
                        "chip_a_qty": chip_a_qty,
                        "chip_b": order.deputyChip,
                        "chip_b_qty": chip_b_qty,
                        "create_by": current_user,
                        "status": order.status
                    }
                )
            
            # 提交事务
            db.commit()
            
            # 返回成功响应
            return AssySubmitOrdersResponse(
                message=f"成功提交{len(data.orders)}个封装单",
                success=True
            )
            
        except Exception as e:
            # 发生异常时回滚事务
            db.rollback()
            logger.error(f"批量提交封装单失败: {str(e)}")
            raise CustomException(f"批量提交封装单失败: {str(e)}")

    def export_assy_orders(self,db:Session)->bytes:
        """导出封装单"""
        try:
            # 准备数据
            params = AssyRequireOrdersQuery(status='0')
            result = self.get_assy_require_orders(db,params)
            orders = result['list']
            
            wb = Workbook()
            ws = wb.active
            ws.title = "封装单"
            
            # 设置表头
            headers = [
                "品名", "品号", "管装/编带", "需求数量", "需求类型", 
                "紧急程度", "销售员", "备注", "A芯片", "A芯片用量", 
                "B芯片", "B芯片用量"
            ]
            
            # 设置列宽
            column_widths = {
                "A": 30,  # 品名
                "B": 40,  # 品号
                "C": 15,  # 管装/编带
                "D": 15,  # 需求数量
                "E": 15,  # 需求类型
                "F": 15,  # 紧急程度
                "G": 15,  # 销售员
                "H": 40,  # 备注
                "I": 30,  # A芯片
                "J": 15,  # A芯片用量
                "K": 30,  # B芯片
                "L": 15,  # B芯片用量
            }
            
            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths.get(get_column_letter(col), 15)
            
            # 写入数据
            for row, order in enumerate(orders, 2):
                data = [
                    order.ITEM_NAME if order.ITEM_NAME else '',
                    order.ITEM_CODE if order.ITEM_CODE else '',
                    order.ABTR if order.ABTR else '',
                    order.BUSINESS_QTY if order.BUSINESS_QTY else '',
                    order.REQUIREMENT_TYPE if order.REQUIREMENT_TYPE else '',
                    order.EMERGENCY if order.EMERGENCY else '',
                    order.SALES if order.SALES else '',
                    order.REMARK if order.REMARK else '',
                    order.CHIP_A if order.CHIP_A else '',
                    order.CHIP_A_QTY if order.CHIP_A_QTY else '',
                    order.CHIP_B if order.CHIP_B else '',
                    order.CHIP_B_QTY if order.CHIP_B_QTY else '',
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border
            
            # 冻结第一行
            ws.freeze_panes = 'A2'
            
            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"导出封装单Excel失败: {str(e)}")
            raise CustomException(f"导出封装单Excel失败: {str(e)}")

    def change_assy_order_status(self,db:Session)->AssySubmitOrdersResponse:
        try:
            query = """
                UPDATE huaxinAdmin_Requirement_DOC
                SET STATUS = '1'
                WHERE STATUS = '0'
            """

            result = db.execute(text(query))
            if result.rowcount == 0:
                raise CustomException("未找到对应的订单记录")
                
            db.commit()
            
            return AssySubmitOrdersResponse(
                message="修改成功！",
                success=True
            )
            
        except CustomException as e:
            db.rollback()
            raise e
        except Exception as e:
            db.rollback()
            logger.error("提交失败: {str(e)}")
            raise CustomException("提交失败")

    def get_cptest_orders_by_params(self,db:Session,params:CpTestOrdersQuery)->Dict[str,Any]:
        """获取CP测试单
        
        Args:
            db: 数据库会话
            params: 查询参数
            
        Returns:
            Dict[str,Any]: 包含数据和总数的字典
            
        Raises:
            CustomException: 当查询失败时抛出
        """
        try:
            # 构建基础查询
            base_query = """
                SELECT 
                ROW_NUMBER() OVER (ORDER BY DOC_DATE,DOC_NO) + 100000 AS ID,
                DOC_NO,
                ITEM_CODE,
                ITEM_NAME,
                LOT_NAME,
                CAST(BUSINESS_QTY AS INT) AS BUSINESS_QTY,
                CAST(RECEIPT_QTY AS INT) AS RECEIPT_QTY,
                CAST(WIP_QTY AS INT) AS WIP_QTY,
                PROGRESS_NAME,
                TESTING_PROGRAM_NAME,
                REMARK,
                CAST(DOC_DATE AS DATE) AS DOC_DATE,
                CAST(FIRST_ARRIVAL_DATE AS DATE) AS FIRST_ARRIVAL_DATE,
                SUPPLIER,
                STATUS
                FROM (
                    SELECT 
                        HC.DOC_NO,
                        LEFT(HC.ITEM_NAME,6) AS ITEM_CODE,
                        HC.ITEM_NAME,
                        HC.LOT_NAME,
                        HC.BUSINESS_QTY,
                        HC.RECEIPT_QTY,
                        0 AS WIP_QTY,
                        HC.PROGRESS_NAME,
                        HC.TESTING_PROGRAM_NAME,
                        HC.REMARK,
                        HC.DOC_DATE,
                        HC.FIRST_ARRIVAL_DATE,
                        HC.SUPPLIER,
                        HC.STATUS
                    FROM huaxinAdmin_cptest HC
                    UNION ALL
                    SELECT
                            PO.DOC_NO,
                            ITEM.ITEM_CODE,
                            ITEM.ITEM_NAME,
                            ITEM_LOT.LOT_CODE,
                            PO_D.BUSINESS_QTY,
                            PO_D.RECEIPTED_PRICE_QTY,
                            PO_D.BUSINESS_QTY - PO_D.RECEIPTED_PRICE_QTY,
                            ZAC.CUSTOM_FIELD10 AS PROGRESS_NAME,
                            ZTP.Z_TESTING_PROGRAM_NAME,
                            Z_PACKAGE.REMARK,
                            PO.PURCHASE_DATE,
                            PR.CreateDate,
                            PO.SUPPLIER_FULL_NAME,
                            PO_SD.RECEIPT_CLOSE
                    FROM PURCHASE_ORDER PO
                    LEFT JOIN PURCHASE_ORDER_D PO_D 
                            ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                    LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                            ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                    LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                            ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                    LEFT JOIN Z_OUT_MO_D 
                            ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                    LEFT JOIN ITEM 
                            ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                    LEFT JOIN ITEM_LOT 
                            ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                    LEFT JOIN Z_ASSEMBLY_CODE 
                            ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_ASSEMBLY_CODE ZAC
                            ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_TESTING_PROGRAM ZTP
                            ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                    LEFT JOIN Z_PACKAGE 
                            ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                    OUTER APPLY (
                            SELECT TOP 1 *
                            FROM PURCHASE_RECEIPT_D PRD
                            WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                            ORDER BY PRD.CreateDate
                    ) PR
                    WHERE PO.PURCHASE_TYPE = 2 
                            AND PO.PURCHASE_DATE > '2024-11-13' 
                            AND ITEM.ITEM_CODE LIKE N'CL%CP' 
                            AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                            AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                            AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                    UNION ALL
                    SELECT 
                    MO.DOC_NO,
                    ITEM.ITEM_CODE,
                    ITEM.ITEM_NAME,
                    IL.LOT_CODE,
                    MO_D.REQUIRED_SECOND_QTY,
                    MP.COMPLETED_SECOND_QTY,
                    MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY,
                    NULL AS PROGRESS_NAME,
                    NULL AS Z_TESTING_PROGRAM_NAME,
                    NULL AS REMARK,
                    MO.DOC_DATE AS PURCHASE_DATE,
                    NULL AS FIRST_ARRIVAL_DATE,
                    '苏工院' AS SUPPLIER,
                    CASE
                      WHEN (MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY) > 0 THEN '0'
                      ELSE '2'
                    END AS RECEIPT_CLOSE
                    FROM MO
                    LEFT JOIN MO_D ON MO.MO_ID = MO_D.MO_ID
                    LEFT JOIN MO_PRODUCT MP ON MP.MO_ID = MO.MO_ID
                    LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = MO.ITEM_ID
                    LEFT JOIN ITEM_LOT IL ON IL.ITEM_LOT_ID = MO_D.ITEM_LOT_ID
                    WHERE ITEM.ITEM_CODE LIKE N'CL%CP' AND MO.DOC_NO LIKE '5105%'
                ) AS CombinedResults
                WHERE 1=1
            """

            # 构建查询条件
            conditions = []
            query_params = {}
            
            # 参数验证和清理
            if params.item_code and isinstance(params.item_code, str):
                conditions.append("AND UPPER(ITEM_CODE) LIKE UPPER(:item_code)")
                query_params["item_code"] = f"%{self._clean_input(params.item_code)}%"
            
            if params.item_name and isinstance(params.item_name, str):
                conditions.append("AND UPPER(ITEM_NAME) LIKE UPPER(:item_name)")
                query_params["item_name"] = f"%{self._clean_input(params.item_name)}%"
            
            if params.lot_name and isinstance(params.lot_name, str):
                conditions.append("AND UPPER(LOT_NAME) LIKE UPPER(:lot_name)")
                query_params["lot_name"] = f"%{self._clean_input(params.lot_name)}%"

            if params.status is not None:
                if params.status == 0:
                    conditions.append("AND STATUS = '0'")
                else:
                    conditions.append("AND STATUS != '0'")
            
            if params.doc_date_start:
                conditions.append("AND DOC_DATE >= :doc_date_start")
                query_params["doc_date_start"] = params.doc_date_start

            if params.doc_date_end:
                conditions.append("AND DOC_DATE <= :doc_date_end")
                query_params["doc_date_end"] = params.doc_date_end

            if params.supplier and isinstance(params.supplier, str):
                conditions.append("AND SUPPLIER LIKE :supplier")
                query_params["supplier"] = f"%{self._clean_input(params.supplier)}%"
                
            if params.progress_name and isinstance(params.progress_name, str):
                conditions.append("AND UPPER(PROGRESS_NAME) LIKE UPPER(:progress_name)")
                query_params["progress_name"] = f"%{self._clean_input(params.progress_name)}%"

            if params.testing_program_name and isinstance(params.testing_program_name, str):
                conditions.append("AND UPPER(TESTING_PROGRAM_NAME) LIKE UPPER(:testing_program_name)")
                query_params["testing_program_name"] = f"%{self._clean_input(params.testing_program_name)}%"

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " ORDER BY ID OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()
            
            # 转换为响应对象
            data = [
                CpTestOrders(
                    ID=row.ID,
                    DOC_NO=row.DOC_NO,
                    ITEM_CODE=row.ITEM_CODE,
                    ITEM_NAME=row.ITEM_NAME,
                    LOT_NAME=row.LOT_NAME,
                    BUSINESS_QTY=row.BUSINESS_QTY,
                    RECEIPT_QTY=row.RECEIPT_QTY,
                    WIP_QTY=row.WIP_QTY,
                    PROGRESS_NAME=row.PROGRESS_NAME,
                    TESTING_PROGRAM_NAME=row.TESTING_PROGRAM_NAME,
                    REMARK=row.REMARK,
                    DOC_DATE=row.DOC_DATE,
                    FIRST_ARRIVAL_DATE=row.FIRST_ARRIVAL_DATE,
                    SUPPLIER=row.SUPPLIER,
                    STATUS=row.STATUS
                ) for row in result
            ]
            
            # 获取总记录数
            count_query = f"""
                SELECT COUNT(1) 
                FROM (
                    SELECT 
                        HC.DOC_NO,
                        LEFT(HC.ITEM_NAME,6) AS ITEM_CODE,
                        HC.ITEM_NAME,
                        HC.LOT_NAME,
                        HC.BUSINESS_QTY,
                        HC.RECEIPT_QTY,
                        0 AS WIP_QTY,
                        HC.PROGRESS_NAME,
                        HC.TESTING_PROGRAM_NAME,
                        HC.REMARK,
                        HC.DOC_DATE,
                        HC.FIRST_ARRIVAL_DATE,
                        HC.SUPPLIER,
                        HC.STATUS
                    FROM huaxinAdmin_cptest HC
                    UNION ALL
                    SELECT
                            PO.DOC_NO,
                            ITEM.ITEM_CODE,
                            ITEM.ITEM_NAME,
                            ITEM_LOT.LOT_CODE,
                            PO_D.BUSINESS_QTY,
                            PO_D.RECEIPTED_PRICE_QTY,
                            PO_D.BUSINESS_QTY - PO_D.RECEIPTED_PRICE_QTY,
                            ZAC.CUSTOM_FIELD10 AS PROGRESS_NAME,
                            ZTP.Z_TESTING_PROGRAM_NAME,
                            Z_PACKAGE.REMARK,
                            PO.PURCHASE_DATE,
                            PR.CreateDate,
                            PO.SUPPLIER_FULL_NAME,
                            PO_SD.RECEIPT_CLOSE
                    FROM PURCHASE_ORDER PO
                    LEFT JOIN PURCHASE_ORDER_D PO_D 
                            ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                    LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                            ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                    LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                            ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                    LEFT JOIN Z_OUT_MO_D 
                            ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                    LEFT JOIN ITEM 
                            ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                    LEFT JOIN ITEM_LOT 
                            ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                    LEFT JOIN Z_ASSEMBLY_CODE 
                            ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_ASSEMBLY_CODE ZAC
                            ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                    LEFT JOIN Z_TESTING_PROGRAM ZTP
                            ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                    LEFT JOIN Z_PACKAGE 
                            ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                    OUTER APPLY (
                            SELECT TOP 1 *
                            FROM PURCHASE_RECEIPT_D PRD
                            WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                            ORDER BY PRD.CreateDate
                    ) PR
                    WHERE PO.PURCHASE_TYPE = 2 
                            AND PO.PURCHASE_DATE > '2024-11-13' 
                            AND ITEM.ITEM_CODE LIKE N'CL%CP' 
                            AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                            AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                            AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                    UNION ALL
                    SELECT 
                    MO.DOC_NO,
                    ITEM.ITEM_CODE,
                    ITEM.ITEM_NAME,
                    IL.LOT_CODE,
                    MO_D.REQUIRED_SECOND_QTY,
                    MP.COMPLETED_SECOND_QTY,
                    MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY,
                    NULL AS PROGRESS_NAME,
                    NULL AS Z_TESTING_PROGRAM_NAME,
                    NULL AS REMARK,
                    MO.DOC_DATE AS PURCHASE_DATE,
                    NULL AS FIRST_ARRIVAL_DATE,
                    '苏工院' AS SUPPLIER,
                    CASE
                      WHEN (MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY) > 0 THEN '0'
                      ELSE '2'
                    END AS RECEIPT_CLOSE
                    FROM MO
                    LEFT JOIN MO_D ON MO.MO_ID = MO_D.MO_ID
                    LEFT JOIN MO_PRODUCT MP ON MP.MO_ID = MO.MO_ID
                    LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = MO.ITEM_ID
                    LEFT JOIN ITEM_LOT IL ON IL.ITEM_LOT_ID = MO_D.ITEM_LOT_ID
                    WHERE ITEM.ITEM_CODE LIKE N'CL%CP' AND MO.DOC_NO LIKE '5105%'
                ) t
                WHERE 1=1 {' '.join(conditions)}
            """
            total = db.execute(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()
            
            return {
                "list": data,
                "total": total
            }
        except Exception as e:
            logger.error(f"获取CP测试单失败: {str(e)}")
            raise CustomException(f"获取CP测试单失败: {str(e)}")

    def export_cptest_orders_excel(self,db:Session,params:CpTestOrdersQuery)->bytes:
        """导出CP测试单Excel
        
        Args:
            db: 数据库会话
            params: 查询参数

        Returns:
            bytes: 导出的Excel文件
            
        Raises:
            CustomException: 当导出失败时抛出
        """
        try:
            # 获取数据
            params.pageIndex = 1
            params.pageSize = 10000000
            data = self.get_cptest_orders_by_params(db, params)
            
            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "CP测试单"
            
            # 设置表头
            headers = [
                "ID",
                "单据号",
                "晶圆编码",
                "晶圆名称",
                "批号",
                "业务数量",
                "收货数量",
                "在制数量",
                "测试流程",
                "测试程序",
                "备注",
                "单据日期",
                "首次到货日期",
                "供应商",
                "状态"
            ]

            # 设置列宽
            column_widths = {
                "A":10,
                "B":30,
                "C":20,
                "D":20,
                "E":20,
                "F":20,
                "G":20,
                "H":20,
                "I":40,
                "J":40,
                "K":40,
                "L":20,
                "M":20,
                "N":40,
                "O":10
            }

            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            for row, result in enumerate(data['list'], 2):
                data = [
                    result.ID,
                    result.DOC_NO,
                    result.ITEM_CODE,
                    result.ITEM_NAME,
                    result.LOT_NAME,
                    result.BUSINESS_QTY,
                    result.RECEIPT_QTY,
                    result.WIP_QTY,
                    result.PROGRESS_NAME,
                    result.TESTING_PROGRAM_NAME,
                    result.REMARK,
                    result.DOC_DATE,
                    result.FIRST_ARRIVAL_DATE,
                    result.SUPPLIER,
                    result.STATUS
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border

            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
        except Exception as e:
            logger.error(f"导出CP测试单Excel失败: {str(e)}")
            raise CustomException(f"导出CP测试单Excel失败: {str(e)}")
        
    def get_chipInfo_trace_by_params(self,db:Session,params:ChipInfoTraceQuery)->Dict[str,Any]:
        """获取芯片信息追溯"""
        try:
            base_query = """
            WITH
            CL AS (
            SELECT 
            ROW_NUMBER() OVER (ORDER BY DOC_DATE,DOC_NO) + 100000 AS ID,
            DOC_NO,
            ITEM_CODE,
            ITEM_NAME,
            LOT_NAME,
            CAST(BUSINESS_QTY AS INT) AS BUSINESS_QTY,
            CAST(RECEIPT_QTY AS INT) AS RECEIPT_QTY,
            CAST(WIP_QTY AS INT) AS WIP_QTY,
            PROGRESS_NAME,
            TESTING_PROGRAM_NAME,
            REMARK,
            CAST(DOC_DATE AS DATE) AS DOC_DATE,
            CAST(FIRST_ARRIVAL_DATE AS DATE) AS FIRST_ARRIVAL_DATE,
            SUPPLIER,
            STATUS
            FROM (
                SELECT 
                    HC.DOC_NO,
                    LEFT(HC.ITEM_NAME,6) AS ITEM_CODE,
                    HC.ITEM_NAME,
                    HC.LOT_NAME,
                    HC.BUSINESS_QTY,
                    HC.RECEIPT_QTY,
                    0 AS WIP_QTY,
                    HC.PROGRESS_NAME,
                    HC.TESTING_PROGRAM_NAME,
                    HC.REMARK,
                    HC.DOC_DATE,
                    HC.FIRST_ARRIVAL_DATE,
                    HC.SUPPLIER,
                    HC.STATUS
                FROM huaxinAdmin_cptest HC
                UNION ALL
                SELECT
                        PO.DOC_NO,
                        ITEM.ITEM_CODE,
                        ITEM.ITEM_NAME,
                        ITEM_LOT.LOT_CODE,
                        PO_D.BUSINESS_QTY,
                        PO_D.RECEIPTED_PRICE_QTY,
                        PO_D.BUSINESS_QTY - PO_D.RECEIPTED_PRICE_QTY,
                        ZAC.CUSTOM_FIELD10 AS PROGRESS_NAME,
                        ZTP.Z_TESTING_PROGRAM_NAME,
                        Z_PACKAGE.REMARK,
                        PO.PURCHASE_DATE,
                        PR.CreateDate,
                        PO.SUPPLIER_FULL_NAME,
                        PO_SD.RECEIPT_CLOSE
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D 
                        ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D 
                        ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                LEFT JOIN ITEM 
                        ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT 
                        ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE 
                        ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC
                        ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_TESTING_PROGRAM ZTP
                        ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                LEFT JOIN Z_PACKAGE 
                        ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                OUTER APPLY (
                        SELECT TOP 1 *
                        FROM PURCHASE_RECEIPT_D PRD
                        WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                        ORDER BY PRD.CreateDate
                ) PR
                WHERE PO.PURCHASE_TYPE = 2 
                        AND PO.PURCHASE_DATE > '2024-11-13' 
                        AND ITEM.ITEM_CODE LIKE N'CL%CP' 
                        AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                UNION ALL
                SELECT 
                MO.DOC_NO,
                ITEM.ITEM_CODE,
                ITEM.ITEM_NAME,
                IL.LOT_CODE,
                MO_D.REQUIRED_SECOND_QTY,
                MP.COMPLETED_SECOND_QTY,
                MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY,
                NULL AS PROGRESS_NAME,
                NULL AS Z_TESTING_PROGRAM_NAME,
                NULL AS REMARK,
                MO.DOC_DATE AS PURCHASE_DATE,
                NULL AS FIRST_ARRIVAL_DATE,
                '苏工院' AS SUPPLIER,
                CASE
                    WHEN (MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY) > 0 THEN '0'
                    ELSE '2'
                END AS RECEIPT_CLOSE
                FROM MO
                LEFT JOIN MO_D ON MO.MO_ID = MO_D.MO_ID
                LEFT JOIN MO_PRODUCT MP ON MP.MO_ID = MO.MO_ID
                LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = MO.ITEM_ID
                LEFT JOIN ITEM_LOT IL ON IL.ITEM_LOT_ID = MO_D.ITEM_LOT_ID
                WHERE ITEM.ITEM_CODE LIKE N'CL%CP' AND MO.DOC_NO LIKE '5105%'
            ) AS CombinedResults
            ),
            AL AS (
            SELECT CombinedResults.*,ALL_BOM.MAIN_CHIP,ALL_BOM.ITEM_CODE AS CHIP_CODE,ALL_BOM.LOT_CODE_NAME,ALL_BOM.BUSINESS_QTY AS WAFER_QTY,ALL_BOM.SECOND_QTY AS S_QTY,ALL_BOM.WAFER_ID
            FROM (
                SELECT
                    hpl.ID,
                    hpl.DOC_NO AS ASSY_DOC_NO,
                    hpl.ITEM_CODE AS CHIP_ITEM_CODE,
                    hpl.Z_PACKAGE_TYPE_NAME,
                    hpl.LOT_CODE AS CHIP_LOT_CODE,
                    hpl.BUSINESS_QTY AS CHIP_BUSINESS_QTY,
                    hpl.RECEIPTED_PRICE_QTY AS CHIP_RECEIPTED_PRICE_QTY,
                    0 AS CHIP_WIP_QTY,
                    hpl.Z_PROCESSING_PURPOSE_NAME,
                    hpl.Z_TESTING_PROGRAM_NAME,
                    hpl.Z_ASSEMBLY_CODE,
                    hpl.Z_WIRE_NAME,
                    hpl.REMARK AS CHIP_REMARK,
                    hpl.PURCHASE_DATE AS CHIP_PURCHASE_DATE,
                    ISNULL(hpl.FIRST_ARRIVAL_DATE, DATEADD(MONTH, 2, hpl.PURCHASE_DATE)) AS CHIP_FIRST_ARRIVAL_DATE,
                    hpl.SUPPLIER_FULL_NAME AS CHIP_SUPPLY_FULL_NAME,
                    hpl.RECEIPT_CLOSE AS CHIP_RECEIPT_CLOSE
                FROM HSUN_PACKAGE_LIST hpl
                UNION ALL
                SELECT
                    ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE, PO.DOC_NO) + 115617 AS ID,
                    PO.DOC_NO,
                    ITEM.ITEM_CODE,
                    ITEM.UDF025 AS Z_PACKAGE_TYPE_NAME,
                    ITEM_LOT.LOT_CODE,
                    CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
                    CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
                    CASE 
                        WHEN PO.[CLOSE] = N'2' THEN 0
                        WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                        ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
                    END AS WIP_QTY,
                    Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_NAME,
                    ZTP.Z_TESTING_PROGRAM_NAME,
                    Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE,
                    Z_WIRE.Z_WIRE_NAME,
                    Z_PACKAGE.REMARK,
                    CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
                    CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
                    PO.SUPPLIER_FULL_NAME,
                    PO_SD.RECEIPT_CLOSE
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D 
                    ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                    ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                    ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D 
                    ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                LEFT JOIN ITEM 
                    ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT 
                    ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE 
                    ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC
                    ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_TESTING_PROGRAM ZTP
                    ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                LEFT JOIN Z_PACKAGE 
                    ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                LEFT JOIN Z_PROCESSING_PURPOSE 
                    ON Z_ASSEMBLY_CODE.Z_PROCESSING_PURPOSE_ID = Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_ID
                LEFT JOIN Z_LOADING_METHOD 
                    ON Z_LOADING_METHOD.Z_LOADING_METHOD_ID = Z_PACKAGE.Z_LOADING_METHOD_ID
                LEFT JOIN Z_WIRE 
                    ON Z_WIRE.Z_WIRE_ID = Z_PACKAGE.Z_WIRE_ID
                LEFT JOIN FEATURE_GROUP 
                    ON FEATURE_GROUP.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                OUTER APPLY (
                    SELECT TOP 1 *
                    FROM PURCHASE_RECEIPT_D PRD
                    WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                    ORDER BY PRD.CreateDate
                ) PR
                WHERE PO.PURCHASE_TYPE = 2 
                    AND PO.PURCHASE_DATE > '2024-10-21' 
                    AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                    AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            ) AS CombinedResults
                LEFT JOIN 
                    (
                        SELECT * FROM HSUN_BOM_LIST
                        UNION ALL
                        SELECT 
                        ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE,PO.DOC_NO) + 16820 AS ID,
                        PO.DOC_NO,
                        ZOMSD.Z_MAIN_CHIP,
                        ITEM.ITEM_CODE,
                        ITEM.ITEM_NAME,
                        IL.LOT_CODE,
                        CAST(ZOMSD.BUSINESS_QTY AS FLOAT) AS BUSINESS_QTY,
                        CAST(ZOMSD.SECOND_QTY AS FLOAT) AS SECOND_QTY,
                        ZOMSD.Z_WF_ID_STRING
                        FROM PURCHASE_ORDER PO
                        LEFT JOIN PURCHASE_ORDER_D PO_D
                        ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                        LEFT JOIN PURCHASE_ORDER_SD PO_SD
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                        LEFT JOIN PURCHASE_ORDER_SSD PO_SSD
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                        LEFT JOIN Z_OUT_MO_D ZOMD
                        ON ZOMD.Z_OUT_MO_D_ID = PO_SSD.REFERENCE_SOURCE_ID_ROid
                        LEFT JOIN Z_OUT_MO_SD ZOMSD
                        ON ZOMSD.Z_OUT_MO_D_ID = ZOMD.Z_OUT_MO_D_ID
                        LEFT JOIN ITEM
                        ON ZOMSD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                        LEFT JOIN ITEM_LOT IL
                        ON IL.ITEM_LOT_ID = ZOMSD.ITEM_LOT_ID
                        WHERE PO_D.PURCHASE_TYPE=2 AND PO.PURCHASE_DATE > '2024-10-21')
                        AS ALL_BOM ON ALL_BOM.DOC_NO = CombinedResults.ASSY_DOC_NO
            )
          

            SELECT AL.*,CL.*
            FROM CL
            INNER JOIN AL ON AL.LOT_CODE_NAME LIKE '%' + CL.LOT_NAME + '%'
            WHERE 1=1 
            """

            # 构建查询条件
            conditions = []
            query_params = {}

            # 参数验证和清理
            if params.CHIP_LOT_CODE and isinstance(params.CHIP_LOT_CODE, str):
                conditions.append("AND UPPER(AL.CHIP_LOT_CODE) LIKE UPPER(:chip_lot_code)")
                query_params["chip_lot_code"] = f"%{self._clean_input(params.CHIP_LOT_CODE)}%"
            
            if params.WAFER_LOT_CODE and isinstance(params.WAFER_LOT_CODE, str):
                conditions.append("AND UPPER(AL.LOT_CODE_NAME) LIKE UPPER(:wafer_lot_code)")
                query_params["wafer_lot_code"] = f"%{self._clean_input(params.WAFER_LOT_CODE)}%"
            
            if params.SUPPLIER and isinstance(params.SUPPLIER, str):
                conditions.append("AND UPPER(CL.SUPPLIER) LIKE UPPER(:supplier)")
                query_params["supplier"] = f"%{self._clean_input(params.SUPPLIER)}%"
            
            if params.CHIP_NAME and isinstance(params.CHIP_NAME, str):
                conditions.append("AND UPPER(AL.CHIP_ITEM_CODE) LIKE UPPER(:chip_name)")
                query_params["chip_name"] = f"%{self._clean_input(params.CHIP_NAME)}%"
            
            if params.WAFER_NAME and isinstance(params.WAFER_NAME, str):
                conditions.append("AND UPPER(CL.CHIP_CODE) LIKE UPPER(:wafer_name)")
                query_params["wafer_name"] = f"%{self._clean_input(params.WAFER_NAME)}%"
            
            if params.TESTING_PROGRAM_NAME and isinstance(params.TESTING_PROGRAM_NAME, str):
                conditions.append("AND UPPER(CL.TESTING_PROGRAM_NAME) LIKE UPPER(:testing_program_name)")
                query_params["testing_program_name"] = f"%{self._clean_input(params.TESTING_PROGRAM_NAME)}%"
            

            # 拼接查询条件
            query = base_query + " " + " ".join(conditions)
            
            # 添加排序
            query += " ORDER BY AL.ASSY_DOC_NO"
            
            # 添加分页
            if params.pageIndex and params.pageSize:
                offset = (params.pageIndex - 1) * params.pageSize
                query += " OFFSET :offset ROWS FETCH NEXT :pageSize ROWS ONLY"
                query_params["offset"] = offset
                query_params["pageSize"] = params.pageSize

            # 执行查询
            stmt = text(query).bindparams(**query_params)
            result = db.execute(stmt).all()

            # 获取总记录数
            count_query = f"""
            WITH
            CL AS (
            SELECT 
            ROW_NUMBER() OVER (ORDER BY DOC_DATE,DOC_NO) + 100000 AS ID,
            DOC_NO,
            ITEM_CODE,
            ITEM_NAME,
            LOT_NAME,
            CAST(BUSINESS_QTY AS INT) AS BUSINESS_QTY,
            CAST(RECEIPT_QTY AS INT) AS RECEIPT_QTY,
            CAST(WIP_QTY AS INT) AS WIP_QTY,
            PROGRESS_NAME,
            TESTING_PROGRAM_NAME,
            REMARK,
            CAST(DOC_DATE AS DATE) AS DOC_DATE,
            CAST(FIRST_ARRIVAL_DATE AS DATE) AS FIRST_ARRIVAL_DATE,
            SUPPLIER,
            STATUS
            FROM (
                SELECT 
                    HC.DOC_NO,
                    LEFT(HC.ITEM_NAME,6) AS ITEM_CODE,
                    HC.ITEM_NAME,
                    HC.LOT_NAME,
                    HC.BUSINESS_QTY,
                    HC.RECEIPT_QTY,
                    0 AS WIP_QTY,
                    HC.PROGRESS_NAME,
                    HC.TESTING_PROGRAM_NAME,
                    HC.REMARK,
                    HC.DOC_DATE,
                    HC.FIRST_ARRIVAL_DATE,
                    HC.SUPPLIER,
                    HC.STATUS
                FROM huaxinAdmin_cptest HC
                UNION ALL
                SELECT
                        PO.DOC_NO,
                        ITEM.ITEM_CODE,
                        ITEM.ITEM_NAME,
                        ITEM_LOT.LOT_CODE,
                        PO_D.BUSINESS_QTY,
                        PO_D.RECEIPTED_PRICE_QTY,
                        PO_D.BUSINESS_QTY - PO_D.RECEIPTED_PRICE_QTY,
                        ZAC.CUSTOM_FIELD10 AS PROGRESS_NAME,
                        ZTP.Z_TESTING_PROGRAM_NAME,
                        Z_PACKAGE.REMARK,
                        PO.PURCHASE_DATE,
                        PR.CreateDate,
                        PO.SUPPLIER_FULL_NAME,
                        PO_SD.RECEIPT_CLOSE
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D 
                        ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D 
                        ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                LEFT JOIN ITEM 
                        ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT 
                        ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE 
                        ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC
                        ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_TESTING_PROGRAM ZTP
                        ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                LEFT JOIN Z_PACKAGE 
                        ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                OUTER APPLY (
                        SELECT TOP 1 *
                        FROM PURCHASE_RECEIPT_D PRD
                        WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                        ORDER BY PRD.CreateDate
                ) PR
                WHERE PO.PURCHASE_TYPE = 2 
                        AND PO.PURCHASE_DATE > '2024-11-13' 
                        AND ITEM.ITEM_CODE LIKE N'CL%CP' 
                        AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                        AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
                UNION ALL
                SELECT 
                MO.DOC_NO,
                ITEM.ITEM_CODE,
                ITEM.ITEM_NAME,
                IL.LOT_CODE,
                MO_D.REQUIRED_SECOND_QTY,
                MP.COMPLETED_SECOND_QTY,
                MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY,
                NULL AS PROGRESS_NAME,
                NULL AS Z_TESTING_PROGRAM_NAME,
                NULL AS REMARK,
                MO.DOC_DATE AS PURCHASE_DATE,
                NULL AS FIRST_ARRIVAL_DATE,
                '苏工院' AS SUPPLIER,
                CASE
                    WHEN (MO_D.REQUIRED_SECOND_QTY-MP.COMPLETED_SECOND_QTY) > 0 THEN '0'
                    ELSE '2'
                END AS RECEIPT_CLOSE
                FROM MO
                LEFT JOIN MO_D ON MO.MO_ID = MO_D.MO_ID
                LEFT JOIN MO_PRODUCT MP ON MP.MO_ID = MO.MO_ID
                LEFT JOIN ITEM ON ITEM.ITEM_BUSINESS_ID = MO.ITEM_ID
                LEFT JOIN ITEM_LOT IL ON IL.ITEM_LOT_ID = MO_D.ITEM_LOT_ID
                WHERE ITEM.ITEM_CODE LIKE N'CL%CP' AND MO.DOC_NO LIKE '5105%'
            ) AS CombinedResults
            ),
            AL AS (
            SELECT CombinedResults.*,ALL_BOM.MAIN_CHIP,ALL_BOM.ITEM_CODE AS CHIP_CODE,ALL_BOM.LOT_CODE_NAME,ALL_BOM.BUSINESS_QTY AS WAFER_QTY,ALL_BOM.SECOND_QTY AS S_QTY,ALL_BOM.WAFER_ID
            FROM (
                SELECT
                    hpl.ID,
                    hpl.DOC_NO AS ASSY_DOC_NO,
                    hpl.ITEM_CODE AS CHIP_ITEM_CODE,
                    hpl.Z_PACKAGE_TYPE_NAME,
                    hpl.LOT_CODE AS CHIP_LOT_CODE,
                    hpl.BUSINESS_QTY AS CHIP_BUSINESS_QTY,
                    hpl.RECEIPTED_PRICE_QTY AS CHIP_RECEIPTED_PRICE_QTY,
                    0 AS CHIP_WIP_QTY,
                    hpl.Z_PROCESSING_PURPOSE_NAME,
                    hpl.Z_TESTING_PROGRAM_NAME,
                    hpl.Z_ASSEMBLY_CODE,
                    hpl.Z_WIRE_NAME,
                    hpl.REMARK AS CHIP_REMARK,
                    hpl.PURCHASE_DATE AS CHIP_PURCHASE_DATE,
                    ISNULL(hpl.FIRST_ARRIVAL_DATE, DATEADD(MONTH, 2, hpl.PURCHASE_DATE)) AS CHIP_FIRST_ARRIVAL_DATE,
                    hpl.SUPPLIER_FULL_NAME AS CHIP_SUPPLY_FULL_NAME,
                    hpl.RECEIPT_CLOSE AS CHIP_RECEIPT_CLOSE
                FROM HSUN_PACKAGE_LIST hpl
                UNION ALL
                SELECT
                    ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE, PO.DOC_NO) + 115617 AS ID,
                    PO.DOC_NO,
                    ITEM.ITEM_CODE,
                    ITEM.UDF025 AS Z_PACKAGE_TYPE_NAME,
                    ITEM_LOT.LOT_CODE,
                    CAST(PO_D.BUSINESS_QTY AS INT) AS BUSINESS_QTY,
                    CAST(PO_D.RECEIPTED_PRICE_QTY AS INT) AS RECEIPTED_PRICE_QTY,
                    CASE 
                        WHEN PO.[CLOSE] = N'2' THEN 0
                        WHEN PO_D.BUSINESS_QTY <> 0 AND (PO_D.RECEIPTED_PRICE_QTY / PO_D.BUSINESS_QTY) > 0.992 THEN 0
                        ELSE CAST(((PO_D.BUSINESS_QTY * 0.996) - PO_D.RECEIPTED_PRICE_QTY) AS INT)
                    END AS WIP_QTY,
                    Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_NAME,
                    ZTP.Z_TESTING_PROGRAM_NAME,
                    Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE,
                    Z_WIRE.Z_WIRE_NAME,
                    Z_PACKAGE.REMARK,
                    CAST(PO.PURCHASE_DATE AS DATE) AS PURCHASE_DATE,
                    CAST(PR.CreateDate AS DATE) AS FIRST_ARRIVAL_DATE,
                    PO.SUPPLIER_FULL_NAME,
                    PO_SD.RECEIPT_CLOSE
                FROM PURCHASE_ORDER PO
                LEFT JOIN PURCHASE_ORDER_D PO_D 
                    ON PO_D.PURCHASE_ORDER_ID = PO.PURCHASE_ORDER_ID
                LEFT JOIN PURCHASE_ORDER_SD PO_SD 
                    ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                LEFT JOIN PURCHASE_ORDER_SSD PO_SSD 
                    ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                LEFT JOIN Z_OUT_MO_D 
                    ON PO_SSD.REFERENCE_SOURCE_ID_ROid = Z_OUT_MO_D.Z_OUT_MO_D_ID
                LEFT JOIN ITEM 
                    ON PO_D.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                LEFT JOIN ITEM_LOT 
                    ON Z_OUT_MO_D.ITEM_LOT_ID = ITEM_LOT.ITEM_LOT_ID
                LEFT JOIN Z_ASSEMBLY_CODE 
                    ON Z_OUT_MO_D.Z_PACKAGE_ASSEMBLY_CODE_ID = Z_ASSEMBLY_CODE.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_ASSEMBLY_CODE ZAC
                    ON Z_OUT_MO_D.Z_TESTING_ASSEMBLY_CODE_ID = ZAC.Z_ASSEMBLY_CODE_ID
                LEFT JOIN Z_TESTING_PROGRAM ZTP
                    ON ZAC.PROGRAM_ROid = ZTP.Z_TESTING_PROGRAM_ID
                LEFT JOIN Z_PACKAGE 
                    ON Z_ASSEMBLY_CODE.PROGRAM_ROid = Z_PACKAGE.Z_PACKAGE_ID
                LEFT JOIN Z_PROCESSING_PURPOSE 
                    ON Z_ASSEMBLY_CODE.Z_PROCESSING_PURPOSE_ID = Z_PROCESSING_PURPOSE.Z_PROCESSING_PURPOSE_ID
                LEFT JOIN Z_LOADING_METHOD 
                    ON Z_LOADING_METHOD.Z_LOADING_METHOD_ID = Z_PACKAGE.Z_LOADING_METHOD_ID
                LEFT JOIN Z_WIRE 
                    ON Z_WIRE.Z_WIRE_ID = Z_PACKAGE.Z_WIRE_ID
                LEFT JOIN FEATURE_GROUP 
                    ON FEATURE_GROUP.FEATURE_GROUP_ID = ITEM.FEATURE_GROUP_ID
                OUTER APPLY (
                    SELECT TOP 1 *
                    FROM PURCHASE_RECEIPT_D PRD
                    WHERE PRD.ORDER_SOURCE_ID_ROid = PO_SD.PURCHASE_ORDER_SD_ID
                    ORDER BY PRD.CreateDate
                ) PR
                WHERE PO.PURCHASE_TYPE = 2 
                    AND PO.PURCHASE_DATE > '2024-10-21' 
                    AND ITEM.ITEM_CODE LIKE N'BC%AB' 
                    AND PO.SUPPLIER_FULL_NAME <> N'温州镁芯微电子有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'苏州荐恒电子科技有限公司'  
                    AND PO.SUPPLIER_FULL_NAME <> N'深圳市华新源科技有限公司'
            ) AS CombinedResults
                LEFT JOIN 
                    (
                        SELECT * FROM HSUN_BOM_LIST
                        UNION ALL
                        SELECT 
                        ROW_NUMBER() OVER (ORDER BY PO.PURCHASE_DATE,PO.DOC_NO) + 16820 AS ID,
                        PO.DOC_NO,
                        ZOMSD.Z_MAIN_CHIP,
                        ITEM.ITEM_CODE,
                        ITEM.ITEM_NAME,
                        IL.LOT_CODE,
                        CAST(ZOMSD.BUSINESS_QTY AS FLOAT) AS BUSINESS_QTY,
                        CAST(ZOMSD.SECOND_QTY AS FLOAT) AS SECOND_QTY,
                        ZOMSD.Z_WF_ID_STRING
                        FROM PURCHASE_ORDER PO
                        LEFT JOIN PURCHASE_ORDER_D PO_D
                        ON PO.PURCHASE_ORDER_ID = PO_D.PURCHASE_ORDER_ID
                        LEFT JOIN PURCHASE_ORDER_SD PO_SD
                        ON PO_SD.PURCHASE_ORDER_D_ID = PO_D.PURCHASE_ORDER_D_ID
                        LEFT JOIN PURCHASE_ORDER_SSD PO_SSD
                        ON PO_SSD.PURCHASE_ORDER_SD_ID = PO_SD.PURCHASE_ORDER_SD_ID
                        LEFT JOIN Z_OUT_MO_D ZOMD
                        ON ZOMD.Z_OUT_MO_D_ID = PO_SSD.REFERENCE_SOURCE_ID_ROid
                        LEFT JOIN Z_OUT_MO_SD ZOMSD
                        ON ZOMSD.Z_OUT_MO_D_ID = ZOMD.Z_OUT_MO_D_ID
                        LEFT JOIN ITEM
                        ON ZOMSD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                        LEFT JOIN ITEM_LOT IL
                        ON IL.ITEM_LOT_ID = ZOMSD.ITEM_LOT_ID
                        WHERE PO_D.PURCHASE_TYPE=2 AND PO.PURCHASE_DATE > '2024-10-21')
                        AS ALL_BOM ON ALL_BOM.DOC_NO = CombinedResults.ASSY_DOC_NO
            )

            SELECT COUNT(*)
            FROM CL
            INNER JOIN AL ON AL.LOT_CODE_NAME LIKE '%' + CL.LOT_NAME + '%'
            WHERE 1=1 {' '.join(conditions)}
            """
            total = db.execute(text(count_query).bindparams(**{k:v for k,v in query_params.items() if k not in ['offset', 'pageSize']})).scalar()
            
            # 构建响应
            chip_trace_list = [
                ChipInfoTrace(
                    ID=row.ID,
                    DOC_NO=row.ASSY_DOC_NO,
                    ITEM_CODE=row.CHIP_ITEM_CODE,
                    Z_PACKAGE_TYPE_NAME=row.Z_PACKAGE_TYPE_NAME,
                    LOT_CODE=row.CHIP_LOT_CODE,
                    BUSINESS_QTY=row.CHIP_BUSINESS_QTY,
                    RECEIPTED_PRICE_QTY=row.CHIP_RECEIPTED_PRICE_QTY,
                    WIP_QTY=row.CHIP_WIP_QTY,
                    Z_PROCESSING_PURPOSE_NAME=row.Z_PROCESSING_PURPOSE_NAME,
                    Z_TESTING_PROGRAM_NAME=row.Z_TESTING_PROGRAM_NAME,
                    Z_ASSEMBLY_CODE=row.Z_ASSEMBLY_CODE,
                    Z_WIRE_NAME=row.Z_WIRE_NAME,
                    REMARK=row.CHIP_REMARK,
                    PURCHASE_DATE=row.CHIP_PURCHASE_DATE,
                    FIRST_ARRIVAL_DATE=row.CHIP_FIRST_ARRIVAL_DATE,
                    SUPPLIER_FULL_NAME=row.CHIP_SUPPLY_FULL_NAME,
                    MAIN_CHIP=row.MAIN_CHIP,
                    CHIP_CODE=row.CHIP_CODE,
                    LOT_CODE_NAME=row.LOT_CODE_NAME,
                    WAFER_QTY=row.BUSINESS_QTY,
                    S_QTY=row.RECEIPT_QTY,
                    WAFER_ID=row.WAFER_ID,
                    PROGRESS_NAME=row.PROGRESS_NAME,
                    TESTING_PROGRAM_NAME=row.TESTING_PROGRAM_NAME,
                    SUPPLIER=row.SUPPLIER
                )
                for row in result
            ]

            return {
                "list":chip_trace_list,
                "total":total
            }

        except Exception as e:
            logger.error(f"获取芯片信息追溯失败: {str(e)}")
            raise CustomException(f"获取芯片信息追溯失败: {str(e)}")