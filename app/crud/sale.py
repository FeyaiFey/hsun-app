import io
import uuid
from uuid import UUID
from datetime import datetime
from typing import List, Optional, Union, Dict, Any
from sqlmodel import Session, select, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.exceptions import CustomException
from app.core.logger import logger
from app.models.sale import Sale
from app.schemas.sale import (
    SaleTableQuery, SaleTableResponse, SaleTargetCreate, SaleTargetUpdate,
    SaleTargetSummaryQuery, SaleTargetSummary, SaleTargetSummaryResponse,
    SaleTargetDetailQuery, SaleTargetDetail, SaleTargetDetailResponse,
    SaleAmountAnalyzeQuery, SaleAmountAnalyze, SaleAmountAnalyzeResponse,
    SaleAnalysisPannel, SaleAnalysisPannelResponse, SaleForecastResponse,
    SaleTable,SaleAmount,SaleAmountResponse,SaleAmountQuery,
    SaleAmountBarChartQuery, SaleAmountBarChartEChartsResponse,
    SaleAmountBarChartEChartsLevelData, SaleAmountBarChartEChartsDataItem
)       
from app.utils.functions import Functions

class CRUSale:
    """销售CRUD操作类"""

    def _clean_input(self, value: str) -> str:
        """清理输入参数，移除潜在的危险字符
        
        Args:
            value: 输入字符串
            
        Returns:
            清理后的字符串
        """
        if value is None:
            return ""
        # 移除常见的 SQL 注入字符
        dangerous_chars = ["'", '"', ';', '--', '/*', '*/', 'xp_']
        cleaned = str(value)
        for char in dangerous_chars:
            cleaned = cleaned.replace(char, '')
        return cleaned
    
    async def get_sale_table(self, db: Session, params: SaleTableQuery) -> SaleTableResponse:
        """获取销售目标列表"""
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)
            admin_unit_name = self._clean_input(params.admin_unit_name)
            employee_name = self._clean_input(params.employee_name)

            # 构建查询条件
            query = select(Sale).order_by(Sale.CreatedAt.desc())
            if year:
                query = query.where(Sale.Year == year)
            if month:
                query = query.where(Sale.Month == month)
            if admin_unit_name:
                query = query.where(Sale.AdminUnitName == admin_unit_name)
            if employee_name:
                query = query.where(Sale.EmployeeName == employee_name)
            
            # 获取总记录数
            total_query = select(text("COUNT(*)")).select_from(query.subquery())
            total = db.exec(total_query).first()
            
            # 添加分页
            query = query.offset((params.pageIndex - 1) * params.pageSize).limit(params.pageSize)
            
            # 执行查询
            sale_targets = db.exec(query).all()
            
            # 将查询结果转换为 SaleTable 对象列表
            sale_table_list = []
            for target in sale_targets:
                sale_table_list.append(SaleTable(
                    Id=str(target.Id),
                    Year=target.Year,
                    Month=target.Month,
                    AdminUnitName=target.AdminUnitName,
                    EmployeeName=target.EmployeeName,
                    MonthlyTarget=target.MonthlyTarget,
                    CreatedBy=target.CreatedBy,
                    CreatedAt=target.CreatedAt,
                    UpdatedAt=target.UpdatedAt
                ))
            
            return SaleTableResponse(
                list=sale_table_list,
                total=total
            )
        except Exception as e:
            logger.error(f"获取销售目标列表失败: {str(e)}")
            raise CustomException(f"获取销售目标列表失败: {str(e)}")
    
    async def create_sale_target(self, db: Session, user_name: str, params: SaleTargetCreate) -> SaleTableResponse:
        """创建销售目标"""
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)
            admin_unit_name = self._clean_input(params.admin_unit_name)
            monthly_target = self._clean_input(params.monthly_target)
            employee_name = self._clean_input(params.employee_name)

            # 检查数据是否已经存在
            if year:
                query = select(Sale).where(Sale.Year == year, Sale.Month == None, Sale.EmployeeName == employee_name, Sale.AdminUnitName == admin_unit_name)
                sale_target = db.exec(query).all()
                if sale_target:
                    raise CustomException(f"{year}年{month}月{employee_name}的销售目标已存在")
            
            # 创建新的销售目标
            if year and month:
                new_target = Sale(
                    Id=uuid.uuid4(),
                    Year=year,
                    Month=month,
                    AdminUnitName=admin_unit_name,
                    EmployeeName=employee_name,
                    MonthlyTarget=monthly_target,
                    CreatedBy=user_name,
                    CreatedAt=datetime.now(),
                    UpdatedAt=datetime.now()
                )
                db.add(new_target)
                db.commit()
                db.refresh(new_target)
                return new_target
            else:
                raise CustomException("无效的输入参数")
        except Exception as e:
            logger.error(f"创建销售目标失败: {str(e)}")
            raise CustomException(f"创建销售目标失败: {str(e)}")

    async def update_sale_target(self, db: Session, params: SaleTargetUpdate) -> SaleTableResponse:
        """更新销售目标"""
        try:
            # 清理输入参数
            id = self._clean_input(params.id)
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)
            admin_unit_name = self._clean_input(params.admin_unit_name)
            employee_name = self._clean_input(params.employee_name)
            monthly_target = self._clean_input(params.monthly_target)


            # 检查数据是否存在
            query = select(Sale).where(Sale.Id == id)
            sale_target = db.exec(query).first()
            if not sale_target:
                raise CustomException("销售目标不存在")
            
            # 更新销售目标
            if year and month and employee_name:
                sale_target.Year = year
                sale_target.Month = month
                sale_target.AdminUnitName = admin_unit_name
                sale_target.EmployeeName = employee_name
                sale_target.MonthlyTarget = monthly_target
            else:
                raise CustomException("无效的输入参数")
            
            sale_target.UpdatedAt = datetime.now()
            db.commit()
            db.refresh(sale_target)
            return sale_target
        except Exception as e:
            logger.error(f"更新销售目标失败: {str(e)}")
            raise CustomException(f"更新销售目标失败: {str(e)}")

    async def delete_sale_target(self, db: Session, id: str) -> SaleTableResponse:
        """删除销售目标"""
        try:
            # 检查数据是否存在
            query = select(Sale).where(Sale.Id == id)
            sale_target = db.exec(query).first()
            if not sale_target:
                raise CustomException("销售目标不存在")
            
            # 删除销售目标
            db.delete(sale_target)
            db.commit()
            return sale_target
        except Exception as e:
            logger.error(f"删除销售目标失败: {str(e)}")
            raise CustomException(f"删除销售目标失败: {str(e)}")

    async def get_sale_target_summary(self, db: Session, params: SaleTargetSummaryQuery) -> List[SaleTargetSummaryResponse]:
        """获取销售目标汇总"""
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)

            # 构建查询语句 - 简化版本，避免使用临时表
            base_query = text(f"""
                WITH SalesTotalSummary AS (
                    SELECT 
                        NT.[YEAR],
                        NT.[MONTH],
                        NT.ADMIN_UNIT_NAME,
                        NT.EMPLOYEE_NAME,
                        SUM(NT.PRICE_QTY) AS PRICE_QTY,
                        SUM(NT.AMOUNT) AS AMOUNT
                    FROM (
                        SELECT 
                            YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                            MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                            AU.ADMIN_UNIT_NAME,
                            E.EMPLOYEE_NAME,
                            SID.PRICE_QTY,
                            SID.AMOUNT
                        FROM SALES_DELIVERY SD
                        JOIN EMPLOYEE E ON SD.Owner_Emp = E.EMPLOYEE_ID
                        JOIN EMPLOYEE_D ED ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                        JOIN ADMIN_UNIT AU ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                        LEFT JOIN SALES_DELIVERY_D SDD ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                        LEFT JOIN SALES_ISSUE_D SID ON SID.SOURCE_ID_ROid = SDD.SALES_DELIVERY_D_ID
                        LEFT JOIN SALES_ISSUE SI ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                        WHERE SD.CATEGORY = '24' 
                            AND YEAR(SI.TRANSACTION_DATE) = {year}
                            AND MONTH(SI.TRANSACTION_DATE) = {month}
                        
                        UNION ALL
                        
                        SELECT 
                            YEAR(SR.TRANSACTION_DATE) AS [YEAR],
                            MONTH(SR.TRANSACTION_DATE) AS [MONTH],
                            AU.ADMIN_UNIT_NAME,
                            E.EMPLOYEE_NAME,
                            SRD.PRICE_QTY * -1 AS PRICE_QTY,
                            SRD.AMOUNT * -1 AS AMOUNT
                        FROM SALES_RETURN SR
                        JOIN EMPLOYEE E ON SR.Owner_Emp = E.EMPLOYEE_ID
                        JOIN EMPLOYEE_D ED ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                        JOIN ADMIN_UNIT AU ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                        LEFT JOIN SALES_RETURN_D SRD ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                        WHERE SR.CATEGORY = '26' 
                            AND SR.RECEIPTED_STATUS = '3'
                            AND YEAR(SR.TRANSACTION_DATE) = {year}
                            AND MONTH(SR.TRANSACTION_DATE) = {month}
                    ) AS NT
                    GROUP BY NT.[YEAR], NT.[MONTH], NT.ADMIN_UNIT_NAME, NT.EMPLOYEE_NAME
                ),
                SalesForecastSummary AS (
                    SELECT 
                        YEAR(SFDD.TIME_BUCKET_SECTION) AS [YEAR],
                        MONTH(SFDD.TIME_BUCKET_SECTION) AS [MONTH],
                        E.EMPLOYEE_NAME,
                        SUM(SFDD.FORECAST_QTY) AS FORECAST_QTY
                    FROM SALES_FORECAST_DOC SFD
                    JOIN EMPLOYEE E ON E.EMPLOYEE_ID = SFD.Owner_Emp
                    LEFT JOIN SALES_FORECAST_DOC_D SFDD ON SFDD.SALES_FORECAST_DOC_ID = SFD.SALES_FORECAST_DOC_ID
                    WHERE SFDD.COLUMN_NO = 'COLUMN001' 
                        AND YEAR(SFDD.TIME_BUCKET_SECTION) = {year}
                        AND MONTH(SFDD.TIME_BUCKET_SECTION) = {month}
                    GROUP BY YEAR(SFDD.TIME_BUCKET_SECTION), MONTH(SFDD.TIME_BUCKET_SECTION), E.EMPLOYEE_NAME
                )
                SELECT 
                    SFS.[YEAR],
                    SFS.[MONTH],
                    ISNULL(STS.ADMIN_UNIT_NAME, '') AS ADMIN_UNIT_NAME,
                    SFS.[EMPLOYEE_NAME],
                    CAST(SFS.FORECAST_QTY AS INT) AS FORECAST_QTY,
                    ISNULL(STS.PRICE_QTY, 0) AS PRICE_QTY,
                    CASE 
                        WHEN SFS.FORECAST_QTY > 0 THEN CAST(ISNULL(STS.PRICE_QTY, 0)/SFS.FORECAST_QTY*100 AS DECIMAL(10,2))
                        ELSE 0 
                    END AS PERCENTAGE
                FROM SalesForecastSummary SFS
                LEFT JOIN SalesTotalSummary STS 
                    ON SFS.[YEAR] = STS.[YEAR] 
                    AND SFS.[MONTH] = STS.[MONTH] 
                    AND SFS.EMPLOYEE_NAME = STS.EMPLOYEE_NAME
                ORDER BY SFS.[YEAR], SFS.[MONTH], STS.ADMIN_UNIT_NAME DESC, SFS.[EMPLOYEE_NAME]
                """)

            # 执行查询
            sale_targets = db.exec(base_query).all()
            
            # 将查询结果转换为响应格式
            result_list = []
            for target in sale_targets:
                result_list.append(SaleTargetSummary(
                    YEAR=target.YEAR,
                    MONTH=target.MONTH,
                    ADMIN_UNIT_NAME=target.ADMIN_UNIT_NAME,
                    EMPLOYEE_NAME=target.EMPLOYEE_NAME,
                    FORECAST_QTY=target.FORECAST_QTY,
                    PRICE_QTY=target.PRICE_QTY,
                    PERCENTAGE=target.PERCENTAGE
                ))
            
            return SaleTargetSummaryResponse(list=result_list)
        except Exception as e:
            logger.error(f"获取销售目标汇总失败: {str(e)}")
            raise CustomException(f"获取销售目标汇总失败: {str(e)}")

    async def export_sale_target_summary(self, db: Session, params: SaleTargetSummaryQuery) -> bytes:
        """导出销售目标汇总"""
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)

            result = await self.get_sale_target_summary(db,params)
            if not result:
                raise CustomException("没有数据可导出")
            
            sale_target_summary = result.list

            # 创建Excel文件
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "封装订单"
            
            # 定义表头
            headers = [
                "年份", "月份", "销售团队", "业务员", "预测销量", "实际销量", "达成率"
            ]

            # 设置列宽
            column_widths = {
                'A': 15,  # 年份
                'B': 15,  # 月份
                'C': 20,  # 销售团队
                'D': 15,  # 业务员
                'E': 15,  # 预测销量
                'F': 15,  # 实际销量
                'G': 15,  # 达成率
            }
            
            # 设置样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = column_widths[get_column_letter(col)]
            
            # 写入数据
            total_forecast = 0
            total_price = 0
            
            for row, order in enumerate(sale_target_summary, 2):
                data = [
                    order.YEAR,
                    order.MONTH,
                    order.ADMIN_UNIT_NAME,
                    order.EMPLOYEE_NAME,
                    order.FORECAST_QTY,
                    order.PRICE_QTY,
                    order.PERCENTAGE/100
                ]
                
                # 累加合计数据
                total_forecast += order.FORECAST_QTY
                total_price += order.PRICE_QTY
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border

                    # 设置数字列的格式
                    if col in [5, 6]:  # 预测销量、实际销量
                        cell.number_format = '#,##0'

                    # 设置百分比列的格式
                    if col == 7:  # 达成率
                        cell.number_format = '0.00%'
            
            # 写入合计行
            total_row = len(sale_target_summary) + 2
            total_data = [
                "合计",
                "",
                "",
                "",
                total_forecast,
                total_price,
                total_price/total_forecast if total_forecast > 0 else 0
            ]
            
            # 设置合计行样式
            total_font = Font(name='微软雅黑', size=11, bold=True)
            
            for col, value in enumerate(total_data, 1):
                cell = ws.cell(row=total_row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = border
                cell.font = total_font
                
                if col in [5, 6]:  # 预测销量、实际销量合计
                    cell.number_format = '#,##0'
                if col == 7:  # 总达成率
                    cell.number_format = '0.00%'
            
            # 冻结首行
            ws.freeze_panes = 'A2'            
            # 保存到内存
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            return excel_file.getvalue()
        
        except Exception as e:
            logger.error(f"导出备货计划Excel失败: {str(e)}")
            raise CustomException("导出备货计划Excel失败")
        
    async def get_sale_target_detail(self, db: Session, params: SaleTargetDetailQuery) -> List[SaleTargetDetailResponse]:
        """获取销售目标详情"""
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)
            employee_name = self._clean_input(params.employee_name)

            # 构建查询条件
            where_clause = ""
            if year:
                where_clause += f" AND SS.[YEAR] = {year}"
            if month:
                where_clause += f" AND SS.[MONTH] = {month}"
            if employee_name:
                where_clause += f" AND SS.[EMPLOYEE_NAME] = '{employee_name}'"

            # 构建查询条件
            base_query = text(f"""
                DECLARE @YYYY INT = {year};
                DECLARE @MM INT = {month};
                DECLARE @EN NVARCHAR(255) = '{employee_name}';

                    WITH
                        SalesSummary AS
                        (
                        SELECT NT.[YEAR],[MONTH],ADMIN_UNIT_NAME,EMPLOYEE_NAME,ITEM_NAME,SHORTCUT,SUM(PRICE_QTY) AS PRICE_QTY,SUM(AMOUNT) AS AMOUNT
                        FROM
                            (
                            ( 
                                SELECT 
                                YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                                MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                                AU.ADMIN_UNIT_NAME,
                                E.EMPLOYEE_NAME,
                                ITEM.ITEM_CODE,
                                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') AS ITEM_NAME,
                                ITEM.SHORTCUT,
                                SID.PRICE_QTY,
                                SID.AMOUNT
                                FROM SALES_DELIVERY SD
                                LEFT JOIN SALES_DELIVERY_D SDD
                                ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                                LEFT JOIN SALES_ISSUE_D SID
                                ON SID.SOURCE_ID_ROid = SDD. SALES_DELIVERY_D_ID
                                LEFT JOIN SALES_ISSUE SI
                                ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                                LEFT JOIN EMPLOYEE E
                                ON SD.Owner_Emp = E.EMPLOYEE_ID
                                LEFT JOIN EMPLOYEE_D ED
                                ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                                LEFT JOIN ADMIN_UNIT AU
                                ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                                LEFT JOIN ITEM
                                ON SDD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                                WHERE SD.CATEGORY = '24' AND YEAR(SI.TRANSACTION_DATE) = @YYYY AND MONTH(SI.TRANSACTION_DATE) = @MM AND E.EMPLOYEE_NAME = @EN
                            )
                            UNION ALL
                            (
                                SELECT 
                                YEAR(SR.TRANSACTION_DATE) AS [YEAR],
                                MONTH(SR.TRANSACTION_DATE) AS [MONTH],
                                AU.ADMIN_UNIT_NAME,
                                E.EMPLOYEE_NAME,
                                ITEM.ITEM_CODE,
                                dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') AS ITEM_NAME,
                                ITEM.SHORTCUT,
                                SRD.PRICE_QTY * -1 AS PRICE_QTY,
                                SRD.AMOUNT * -1 AS AMOUNT
                                FROM SALES_RETURN SR
                                LEFT JOIN SALES_RETURN_D SRD
                                ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                                LEFT JOIN EMPLOYEE E
                                ON SR.Owner_Emp = E.EMPLOYEE_ID
                                LEFT JOIN EMPLOYEE_D ED
                                ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                                LEFT JOIN ADMIN_UNIT AU
                                ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                                LEFT JOIN ITEM
                                ON SRD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                                WHERE SR.CATEGORY = '26' AND SR.RECEIPTED_STATUS = '3' AND YEAR(SR.TRANSACTION_DATE) = @YYYY AND MONTH(SR.TRANSACTION_DATE) = @MM AND E.EMPLOYEE_NAME = @EN
                            )
                            ) AS NT
                            WHERE NT.YEAR > 0
                            GROUP BY NT.[YEAR],[MONTH],ADMIN_UNIT_NAME,EMPLOYEE_NAME,ITEM_NAME,SHORTCUT
                        ),
                    SalesForecast AS 
                        (
                            SELECT 
                            SFD.DOC_NO,
                            SFD.DOC_DATE,
                            SFD.CATEGORY,
                            SFD.BEGIN_DATE,
                            SFD.END_DATE,
                            SFD.FORECAST_VERSION,
                            SFD.BUCKET_SECTION_STRING,
                            SFD.FORECAST_DATE,
                            E.EMPLOYEE_NAME,
                            AU.ADMIN_UNIT_NAME,
                            ITEM.ITEM_CODE,
                            ITEM.SHORTCUT,
                            dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') AS ITEM_NAME,
                            SFDD.COLUMN_NO,
                            SFDD.TIME_BUCKET_SECTION,
                            SFDD.SECTION_END_DATE,
                            SFDD.FORECAST_QTY
                            FROM SALES_FORECAST_DOC SFD
                            LEFT JOIN SALES_FORECAST_DOC_D SFDD
                            ON SFDD.SALES_FORECAST_DOC_ID = SFD.SALES_FORECAST_DOC_ID
                            LEFT JOIN EMPLOYEE E
                            ON E.EMPLOYEE_ID = SFD.Owner_Emp
                            LEFT JOIN EMPLOYEE_D ED
                            ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                            LEFT JOIN ADMIN_UNIT AU
                            ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                            LEFT JOIN ITEM 
                            ON ITEM.ITEM_BUSINESS_ID = SFDD.ITEM_ID
                            WHERE SFDD.FORECAST_QTY > 0 AND SFDD.COLUMN_NO = 'COLUMN001' AND YEAR(SFDD.TIME_BUCKET_SECTION) = @YYYY AND MONTH(SFDD.TIME_BUCKET_SECTION) = @MM AND E.EMPLOYEE_NAME = @EN
                        )
                        

                    SELECT 
                        YEAR(SF.TIME_BUCKET_SECTION) AS [YEAR],
                        MONTH(SF.TIME_BUCKET_SECTION) AS [MONTH],
                        SF.ADMIN_UNIT_NAME,
                        SF.EMPLOYEE_NAME,
                        SF.SHORTCUT,
                        SF.ITEM_NAME,
                        CAST(SF.FORECAST_QTY AS INT) AS FORECAST_QTY,
                        CAST(SS.PRICE_QTY AS INT) AS PRICE_QTY,
                        CAST(SS.PRICE_QTY/SF.FORECAST_QTY*100 AS DECIMAL(18,2)) AS PERCENTAGE
                    FROM SalesForecast SF
                    LEFT JOIN SalesSummary SS
                    ON SF.EMPLOYEE_NAME = SS.EMPLOYEE_NAME AND SF.ITEM_NAME = SS.ITEM_NAME AND YEAR(SF.TIME_BUCKET_SECTION) = SS.[YEAR] AND MONTH(SF.TIME_BUCKET_SECTION) = SS.[MONTH]
                    ORDER BY YEAR(SF.TIME_BUCKET_SECTION),MONTH(SF.TIME_BUCKET_SECTION),SF.EMPLOYEE_NAME;
                """)

            # 执行查询
            sale_targets = db.exec(base_query).all()
            
            # 将查询结果转换为响应格式
            result_list = []
            for target in sale_targets:
                result_list.append(SaleTargetDetail(
                    YEAR=target.YEAR,
                    MONTH=target.MONTH,
                    ADMIN_UNIT_NAME=target.ADMIN_UNIT_NAME,
                    EMPLOYEE_NAME=target.EMPLOYEE_NAME,
                    SHORTCUT=target.SHORTCUT,
                    ITEM_NAME=target.ITEM_NAME,
                    FORECAST_QTY=target.FORECAST_QTY,
                    PRICE_QTY=target.PRICE_QTY,
                    PERCENTAGE=target.PERCENTAGE
                ))
            
            return SaleTargetDetailResponse(list=result_list)
        except Exception as e:
            logger.error(f"获取销售目标详情失败: {str(e)}")
            raise CustomException(f"获取销售目标详情失败: {str(e)}")

    async def get_sale_amount_analyze(self, db: Session, params: SaleAmountAnalyzeQuery) -> List[SaleAmountAnalyzeResponse]:
        """获取销售金额分析"""
        try:
            # 初始化变量
            where_clause_delivery = ""
            where_clause_return = ""
            
            # 清理输入参数并构建WHERE子句
            if params.year:
                year = self._clean_input(params.year)
                where_clause_delivery += f" AND YEAR(SI.TRANSACTION_DATE) = {year}"
                where_clause_return += f" AND YEAR(SR.TRANSACTION_DATE) = {year}"
            if params.month:
                month = self._clean_input(params.month)
                where_clause_delivery += f" AND MONTH(SI.TRANSACTION_DATE) = {month}"
                where_clause_return += f" AND MONTH(SR.TRANSACTION_DATE) = {month}"
            if params.shortcut:
                shortcut = self._clean_input(params.shortcut)
                where_clause_delivery += f" AND ITEM.SHORTCUT = '{shortcut}'"
                where_clause_return += f" AND ITEM.SHORTCUT = '{shortcut}'"
            if params.admin_unit_name:
                admin_unit_name = self._clean_input(params.admin_unit_name)
                where_clause_delivery += f" AND AU.ADMIN_UNIT_NAME = '{admin_unit_name}'"
                where_clause_return += f" AND AU.ADMIN_UNIT_NAME = '{admin_unit_name}'"
            if params.employee_name:
                employee_name = self._clean_input(params.employee_name)
                where_clause_delivery += f" AND E.EMPLOYEE_NAME = '{employee_name}'"
                where_clause_return += f" AND E.EMPLOYEE_NAME = '{employee_name}'"
            if params.item_name:
                item_name = self._clean_input(params.item_name)
                where_clause_delivery += f" AND dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') = '{item_name}'"
                where_clause_return += f" AND dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') = '{item_name}'"
            
            # 构建GROUP BY子句
            group_by_parts = []
            
            # 根据参数添加分组字段
            if params.group_by_year:
                group_by_parts.append("NT.[YEAR]")
            if params.group_by_month:
                group_by_parts.append("NT.[MONTH]")
            if params.group_by_shortcut:
                group_by_parts.append("NT.SHORTCUT")
            if params.group_by_admin_unit_name:
                group_by_parts.append("NT.ADMIN_UNIT_NAME")
            if params.group_by_employee_name:
                group_by_parts.append("NT.EMPLOYEE_NAME")
            if params.group_by_item_name:
                group_by_parts.append("NT.ITEM_NAME")
                
            # 确保至少有一个分组字段
            if len(group_by_parts) == 0:
                group_by_parts.append("NT.[YEAR]")  # 默认按年份分组
                
            group_by = ", ".join(group_by_parts)

            base_query = text(f"""
                SELECT 
                    {group_by},
                    SUM(NT.PRICE_QTY) AS PRICE_QTY,
                    SUM(NT.AMOUNT) AS AMOUNT
                FROM
                (
                ( 
                        SELECT 
                        YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                        MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                        AU.ADMIN_UNIT_NAME,
                        E.EMPLOYEE_NAME,
                        ITEM.ITEM_CODE,
                        dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') AS ITEM_NAME,
                        ITEM.SHORTCUT,
                        SID.PRICE_QTY,
                        SID.AMOUNT
                        FROM SALES_DELIVERY SD
                        LEFT JOIN SALES_DELIVERY_D SDD
                        ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                        LEFT JOIN SALES_ISSUE_D SID
                        ON SID.SOURCE_ID_ROid = SDD. SALES_DELIVERY_D_ID
                        LEFT JOIN SALES_ISSUE SI
                        ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                        LEFT JOIN EMPLOYEE E
                        ON SD.Owner_Emp = E.EMPLOYEE_ID
                        LEFT JOIN EMPLOYEE_D ED
                        ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                        LEFT JOIN ADMIN_UNIT AU
                        ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                        LEFT JOIN ITEM
                        ON SDD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                        WHERE SID.PRICE_QTY > 0 AND SD.CATEGORY = '24' {where_clause_delivery}
                )
                UNION ALL
                (
                        SELECT 
                        YEAR(SR.TRANSACTION_DATE) AS [YEAR],
                        MONTH(SR.TRANSACTION_DATE) AS [MONTH],
                        AU.ADMIN_UNIT_NAME,
                        E.EMPLOYEE_NAME,
                        ITEM.ITEM_CODE,
                        dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR') AS ITEM_NAME,
                        ITEM.SHORTCUT,
                        SRD.PRICE_QTY * -1 AS PRICE_QTY,
                        SRD.AMOUNT * -1 AS AMOUNT
                        FROM SALES_RETURN SR
                        LEFT JOIN SALES_RETURN_D SRD
                        ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                        LEFT JOIN EMPLOYEE E
                        ON SR.Owner_Emp = E.EMPLOYEE_ID
                        LEFT JOIN EMPLOYEE_D ED
                        ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                        LEFT JOIN ADMIN_UNIT AU
                        ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                        LEFT JOIN ITEM
                        ON SRD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                        WHERE SR.CATEGORY = '26' AND SR.RECEIPTED_STATUS = '3' {where_clause_return}
                )
                ) AS NT
                GROUP BY {group_by}
                ORDER BY {group_by}
                """)

            # 执行查询
            sale_amounts = db.exec(base_query).all()

            # 将查询结果转换为响应格式
            result_list = []
            for amount in sale_amounts:
                # 创建响应对象，只设置需要的字段
                sale_amount = SaleAmountAnalyze(
                    PRICE_QTY=amount.PRICE_QTY,
                    AMOUNT=amount.AMOUNT
                )
                
                # 根据分组字段设置相应的字段值
                if params.group_by_year and hasattr(amount, 'YEAR'):
                    sale_amount.YEAR = amount.YEAR
                if params.group_by_month and hasattr(amount, 'MONTH'):
                    sale_amount.MONTH = amount.MONTH
                if params.group_by_shortcut and hasattr(amount, 'SHORTCUT'):
                    sale_amount.SHORTCUT = amount.SHORTCUT
                if params.group_by_admin_unit_name and hasattr(amount, 'ADMIN_UNIT_NAME'):
                    sale_amount.ADMIN_UNIT_NAME = amount.ADMIN_UNIT_NAME
                if params.group_by_employee_name and hasattr(amount, 'EMPLOYEE_NAME'):
                    sale_amount.EMPLOYEE_NAME = amount.EMPLOYEE_NAME
                if params.group_by_item_name and hasattr(amount, 'ITEM_NAME'):
                    sale_amount.ITEM_NAME = amount.ITEM_NAME
                
                result_list.append(sale_amount)
            
            return SaleAmountAnalyzeResponse(list=result_list)
        except Exception as e:
            logger.error(f"获取销售金额分析失败: {str(e)}")
            raise CustomException(f"获取销售金额分析失败: {str(e)}")

    async def get_sale_analysis_pannel(self, db: Session) -> SaleAnalysisPannelResponse:
        """获取销售分析面板"""
        try:
            # 获取本年销售额
            base_query = text(f"""
                                WITH
                SS AS
                    (
                        SELECT 
                            CAST(NT.[DATE] AS DATE) AS [DATE],
                            SUM(NT.PRICE_QTY) AS PRICE_QTY,
                            SUM(NT.AMOUNT) AS AMOUNT
                            FROM
                                (
                                    (
                                        SELECT 
                                            SI.TRANSACTION_DATE AS [DATE],
                                            SID.PRICE_QTY,
                                            SID.AMOUNT
                                        FROM SALES_DELIVERY SD
                                        LEFT JOIN SALES_DELIVERY_D SDD
                                        ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                                        LEFT JOIN SALES_ISSUE_D SID
                                        ON SID.SOURCE_ID_ROid = SDD. SALES_DELIVERY_D_ID
                                        LEFT JOIN SALES_ISSUE SI
                                        ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                                        WHERE SID.PRICE_QTY > 0
                                    )
                                    UNION ALL
                                    (
                                        SELECT 
                                        SR.TRANSACTION_DATE AS [DATE],
                                        SRD.PRICE_QTY * -1 AS PRICE_QTY,
                                        SRD.AMOUNT * -1 AS AMOUNT
                                        FROM SALES_RETURN SR
                                        LEFT JOIN SALES_RETURN_D SRD
                                        ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                                        WHERE SR.CATEGORY = '26' AND SR.RECEIPTED_STATUS = '3'
                                    ) 
                                )AS NT
                            GROUP BY NT.[DATE]
                        )
        
                
                        
                SELECT
                    -- 今日数据
                    SUM(CASE WHEN [DATE] = CAST(GETDATE() AS DATE) THEN PRICE_QTY ELSE 0 END) AS today_sale_qty,
                    SUM(CASE WHEN [DATE] = CAST(GETDATE() AS DATE) THEN AMOUNT ELSE 0 END) AS today_sale_amount,
                    
                    -- 昨日数据
                    SUM(CASE WHEN [DATE] = CAST(DATEADD(DAY,-1,GETDATE()) AS DATE) THEN PRICE_QTY ELSE 0 END) AS yesterday_sale_qty,
                    SUM(CASE WHEN [DATE] = CAST(DATEADD(DAY,-1,GETDATE()) AS DATE) THEN AMOUNT ELSE 0 END) AS yesterday_sale_amount,
                    
                    -- 本年数据
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) THEN PRICE_QTY ELSE 0 END) AS this_year_sale_qty,
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) THEN AMOUNT ELSE 0 END) AS this_year_sale_amount,
                    
                    -- 本月数据
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) THEN PRICE_QTY ELSE 0 END) AS this_month_sale_qty,
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) THEN AMOUNT ELSE 0 END) AS this_month_sale_amount,
                    
                    -- 去年数据
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) - 1 THEN PRICE_QTY ELSE 0 END) AS last_year_sale_qty,
                    SUM(CASE WHEN YEAR([DATE]) = YEAR(GETDATE()) - 1 THEN AMOUNT ELSE 0 END) AS last_year_sale_amount,
                    
                    -- 上月数据
                    SUM(CASE WHEN 
                        [DATE] BETWEEN 
                          DATEFROMPARTS(YEAR(DATEADD(MONTH, -1, GETDATE())),MONTH(DATEADD(MONTH, -1, GETDATE())),1)
                        AND
                          EOMONTH(DATEADD(MONTH, -1, GETDATE()))
                    THEN PRICE_QTY ELSE 0 END) AS last_month_sale_qty,
                        
                    SUM(CASE WHEN 
                        [DATE] BETWEEN 
                          DATEFROMPARTS(YEAR(DATEADD(MONTH, -1, GETDATE())),MONTH(DATEADD(MONTH, -1, GETDATE())),1)
                        AND
                          EOMONTH(DATEADD(MONTH, -1, GETDATE()))
                    THEN AMOUNT ELSE 0 END) AS last_month_sale_amount,
                        
                        -- 上上月数据
                    SUM(CASE WHEN 
                        (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                    THEN PRICE_QTY ELSE 0 END) AS last_last_month_sale_qty,
                        
                        SUM(CASE WHEN 
                        (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                    THEN AMOUNT ELSE 0 END) AS last_last_month_sale_amount,
                    
                    -- 环比计算 (上月 vs 上上月)
                    ROUND(
                        (SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN PRICE_QTY ELSE 0 END) 
                        - SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                        THEN PRICE_QTY ELSE 0 END)) 
                        / NULLIF(SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                        THEN PRICE_QTY ELSE 0 END), 0) * 100, 2
                    ) AS month_on_month_qty,
                    
                    ROUND(
                        (SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN AMOUNT ELSE 0 END) 
                        - SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                        THEN AMOUNT ELSE 0 END)) 
                        / NULLIF(SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(DATEADD(MONTH, -2, GETDATE())) AND MONTH([DATE]) = MONTH(DATEADD(MONTH, -2, GETDATE())))
                        THEN AMOUNT ELSE 0 END), 0) * 100, 2
                    ) AS month_on_month_amount,
                    
                    -- 同比计算 (上月 vs 去年同月)
                    ROUND(
                        (SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN PRICE_QTY ELSE 0 END) 
                        - SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 2 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN PRICE_QTY ELSE 0 END)) 
                        / NULLIF(SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 2 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN PRICE_QTY ELSE 0 END), 0) * 100, 2
                    ) AS year_on_year_qty,
                    
                    ROUND(
                        (SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN AMOUNT ELSE 0 END) 
                        - SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 2 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN AMOUNT ELSE 0 END)) 
                        / NULLIF(SUM(CASE WHEN 
                            (YEAR([DATE]) = YEAR(GETDATE()) - 1 AND MONTH([DATE]) = MONTH(GETDATE()) - 1) 
                            OR (YEAR([DATE]) = YEAR(GETDATE()) - 2 AND MONTH([DATE]) = 12 AND MONTH(GETDATE()) = 1) 
                        THEN AMOUNT ELSE 0 END), 0) * 100, 2
                    ) AS year_on_year_amount
                FROM SS;
            """)

            # 执行查询
            result = db.exec(base_query).first()
            
            # 直接使用查询结果创建对象
            if result:
                panel = SaleAnalysisPannel(
                    today_sale_qty=int(result.today_sale_qty or 0),
                    today_sale_amount=float(result.today_sale_amount or 0),
                    yesterday_sale_qty=int(result.yesterday_sale_qty or 0),
                    yesterday_sale_amount=float(result.yesterday_sale_amount or 0),
                    this_year_sale_qty=int(result.this_year_sale_qty or 0),
                    this_year_sale_amount=float(result.this_year_sale_amount or 0),
                    this_month_sale_qty=int(result.this_month_sale_qty or 0),
                    this_month_sale_amount=float(result.this_month_sale_amount or 0),
                    last_year_sale_qty=int(result.last_year_sale_qty or 0),
                    last_year_sale_amount=float(result.last_year_sale_amount or 0),
                    last_month_sale_qty=int(result.last_month_sale_qty or 0),
                    last_month_sale_amount=float(result.last_month_sale_amount or 0),
                    last_last_month_sale_qty=int(result.last_last_month_sale_qty or 0),
                    last_last_month_sale_amount=float(result.last_last_month_sale_amount or 0),
                    month_on_month_qty=float(result.month_on_month_qty or 0),
                    month_on_month_amount=float(result.month_on_month_amount or 0),
                    year_on_year_qty=float(result.year_on_year_qty or 0),
                    year_on_year_amount=float(result.year_on_year_amount or 0)
                )
                return SaleAnalysisPannelResponse(list=[panel])
            else:
                # 如果没有结果，返回空列表
                return SaleAnalysisPannelResponse(list=[])
        except Exception as e:
            logger.error(f"获取销售分析面板失败: {str(e)}")
            raise CustomException(f"获取销售分析面板失败: {str(e)}")

    async def get_sale_forecast(self, db: Session) -> SaleForecastResponse:
        try:
            # 获取本年销售额
            base_query = text(f"""
                SELECT 
                    SUM(CASE WHEN Year = YEAR(GETDATE()) THEN MonthlyTarget ELSE 0 END) AS YearlyTotal,
                    SUM(CASE WHEN Year = YEAR(GETDATE()) AND Month = MONTH(GETDATE()) THEN MonthlyTarget ELSE 0 END) AS MonthlyTotal
                FROM huaxinAdmin_SaleTarget;
            """)

            # 执行查询
            result = db.exec(base_query).first()

            return SaleForecastResponse(
                YearForecast=int(result.YearlyTotal) if result.YearlyTotal else 0,
                MonthForecast=int(result.MonthlyTotal) if result.MonthlyTotal else 0
            )
        except Exception as e:
            logger.error(f"获取销售预测失败: {str(e)}")
            raise CustomException(f"获取销售预测失败: {str(e)}")

    async def get_sale_analyze_amount(self, db: Session, params: SaleAmountQuery) -> List[SaleAmountResponse]:
        try:
            # 初始化变量
            where_clause_sale_actual = ''
            where_clause_sale_return = ''
            where_clause_sale_forecast = ''

            # 清理输入参数并构建WHERE子句
            if params.year:
                year = self._clean_input(params.year)
                where_clause_sale_actual += f"AND YEAR(SI.TRANSACTION_DATE) = {year}"
                where_clause_sale_return += f"AND YEAR(SR.TRANSACTION_DATE) = {year}"
                where_clause_sale_forecast += f"AND SF.[Year] = {year}"
            if params.month:
                month = self._clean_input(params.month)
                where_clause_sale_actual += f"AND MONTH(SI.TRANSACTION_DATE) = {month}"
                where_clause_sale_return += f"AND MONTH(SR.TRANSACTION_DATE) = {month}"
                where_clause_sale_forecast += f"AND SF.[Month] = {month}"
            if params.admin_unit_name:
                admin_unit_name = self._clean_input(params.admin_unit_name)
                where_clause_sale_actual += f"AND AU.ADMIN_UNIT_NAME = '{admin_unit_name}'"
                where_clause_sale_return += f"AND AU.ADMIN_UNIT_NAME = '{admin_unit_name}'"
                where_clause_sale_forecast += f"AND SF.[AdminUnitName] = '{admin_unit_name}'"
            if params.employee_name:
                employee_name = self._clean_input(params.employee_name)
                where_clause_sale_actual += f"AND E.EMPLOYEE_NAME = '{employee_name}'"
                where_clause_sale_return += f"AND E.EMPLOYEE_NAME = '{employee_name}'"
                where_clause_sale_forecast += f"AND SF.[EmployeeName] = '{employee_name}'"
                
            # 构建GROUP BY子句
            group_by = []

            # 根据参数添加分组字段
            if params.group_by_year:
                group_by.append("[YEAR]")
            if params.group_by_month:
                group_by.append("[MONTH]")
            if params.group_by_admin_unit_name:
                group_by.append("[ADMIN_UNIT_NAME]")
            if params.group_by_employee_name:
                group_by.append("[EMPLOYEE_NAME]")
            
            if len(group_by) == 0:
                group_by.append("[YEAR]")
            
            group_by_clause = ", ".join(group_by)

            # 构建查询语句
            base_query = text(f"""
                WITH 
                FSD AS (
                    SELECT SF.[Year] AS [YEAR],SF.[Month] AS [MONTH],SF.AdminUnitName AS ADMIN_UNIT_NAME,SF.EmployeeName AS EMPLOYEE_NAME,SF.MonthlyTarget AS FORECAST_AMOUNT,
                    CB.PRICE_AMOUNT,
                    CB.PRICE_QTY
                    FROM huaxinAdmin_SaleTarget SF
                    LEFT JOIN (
                    SELECT 
                        [YEAR],
                        [MONTH],
                        ADMIN_UNIT_NAME,
                        EMPLOYEE_NAME,
                        SUM(NT.AMOUNT) AS PRICE_AMOUNT,
                        SUM(NT.PRICE_QTY) AS PRICE_QTY
                    FROM
                    (
                    ( 
                            SELECT 
                            YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                            MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                            AU.ADMIN_UNIT_NAME,
                            E.EMPLOYEE_NAME,
                            SID.PRICE_QTY,
                            SID.AMOUNT
                            FROM SALES_DELIVERY SD
                            LEFT JOIN SALES_DELIVERY_D SDD
                            ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                            LEFT JOIN SALES_ISSUE_D SID
                            ON SID.SOURCE_ID_ROid = SDD. SALES_DELIVERY_D_ID
                            LEFT JOIN SALES_ISSUE SI
                            ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                            LEFT JOIN EMPLOYEE E
                            ON SD.Owner_Emp = E.EMPLOYEE_ID
                            LEFT JOIN EMPLOYEE_D ED
                            ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                            LEFT JOIN ADMIN_UNIT AU
                            ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                            WHERE SID.PRICE_QTY > 0 AND SD.CATEGORY = '24' {where_clause_sale_actual}
                    )
                    UNION ALL
                    (
                            SELECT 
                            YEAR(SR.TRANSACTION_DATE) AS [YEAR],
                            MONTH(SR.TRANSACTION_DATE) AS [MONTH],
                            AU.ADMIN_UNIT_NAME,
                            E.EMPLOYEE_NAME,
                            SRD.PRICE_QTY * -1 AS PRICE_QTY,
                            SRD.AMOUNT * -1 AS AMOUNT
                            FROM SALES_RETURN SR
                            LEFT JOIN SALES_RETURN_D SRD
                            ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                            LEFT JOIN EMPLOYEE E
                            ON SR.Owner_Emp = E.EMPLOYEE_ID
                            LEFT JOIN EMPLOYEE_D ED
                            ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                            LEFT JOIN ADMIN_UNIT AU
                            ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                            WHERE SR.CATEGORY = '26' AND SR.RECEIPTED_STATUS = '3' {where_clause_sale_return}
                    )
                    ) AS NT
                    GROUP BY [YEAR],[MONTH],EMPLOYEE_NAME,ADMIN_UNIT_NAME
                    ) AS CB
                    ON SF.[Year] = CB.[YEAR] AND SF.[Month] = CB.[MONTH] AND SF.AdminUnitName = CB.ADMIN_UNIT_NAME AND SF.EmployeeName = CB.EMPLOYEE_NAME
                    WHERE 1 = 1 {where_clause_sale_forecast}
                )
                
                SELECT
                {group_by_clause},
                SUM(PRICE_QTY) AS PRICE_QTY,
                SUM(FORECAST_AMOUNT) AS FORECAST_AMOUNT,
                SUM(PRICE_AMOUNT) AS PRICE_AMOUNT,
                SUM(PRICE_AMOUNT)/SUM(FORECAST_AMOUNT)*100 AS PERCENTAGE
                FROM FSD
                GROUP BY {group_by_clause}
            """)

            # 执行查询
            result = db.exec(base_query).all()

            # 将查询结果转换为响应格式
            result_list = []
            for amount in result:
                analysis_amount = SaleAmount(
                    FORECAST_AMOUNT=amount.FORECAST_AMOUNT if amount.FORECAST_AMOUNT else 0,
                    PRICE_AMOUNT=amount.PRICE_AMOUNT if amount.PRICE_AMOUNT else 0,
                    PERCENTAGE=amount.PERCENTAGE if amount.PERCENTAGE else 0,
                    PRICE_QTY=amount.PRICE_QTY if amount.PRICE_QTY else 0
                )
                # 根据分组字段设置相应的字段值
                if params.group_by_year and hasattr(amount, 'YEAR'):
                    analysis_amount.YEAR = amount.YEAR
                if params.group_by_month and hasattr(amount, 'MONTH'):
                    analysis_amount.MONTH = amount.MONTH
                if params.group_by_admin_unit_name and hasattr(amount, 'ADMIN_UNIT_NAME'):
                    analysis_amount.ADMIN_UNIT_NAME = amount.ADMIN_UNIT_NAME
                if params.group_by_employee_name and hasattr(amount, 'EMPLOYEE_NAME'):
                    analysis_amount.EMPLOYEE_NAME = amount.EMPLOYEE_NAME
                
                result_list.append(analysis_amount)
                
            return SaleAmountResponse(list=result_list)
        except Exception as e:
            logger.error(f"获取销售金额详情失败: {str(e)}")
            raise CustomException(f"获取销售金额详情失败: {str(e)}")

    async def get_sale_amount_bar_chart(self,db: Session,params: SaleAmountBarChartQuery) -> SaleAmountBarChartEChartsResponse:
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)

            # 构建查询条件
            where_clause = ""
            if year:
                where_clause += f"AND YEAR(SI.TRANSACTION_DATE) = {year}"
            if month:
                where_clause += f"AND MONTH(SI.TRANSACTION_DATE) = {month}"
            
            # 构建查询语句
            base_query = text(f"""
                WITH SD AS (
                SELECT 
                    YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                    MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                    ISNULL(AU.ADMIN_UNIT_NAME, '') AS ADMIN_UNIT_NAME,
                    ISNULL(E.EMPLOYEE_NAME, '') AS EMPLOYEE_NAME,
                    ISNULL(ITEM.ITEM_CODE, '') AS ITEM_CODE,
                    ISNULL(dbo.RemoveSpecificStrings(ITEM.ITEM_CODE, 'CP-,DG-,-Blank+TR,- Blank+TR,-Blank+TS,-FT+TR,-PGM+TR,-WG+TR,-PGM+TS,-PGM,- PGM,-FT,-TR,+TS+TR,-Blank,-WG,-AB,+TR'), '') AS ITEM_NAME,
                    ISNULL(ITEM.SHORTCUT, '') AS SHORTCUT,
                    ISNULL(SDD.PRICE, 0) AS PRICE,
                    ISNULL(SID.PRICE_QTY, 0) AS PRICE_QTY,
                    ISNULL(SID.AMOUNT, 0) AS AMOUNT
                FROM SALES_DELIVERY SD
                LEFT JOIN SALES_DELIVERY_D SDD ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                LEFT JOIN SALES_ISSUE_D SID ON SID.SOURCE_ID_ROid = SDD.SALES_DELIVERY_D_ID
                LEFT JOIN SALES_ISSUE SI ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                LEFT JOIN EMPLOYEE E ON SD.Owner_Emp = E.EMPLOYEE_ID
                LEFT JOIN EMPLOYEE_D ED ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                LEFT JOIN ADMIN_UNIT AU ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                LEFT JOIN ITEM ON SDD.ITEM_ID = ITEM.ITEM_BUSINESS_ID
                WHERE SD.CATEGORY = '24' { where_clause }
                )

                SELECT 
                    ADMIN_UNIT_NAME,
                    EMPLOYEE_NAME,
                    SHORTCUT,
                    ITEM_NAME,
                    SUM(PRICE_QTY) AS PRICE_QTY,
                    SUM(AMOUNT) AS AMOUNT
                FROM SD
                GROUP BY 
                    ADMIN_UNIT_NAME,
                    EMPLOYEE_NAME,
                    SHORTCUT,
                    ITEM_NAME
            """)

            # 执行查询
            result = db.exec(base_query).all()

            # 确保结果中没有 None 值
            for row in result:
                if row.ADMIN_UNIT_NAME is None:
                    row.ADMIN_UNIT_NAME = ""
                if row.EMPLOYEE_NAME is None:
                    row.EMPLOYEE_NAME = ""
                if row.SHORTCUT is None:
                    row.SHORTCUT = ""
                if row.ITEM_NAME is None:
                    row.ITEM_NAME = ""
                if row.PRICE_QTY is None:
                    row.PRICE_QTY = 0
                if row.AMOUNT is None:
                    row.AMOUNT = 0

            functions = Functions()
            response_dict = functions.process_amount_data_for_echarts(result)
            
            # 将字典中的项转换为 SaleAmountBarChartEChartsDataItem 对象
            processed_dict = {}
            for level_id, items in response_dict.items():
                data_items = []
                for item in items:
                    data_items.append(SaleAmountBarChartEChartsDataItem(
                        name=item['name'],
                        value=float(item['value']),
                        quantity=item.get('quantity'),
                        group_id=item['group_id'],
                        child_group_id=item.get('child_group_id')
                    ))
                
                processed_dict[level_id] = data_items
            
            # 将字典转换为列表
            response_list = [
                SaleAmountBarChartEChartsLevelData(
                    level_id=level_id,
                    items=data_items
                )
                for level_id, data_items in processed_dict.items()
            ]
            

            return SaleAmountBarChartEChartsResponse(list=response_list)
            
            
            
        except Exception as e:
            logger.error(f"获取销售金额柱状图失败: {str(e)}")
            raise CustomException(f"获取销售金额柱状图失败: {str(e)}")

    async def get_sale_percentage_bar_chart(self,db: Session,params: SaleAmountBarChartQuery) -> SaleAmountBarChartEChartsResponse:
        try:
            # 清理输入参数
            year = self._clean_input(params.year)
            month = self._clean_input(params.month)

            # 构建查询条件
            where_clause = ""
            if year:
                where_clause += f"AND SF.[Year] = {year} "
            if month:
                where_clause += f"AND SF.[Month] = {month} "
            
            # 构建查询语句
            base_query = text(f"""
                WITH P AS (
                    SELECT 
                        SF.[Year] AS [YEAR],
                        SF.[Month] AS [MONTH],
                        SF.AdminUnitName AS ADMIN_UNIT_NAME,
                        SF.EmployeeName AS EMPLOYEE_NAME,
                        SF.MonthlyTarget AS FORECAST_AMOUNT,
                        ISNULL(CB.PRICE_AMOUNT,0) AS PRICE_AMOUNT,
                        ISNULL(CB.PRICE_QTY,0) AS PRICE_QTY
                    FROM huaxinAdmin_SaleTarget SF
                    LEFT JOIN (
                        SELECT 
                            [YEAR],
                            [MONTH],
                            ADMIN_UNIT_NAME,
                            EMPLOYEE_NAME,
                            SUM(NT.AMOUNT) AS PRICE_AMOUNT,
                            SUM(NT.PRICE_QTY) AS PRICE_QTY
                        FROM (
                            SELECT 
                                YEAR(SI.TRANSACTION_DATE) AS [YEAR],
                                MONTH(SI.TRANSACTION_DATE) AS [MONTH],
                                AU.ADMIN_UNIT_NAME,
                                E.EMPLOYEE_NAME,
                                SID.PRICE_QTY,
                                SID.AMOUNT
                            FROM SALES_DELIVERY SD
                            LEFT JOIN SALES_DELIVERY_D SDD 
                                ON SD.SALES_DELIVERY_ID = SDD.SALES_DELIVERY_ID
                            LEFT JOIN SALES_ISSUE_D SID 
                                ON SID.SOURCE_ID_ROid = SDD.SALES_DELIVERY_D_ID -- 修正了表别名间距
                            LEFT JOIN SALES_ISSUE SI 
                                ON SI.SALES_ISSUE_ID = SID.SALES_ISSUE_ID
                            LEFT JOIN EMPLOYEE E 
                                ON SD.Owner_Emp = E.EMPLOYEE_ID
                            LEFT JOIN EMPLOYEE_D ED 
                                ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                            LEFT JOIN ADMIN_UNIT AU 
                                ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                            WHERE SID.PRICE_QTY > 0 AND SD.CATEGORY = '24'
                            
                            UNION ALL
                            
                            SELECT 
                                YEAR(SR.TRANSACTION_DATE) AS [YEAR],
                                MONTH(SR.TRANSACTION_DATE) AS [MONTH],
                                AU.ADMIN_UNIT_NAME,
                                E.EMPLOYEE_NAME,
                                SRD.PRICE_QTY * -1 AS PRICE_QTY,
                                SRD.AMOUNT * -1 AS AMOUNT
                            FROM SALES_RETURN SR
                            LEFT JOIN SALES_RETURN_D SRD 
                                ON SR.SALES_RETURN_ID = SRD.SALES_RETURN_ID
                            LEFT JOIN EMPLOYEE E 
                                ON SR.Owner_Emp = E.EMPLOYEE_ID
                            LEFT JOIN EMPLOYEE_D ED 
                                ON ED.EMPLOYEE_ID = E.EMPLOYEE_ID
                            LEFT JOIN ADMIN_UNIT AU 
                                ON AU.ADMIN_UNIT_ID = ED.ADMIN_UNIT_ID
                            WHERE SR.CATEGORY = '26' AND SR.RECEIPTED_STATUS = '3'
                        ) AS NT
                        GROUP BY [YEAR], [MONTH], EMPLOYEE_NAME, ADMIN_UNIT_NAME
                    ) AS CB
                        ON SF.[Year] = CB.[YEAR] 
                        AND SF.[Month] = CB.[MONTH] 
                        AND SF.AdminUnitName = CB.ADMIN_UNIT_NAME 
                        AND SF.EmployeeName = CB.EMPLOYEE_NAME
                    WHERE ((SF.[Year] < YEAR(GETDATE())) OR (SF.[Year] = YEAR(GETDATE()) AND SF.[Month] <= MONTH(GETDATE()))) { where_clause } )

                    SELECT
                    ADMIN_UNIT_NAME,
                    EMPLOYEE_NAME,
                    PRICE_AMOUNT,
                    FORECAST_AMOUNT
                    FROM P
            """)

            # 执行查询
            result = db.exec(base_query).all()

            # 确保结果中没有 None 值
            for row in result:
                if row.ADMIN_UNIT_NAME is None:
                    row.ADMIN_UNIT_NAME = ""
                if row.EMPLOYEE_NAME is None:
                    row.EMPLOYEE_NAME = ""
                if row.FORECAST_AMOUNT is None:
                    row.FORECAST_AMOUNT = 0
                if row.PRICE_AMOUNT is None:
                    row.PRICE_AMOUNT = 0

            functions = Functions()
            response_dict = functions.process_percentage_data_for_echarts(result)
            
            # 将字典中的项转换为 SaleAmountBarChartEChartsDataItem 对象
            processed_dict = {}
            for level_id, items in response_dict.items():
                data_items = []
                for item in items:
                    data_items.append(SaleAmountBarChartEChartsDataItem(
                        name=item['name'],
                        value=float(item['value']),
                        group_id=item['group_id'],
                        child_group_id=item.get('child_group_id')
                    ))
                
                processed_dict[level_id] = data_items
            
            # 将字典转换为列表
            response_list = [
                SaleAmountBarChartEChartsLevelData(
                    level_id=level_id,
                    items=data_items
                )
                for level_id, data_items in processed_dict.items()
            ]
            

            return SaleAmountBarChartEChartsResponse(list=response_list)
        except Exception as e:
            logger.error(f"获取销售金额完成率柱状图失败: {str(e)}")
            raise CustomException(f"获取销售金额完成率柱状图失败: {str(e)}")

